#!/usr/bin/env python3
import os as _os
_HERE   = _os.path.dirname(_os.path.abspath(__file__))
_QC     = _os.path.dirname(_HERE)
_CACHE  = _os.path.join(_QC, 'cache')
_OUTPUT = _os.path.join(_QC, 'output')

"""Build L2PHL Cross-Round Questionnaire Change Report"""
import json
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

ROUNDS = ['R1','R2','R3','R4','R5']
ROUND_LABELS = {'R1':'R1 (Nov 2025)','R2':'R2 (Dec 2025)','R3':'R3 (Jan 2026)','R4':'R4 (Feb 2026)','R5':'R5 (Mar 2026)'}
MODULES = ['M00','M01','M02','M03','M04','M05','M06','M07','M08','M09']
MOD_NAMES = {
    'M00':'Introduction / Passport','M01':'Demographics / Roster',
    'M02':'Education','M03':'Shocks','M04':'Employment',
    'M05':'Income','M06':'Finance','M07':'Health',
    'M08':'Food & Non-Food','M09':'Opinions & Views'
}

# Load data
with open(_os.path.join(_CACHE, 'module_tables.json')) as f:
    module_tables = json.load(f)
with open(_os.path.join(_CACHE, 'all_questions.json')) as f:
    all_qs = json.load(f)
with open(_os.path.join(_CACHE, 'do_modules.json')) as f:
    do_mods = json.load(f)

# ── Styles ──────────────────────────────────────────────────────────────────
FONT_NAME = 'Arial'

def font(bold=False, size=10, color='000000', italic=False):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style='medium', color='999999')
    return Border(left=s, right=s, top=s, bottom=s)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Color palette
CLR_HEADER_DARK   = '1F3864'  # Dark navy
CLR_HEADER_MID    = '2E5090'  # Medium blue
CLR_MOD_HEADER    = '2E5090'
CLR_SUBHEADER     = 'B8CCE4'  # Light blue
CLR_NEW           = 'E2EFDA'  # Light green - new question
CLR_DROPPED       = 'FCE4D6'  # Light orange - dropped
CLR_CHANGED       = 'FFF2CC'  # Light yellow - changed
CLR_CONSISTENT    = 'F2F2F2'  # Light grey - consistent
CLR_PRESENT       = 'C6EFCE'  # Green tick cell
CLR_ABSENT        = 'FFCCCC'  # Red absent cell
CLR_ROW_ALT       = 'F7FAFF'  # Very light blue alt row
CLR_WHITE         = 'FFFFFF'
CLR_SECTION_HDR   = 'D9E1F2'  # Section divider


def style_cell(cell, bold=False, size=10, color='000000', bg=None,
               h_align='left', v_align='center', wrap=False, italic=False):
    cell.font = font(bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align(h_align, v_align, wrap)
    cell.border = border_thin()


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def write_header_row(ws, row, cols, bg=CLR_HEADER_DARK, txt_color='FFFFFF', size=10):
    for c_idx, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c_idx, value=text)
        style_cell(cell, bold=True, size=size, color=txt_color, bg=bg, h_align='center', v_align='center', wrap=True)


# ── COVER SHEET ──────────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.active
    ws.title = 'Cover'
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20

    # Title block
    ws.merge_cells('B2:C2')
    c = ws['B2']
    c.value = 'L2PHL CATI Survey\nCross-Round Questionnaire Change Report'
    c.font = Font(name=FONT_NAME, bold=True, size=20, color='FFFFFF')
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.merge_cells('B3:C3')
    c = ws['B3']
    c.value = 'Rounds R1 (Nov 2025) – R5 (Mar 2026)  |  Modules M00–M09'
    c.font = Font(name=FONT_NAME, size=12, color='FFFFFF', italic=True)
    c.fill = fill(CLR_HEADER_MID)
    c.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('B4:C4')
    c = ws['B4']
    c.value = 'This report tracks all questionnaire variables across 5 CATI rounds: new questions, dropped questions, title/type changes, skip-logic changes, and data-check notes.'
    c.font = Font(name=FONT_NAME, size=10, color='333333', italic=True)
    c.fill = fill('EEF3FB')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Module index table
    r = 7
    ws.merge_cells(f'B{r}:C{r}')
    c = ws[f'B{r}']
    c.value = 'MODULE INDEX'
    style_cell(c, bold=True, size=11, bg=CLR_HEADER_DARK, color='FFFFFF', h_align='left')

    r += 1
    for col_hdr, col_idx in [('Module', 'B'), ('Sheet Name', 'C')]:
        c = ws[f'{col_idx}{r}']
        c.value = col_hdr
        style_cell(c, bold=True, size=10, bg=CLR_SUBHEADER, h_align='left')

    r += 1
    for i, mod in enumerate(MODULES):
        bg = CLR_ROW_ALT if i % 2 == 0 else CLR_WHITE
        c = ws[f'B{r}']
        c.value = f'{mod} – {MOD_NAMES[mod]}'
        style_cell(c, size=10, bg=bg)
        c2 = ws[f'C{r}']
        c2.value = f'{mod} {MOD_NAMES[mod][:20]}'
        style_cell(c2, size=10, bg=bg)
        r += 1

    # Legend
    r += 2
    ws.merge_cells(f'B{r}:C{r}')
    c = ws[f'B{r}']
    c.value = 'ROW COLOUR LEGEND'
    style_cell(c, bold=True, size=10, bg=CLR_HEADER_DARK, color='FFFFFF')
    r += 1

    legend = [
        (CLR_NEW,        'New question (not in R1)'),
        (CLR_DROPPED,    'Question dropped before R5'),
        (CLR_CHANGED,    'Variable changed across rounds'),
        (CLR_CONSISTENT, 'Consistent across all rounds'),
        (CLR_PRESENT,    '✓  Present in this round'),
        (CLR_ABSENT,     '—  Absent in this round'),
    ]
    for bg_clr, desc in legend:
        c = ws[f'B{r}']
        c.value = '    ' + desc
        style_cell(c, size=10, bg=bg_clr)
        c2 = ws[f'C{r}']
        c2.value = ''
        style_cell(c2, size=10, bg=bg_clr)
        r += 1


# ── MODULE SUMMARY SHEET ─────────────────────────────────────────────────────
def build_summary(wb):
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'

    # Title
    ws.merge_cells('A1:P1')
    c = ws['A1']
    c.value = 'L2PHL CATI – Cross-Round Variable Count Summary (All Modules)'
    c.font = Font(name=FONT_NAME, bold=True, size=13, color='FFFFFF')
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Sub-header explanation
    ws.merge_cells('A2:P2')
    c = ws['A2']
    c.value = 'Counts show number of unique question items per module per round (from questionnaire). Δ = change vs previous round.'
    c.font = Font(name=FONT_NAME, size=9, italic=True, color='555555')
    c.fill = fill('EEF3FB')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16

    # Headers
    ws.row_dimensions[3].height = 32
    headers = ['Module', 'Module Name'] + [ROUND_LABELS[r] for r in ROUNDS] + ['New in R2','New in R3','New in R4','New in R5','Dropped','Changed','Notes']
    write_header_row(ws, 3, headers, bg=CLR_HEADER_DARK)

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    for ci in range(3, 3+5):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    for ci in range(8, 8+7):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.column_dimensions[get_column_letter(15)].width = 40

    r = 4
    for i, mod in enumerate(MODULES):
        rows = module_tables.get(mod, [])
        counts_by_round = {}
        for rnd in ROUNDS:
            counts_by_round[rnd] = sum(1 for row in rows if row.get(f'in_{rnd}','') == '✓')

        new_counts = {}
        for rnd in ROUNDS[1:]:
            new_counts[rnd] = sum(1 for row in rows
                                   if row.get(f'in_{rnd}','') == '✓'
                                   and row.get(f'in_{ROUNDS[ROUNDS.index(rnd)-1]}','') != '✓'
                                   and all(row.get(f'in_{ROUNDS[j]}','') != '✓' for j in range(ROUNDS.index(rnd))))

        dropped = sum(1 for row in rows if row.get('status','').startswith('Dropped'))
        changed = sum(1 for row in rows if row.get('title_changes','') or row.get('skip_changes',''))

        bg = CLR_ROW_ALT if i % 2 == 0 else CLR_WHITE

        row_data = [
            mod, MOD_NAMES[mod],
            counts_by_round['R1'], counts_by_round['R2'],
            counts_by_round['R3'], counts_by_round['R4'], counts_by_round['R5'],
            new_counts.get('R2', 0), new_counts.get('R3', 0),
            new_counts.get('R4', 0), new_counts.get('R5', 0),
            dropped, changed, ''
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            h = 'center' if ci > 2 else 'left'
            style_cell(cell, size=10, bg=bg, h_align=h)

        # Highlight modules with changes
        if changed > 0 or dropped > 0:
            ws.cell(row=r, column=1).fill = fill(CLR_CHANGED)
            ws.cell(row=r, column=1).font = Font(name=FONT_NAME, bold=True, size=10)

        r += 1


# ── PER-MODULE SHEET ─────────────────────────────────────────────────────────
def build_module_sheet(wb, mod):
    mod_name = MOD_NAMES[mod]
    sheet_title = f'{mod} {mod_name[:20]}'.replace('/', '-')
    ws = wb.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'D5'

    # Title row
    ws.merge_cells('A1:R1')
    c = ws['A1']
    c.value = f'{mod} – {mod_name}  |  Cross-Round Variable Tracker  (R1–R5)'
    c.font = Font(name=FONT_NAME, bold=True, size=13, color='FFFFFF')
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    # Sub-header row: round dates
    dates_row = ['','','','R1\nNov 2025','R2\nDec 2025','R3\nJan 2026','R4\nFeb 2026','R5\nMar 2026','','','','','','','','','','']
    ws.row_dimensions[2].height = 32
    for ci, val in enumerate(dates_row[:18], 1):
        cell = ws.cell(row=2, column=ci, value=val)
        if val.startswith('R'):
            style_cell(cell, bold=True, size=10, color='FFFFFF', bg=CLR_HEADER_MID, h_align='center', v_align='center', wrap=True)
        else:
            cell.fill = fill('EEF3FB')

    # Merge presence header over R1-R5
    ws.merge_cells('D2:H2')

    # Column headers
    ws.row_dimensions[3].height = 36
    headers = [
        'Variable', 'Question Title', 'Status',
        'R1', 'R2', 'R3', 'R4', 'R5',
        'Question Text (English)', 'Question Type',
        '# Codes R1', '# Codes R5',
        'Title / Wording Change', 'Skip Logic Change',
        'Skip Logic (R1)', 'Skip Logic (R5)',
        'Data Check Notes (R5)', 'Remarks'
    ]
    write_header_row(ws, 3, headers, bg=CLR_HEADER_MID, size=9)
    ws.row_dimensions[3].height = 40

    # Column widths
    col_widths = [12, 35, 18, 5, 5, 5, 5, 5, 50, 16, 9, 9, 42, 30, 35, 35, 40, 35]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze
    ws.freeze_panes = 'D4'

    rows = module_tables.get(mod, [])
    if not rows:
        cell = ws.cell(row=4, column=1, value='No data found for this module.')
        style_cell(cell, italic=True, color='888888')
        return

    data_row = 4
    for i, row in enumerate(rows):
        status = row.get('status', 'All rounds')

        # Row background based on status
        if 'New in' in status:
            row_bg = CLR_NEW
        elif 'Dropped' in status:
            row_bg = CLR_DROPPED
        elif row.get('title_changes','') or row.get('skip_changes',''):
            row_bg = CLR_CHANGED
        else:
            row_bg = CLR_ROW_ALT if i % 2 == 0 else CLR_WHITE

        ws.row_dimensions[data_row].height = 40

        # Get skip logic for R1 and R5
        skip_r1 = row.get('skip_r1','')
        skip_r5 = row.get('skip_r5','') or row.get('skip_r4','') or row.get('skip_r3','')

        # Get data check from R5 (most complete)
        data_check = row.get('data_check_r5','') or row.get('data_check_r4','') or row.get('data_check_r3','')

        row_vals = [
            row.get('variable',''),
            row.get('question_title',''),
            status,
            row.get('in_R1',''),
            row.get('in_R2',''),
            row.get('in_R3',''),
            row.get('in_R4',''),
            row.get('in_R5',''),
            row.get('english_text',''),
            row.get('question_type',''),
            row.get('codes_r1','') if row.get('codes_r1','') else '',
            row.get('codes_r5','') if row.get('codes_r5','') else '',
            row.get('title_changes',''),
            row.get('skip_changes',''),
            skip_r1,
            skip_r5,
            data_check,
            row.get('remarks',''),
        ]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=data_row, column=ci, value=val)
            h = 'center' if ci in (1, 3, 4, 5, 6, 7, 8, 11, 12) else 'left'
            wrap = ci not in (1, 2, 3, 4, 5, 6, 7, 8, 11, 12)
            style_cell(cell, size=9, bg=row_bg, h_align=h, v_align='top', wrap=wrap)

            # Presence cells: color individually
            if ci in (4, 5, 6, 7, 8):
                if val == '✓':
                    cell.fill = fill(CLR_PRESENT)
                    cell.font = Font(name=FONT_NAME, bold=True, size=11, color='006100')
                else:
                    cell.fill = fill(CLR_ABSENT)
                    cell.font = Font(name=FONT_NAME, size=10, color='9C0006')
                    cell.value = '—'

            # Variable name: bold
            if ci == 1:
                cell.font = Font(name=FONT_NAME, bold=True, size=10, color='000000')

            # Status coloring
            if ci == 3:
                if 'New' in str(val):
                    cell.fill = fill(CLR_NEW)
                    cell.font = Font(name=FONT_NAME, bold=True, size=9, color='375623')
                elif 'Dropped' in str(val):
                    cell.fill = fill(CLR_DROPPED)
                    cell.font = Font(name=FONT_NAME, bold=True, size=9, color='843C0C')
                elif 'All rounds' in str(val):
                    cell.fill = fill(CLR_CONSISTENT)
                    cell.font = Font(name=FONT_NAME, size=9, color='404040')

            # Highlight changes
            if ci in (13, 14) and val:
                cell.fill = fill('FFF2CC')
                cell.font = Font(name=FONT_NAME, size=9, color='7F4F00', italic=True)

            # Data check notes: highlight
            if ci == 17 and val:
                cell.fill = fill('EAF2FF')
                cell.font = Font(name=FONT_NAME, size=9, color='003366', italic=True)

        data_row += 1

    # Add a "DO-file variables" section at the bottom
    data_row += 1
    ws.merge_cells(f'A{data_row}:R{data_row}')
    c = ws.cell(row=data_row, column=1, value=f'DO-FILE VARIABLE DETAILS (destring/generate operations per round)')
    c.font = Font(name=FONT_NAME, bold=True, size=10, color='FFFFFF')
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[data_row].height = 20
    data_row += 1

    # Headers for do-file section
    do_headers = ['Round', 'Variables processed (destring/tostring)', 'Generated variables', 'Notes']
    for ci, h in enumerate(do_headers, 1):
        cell = ws.cell(row=data_row, column=ci, value=h)
        style_cell(cell, bold=True, size=9, bg=CLR_SUBHEADER, h_align='left')
    ws.merge_cells(f'B{data_row}:J{data_row}')
    ws.merge_cells(f'K{data_row}:N{data_row}')
    ws.merge_cells(f'O{data_row}:R{data_row}')
    ws.row_dimensions[data_row].height = 18
    data_row += 1

    for ri, rnd in enumerate(ROUNDS):
        mod_data = do_mods.get(rnd, {}).get(mod, {})
        destr_vars = mod_data.get('destring', [])
        gen_vars = mod_data.get('generate', [])

        bg = CLR_ROW_ALT if ri % 2 == 0 else CLR_WHITE
        ws.row_dimensions[data_row].height = 25

        cell = ws.cell(row=data_row, column=1, value=ROUND_LABELS[rnd])
        style_cell(cell, bold=True, size=9, bg=bg)

        cell2 = ws.cell(row=data_row, column=2, value=', '.join(destr_vars[:30]) + ('...' if len(destr_vars)>30 else ''))
        style_cell(cell2, size=9, bg=bg, wrap=True)
        ws.merge_cells(f'B{data_row}:J{data_row}')

        gen_str = ', '.join([g['var'] for g in gen_vars[:20]]) if gen_vars else ''
        cell3 = ws.cell(row=data_row, column=11, value=gen_str)
        style_cell(cell3, size=9, bg=bg, wrap=True)
        ws.merge_cells(f'K{data_row}:N{data_row}')

        # Notes about differences
        q_vars = set(q['qnum'].lower() for q in all_qs.get(rnd, {}).get(mod, []))
        do_vars_set = set(destr_vars)
        in_do_not_q = sorted(do_vars_set - q_vars)[:15]
        note = ('Raw/derived only: ' + ', '.join(in_do_not_q)) if in_do_not_q else 'Matches questionnaire'
        cell4 = ws.cell(row=data_row, column=15, value=note)
        style_cell(cell4, size=9, bg=bg, wrap=True, italic=bool(in_do_not_q))
        ws.merge_cells(f'O{data_row}:R{data_row}')

        data_row += 1


# ── CHANGES SUMMARY SHEET ────────────────────────────────────────────────────
def build_changes_sheet(wb):
    ws = wb.create_sheet('All Changes by Round')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'L2PHL CATI – All Questionnaire Changes by Round (R2–R5 vs Previous)'
    c.font = Font(name=FONT_NAME, bold=True, size=13, color='FFFFFF')
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = 'Each row is a change event: new question, dropped question, wording change, skip-logic change, or new data-check note.'
    c.font = Font(name=FONT_NAME, size=9, italic=True, color='555555')
    c.fill = fill('EEF3FB')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16

    headers = ['Round Introduced', 'Module', 'Variable', 'Change Type', 'Detail', 'Question Title', 'English Text (truncated)', 'Skip Logic']
    write_header_row(ws, 3, headers, bg=CLR_HEADER_DARK, size=9)
    ws.row_dimensions[3].height = 36

    col_widths = [16, 22, 12, 22, 55, 35, 55, 40]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    CHANGE_COLORS = {
        'New Question':         CLR_NEW,
        'Question Dropped':     CLR_DROPPED,
        'Title/Wording Change': CLR_CHANGED,
        'Skip Logic Change':    'FFF2CC',
        'Code List Change':     'DDEBF7',
        'Data Check Added':     'EAF2FF',
        'Type Change':          'FCE4D6',
    }

    data_row = 4
    for rnd in ROUNDS[1:]:
        prev_rnd = ROUNDS[ROUNDS.index(rnd)-1]
        section_written = False

        for mod in MODULES:
            rows = module_tables.get(mod, [])

            for row in rows:
                var = row.get('variable','')
                in_curr = row.get(f'in_{rnd}','') == '✓'
                in_prev = row.get(f'in_{prev_rnd}','') == '✓'

                changes_to_log = []

                # New question
                if in_curr and not in_prev and all(row.get(f'in_{ROUNDS[j]}','') != '✓' for j in range(ROUNDS.index(rnd))):
                    changes_to_log.append({
                        'type': 'New Question',
                        'detail': f'First appears in {rnd}',
                        'title': row.get('question_title',''),
                        'eng': row.get('english_text','')[:120],
                        'skip': '',
                    })

                # Question dropped
                if not in_curr and in_prev:
                    changes_to_log.append({
                        'type': 'Question Dropped',
                        'detail': f'Present in {prev_rnd}, absent from {rnd}+',
                        'title': row.get('question_title',''),
                        'eng': row.get('english_text','')[:120],
                        'skip': '',
                    })

                if in_curr and in_prev:
                    # Title change
                    tc = row.get('title_changes','')
                    if tc and f'{prev_rnd}→{rnd}' in tc:
                        changes_to_log.append({
                            'type': 'Title/Wording Change',
                            'detail': [x for x in tc.split(' | ') if f'{prev_rnd}→{rnd}' in x][0] if tc else '',
                            'title': row.get('question_title',''),
                            'eng': row.get('english_text','')[:120],
                            'skip': '',
                        })

                    # Skip logic change
                    sc = row.get('skip_changes','')
                    if sc and rnd in sc:
                        skip_detail = [x for x in sc.split(' | ') if rnd in x]
                        if skip_detail:
                            changes_to_log.append({
                                'type': 'Skip Logic Change',
                                'detail': skip_detail[0],
                                'title': row.get('question_title',''),
                                'eng': '',
                                'skip': row.get(f'skip_{rnd.lower()}','')[:120],
                            })

                    # Code count change
                    c_prev = row.get(f'codes_{prev_rnd.lower()}', '')
                    c_curr = row.get(f'codes_{rnd.lower()}', '')
                    if c_prev and c_curr and str(c_prev) != str(c_curr):
                        changes_to_log.append({
                            'type': 'Code List Change',
                            'detail': f'# answer codes: {c_prev} → {c_curr}',
                            'title': row.get('question_title',''),
                            'eng': '',
                            'skip': '',
                        })

                    # Data check added
                    dc_curr = row.get(f'data_check_{rnd.lower()}','')
                    dc_prev = row.get(f'data_check_{prev_rnd.lower()}','')
                    if dc_curr and not dc_prev:
                        changes_to_log.append({
                            'type': 'Data Check Added',
                            'detail': dc_curr[:80],
                            'title': row.get('question_title',''),
                            'eng': '',
                            'skip': '',
                        })

                for chg in changes_to_log:
                    bg = CHANGE_COLORS.get(chg['type'], CLR_WHITE)
                    row_data = [
                        rnd, f'{mod} – {MOD_NAMES[mod]}', var,
                        chg['type'], chg['detail'],
                        chg['title'], chg['eng'], chg['skip']
                    ]
                    ws.row_dimensions[data_row].height = 35
                    for ci, val in enumerate(row_data, 1):
                        cell = ws.cell(row=data_row, column=ci, value=val)
                        h = 'center' if ci in (1, 3) else 'left'
                        style_cell(cell, size=9, bg=bg, h_align=h, v_align='top', wrap=(ci > 3))
                        if ci == 3:
                            cell.font = Font(name=FONT_NAME, bold=True, size=9)
                        if ci == 4:
                            cell.font = Font(name=FONT_NAME, bold=True, size=9, color='333333')
                    data_row += 1


# ── DQ ISSUES SHEET ──────────────────────────────────────────────────────────
def build_dq_issues_sheet(wb):
    """Build a sheet summarising data quality / logic issues found."""
    ws = wb.create_sheet('DQ & Logic Issues')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'

    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = 'L2PHL CATI – Data Quality & Logic Issues by Module and Round'
    c.font = Font(name=FONT_NAME, bold=True, size=13, color='FFFFFF')
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    c = ws['A2']
    c.value = 'Issues identified from questionnaire structure: skip-logic inconsistencies, variable naming changes, module additions/removals, and data-check requirements.'
    c.font = Font(name=FONT_NAME, size=9, italic=True, color='555555')
    c.fill = fill('EEF3FB')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 20

    headers = ['Module', 'Variable(s)', 'Issue Type', 'Rounds Affected', 'Description', 'Questionnaire Evidence', 'DO-file Evidence', 'Severity', 'Recommended Action']
    write_header_row(ws, 3, headers, bg=CLR_HEADER_DARK, size=9)
    ws.row_dimensions[3].height = 36

    col_widths = [22, 18, 20, 16, 55, 45, 40, 12, 45]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    SEV_COLORS = {'High': 'FF0000', 'Medium': 'FF6600', 'Low': '336600'}
    SEV_BGS = {'High': 'FCE4D6', 'Medium': 'FFF2CC', 'Low': 'E2EFDA'}

    issues = [
        # M00
        ('M00 – Introduction','FMID','New variable in R2','R1 missing, R2–R5',
         'FMID (follow-up member ID) was not included in R1 questionnaire but appears from R2 onward.',
         'Not present in R1 Introduction sheet; added in R2.',
         'destring fmid in R2 M01 section; R1 do-file lacks fmid destring.',
         'Medium','Verify R1 FMID is backfilled from baseline/trailer data before merging.'),

        # M01
        ('M01 – Demographics','D5a','Title/scope change','R1 vs R2+',
         "R1: 'NAME & AGE OF HH MEMBERS'. R2+: 'NAME, AGE & SEX OF HH MEMBERS'. Gender/sex added to the roster confirmation question.",
         'D5a title differs between R1 and R2 questionnaires.',
         'R1 do-file: no gender variable in M01 destring block; R2+: gender added.',
         'Medium','Ensure gender (D31/D28) is consistently coded for R1 members. Check if sex was captured in R1 Trailer.'),

        ('M01 – Demographics','D31','Label change: Gender → Sex','R1 vs R2+',
         "Variable D31 title changed from 'GENDER - NEW HH MEMBER' (R1) to 'SEX - NEW HH MEMBER' (R2+). Conceptual/labelling shift.",
         'D31 title differs across questionnaire rounds.',
         'Variable name unchanged but label semantics shifted.',
         'Low','Update variable label in do-file from gender to sex for R2+ consistency.'),

        ('M01 – Demographics','D33','New variable R2+','R1 missing, R2–R5',
         'D33 added in R2 (likely a new member characteristic). Absent in R1.',
         'Not in R1 Demographics sheet; present from R2.',
         'destring d33 appears in R2 M01 block.',
         'Medium','Check if D33 is needed for R1. If so, backfill with missing (.); if not applicable flag clearly.'),

        ('M01 – Demographics','M13','Code list expanded in R3','R2→R3',
         'Number of answer codes for M13 increased from 18 (R2) to 19 (R3). A new reason/category added.',
         'Code count differs: R1/R2=18, R3+=19.',
         'R1: m13_21 appears as generated var; R2: not present.',
         'Low','Check what new code was added in R3 and ensure value label is consistent across rounds in final merged dataset.'),

        # M02
        ('M02 – Education','ED16','Code list reduced R2→R3','R2→R3',
         'ED16 (dropout reasons, select_multiple) had 16 codes in R2 but 15 in R3. One response option removed or merged.',
         'Code count: R2=16, R3+=15.',
         'No corresponding code variable change found in do-files.',
         'Medium','Identify which code was dropped. Recode merged R1–R5 dataset to handle the removed category consistently.'),

        # M03
        ('M03 – Shocks','NH2,NH3,NH7,NH10,NH14–NH17','New natural hazard sub-module in R3','R3 only (R1,R2,R4,R5 absent)',
         'A full natural hazard impact module (NH series: early warning, damage, repair, assistance) was added exclusively in R3 and removed in R4. This is a round-specific expansion.',
         'NH2/NH3/NH7/NH10/NH14-NH17 present only in R3 questionnaire.',
         'NH variables not found in R4/R5 do-file destring blocks.',
         'High','Clearly document that NH variables are R3-only. Do NOT attempt to impute for other rounds. Keep as separate R3 dataset or flag with round indicator.'),

        ('M03 – Shocks','N1, N3','Internet access questions in R3','R3 only',
         'N1 (internet at home) and N3 (internet subscription type) appear only in R3, placed within the Shocks module.',
         'N1/N3 present only in R3 Shocks sheet.',
         'Not found in R4/R5 do-file M03 blocks.',
         'High','Document as R3-only variables. Consider moving to a separate module in the merged dataset rather than keeping under Shocks.'),

        ('M03 – Shocks','SH1','Skip logic changed R1→R2','R1 vs R2+',
         'SH1 (any shock experienced) has different skip routing between R1 and R2.',
         'SH1 skip_rules differ in questionnaire rows R1 vs R2.',
         'Not directly reflected in do-files (skip logic handled in CAPI/KoboCollect).',
         'Medium','Verify that the routing change did not cause differential missingness in SH1b or downstream shock variables between R1 and R2.'),

        # M05
        ('M05 – Income','IC1–IC9, ID1–ID2','Expanded income sub-modules in R3','R3 only (R1,R2,R4,R5 absent)',
         'R3 expanded income module with: IC series (remittances from abroad, domestic support, rental income, pension) and ID series (other income). These 10 new questions are R3-only.',
         'IC1–IC9 and ID1–ID2 present only in R3 Income sheet.',
         'ic4_index_str generated in R3 do-file only.',
         'High','IC/ID variables are R3-only. Flag in merged dataset. Consider whether these should be collected in future rounds.'),

        ('M05 – Income','IA3, IA5','Skip logic added/changed in R2–R3','R2+, R3',
         'IA3 (regular employment income) and IA5 (seasonal income) had skip logic changes between rounds.',
         'skip_rules differ across rounds in questionnaire.',
         'IA3_TOTAL_INCOME and IA5_TOTAL_INCOME sections noted in R1 do-file comments.',
         'Medium','Check whether income zero-values or missing values differ across rounds due to routing changes.'),

        # M06
        ('M06 – Finance','F17, F18','New financial inclusion variables in R5','R5 only',
         'F17 (formal bank account ownership) and F18 (mobile money account: GCash/Maya) added in R5.',
         'F17/F18 present only in R5 Finance sheet.',
         'f17/f18 appear in R5 do-file M06 destring block.',
         'Medium','R5-only; document clearly. These reflect financial inclusion add-on. Not retroactively collectable.'),

        # M07
        ('M07 – Health','H4,H7–H17','Major health module expansion in R5','R5 only (R1–R4 minimal)',
         'R1–R4 had only 3 health questions (H2, H2a, H3). R5 added 13 new questions covering: facility type (H4), transport costs (H7), out-of-pocket payments (H8), prescriptions (H9–H11), hospitalization (H12–H16), and PhilHealth membership (H17).',
         'R5 Health sheet has 16 questions vs 3 in R1–R4.',
         'R5 do-file M07 block has 26 vars vs 3–4 in R1–R4. Health_index_str generated.',
         'High','Health module substantially different in R5. Cannot compare most health variables across rounds. R5 health data is far richer. Maintain separate R5 health file or clearly flag round coverage.'),

        ('M07 – Health','H9','Questionnaire vs do-file mismatch in R5','R5',
         "H9 (prescriptions) is in R5 questionnaire but do-file generates sub-variables (h9a, h9b, h9c, h9_ans) via computation rather than direct destring.",
         'H9 in R5 questionnaire, but do-file splits to h9a/h9b/h9c.',
         'h9a, h9b, h9c, h9_ans generated in R5 M07 block; H9 not destringed directly.',
         'Medium','Document the H9 → h9a/h9b/h9c decomposition. Ensure original H9 response is preserved before splitting.'),

        # M09
        ('M09 – Opinions & Views','V11, V12','Rice tariff questions dropped in R4','R1–R3 present, R4–R5 absent',
         'V11 (support rice tariff) and V12 (support if price rises 10%) were present in R1–R3 but removed from R4 and R5 questionnaires.',
         'V11/V12 absent from R4 and R5 Opinions sheets.',
         'v9_c (rice-related sub-var) disappears in R5 do-file M09 block.',
         'Medium','Note that rice tariff opinion data is only available for R1–R3 in the longitudinal dataset. Flag in panel analysis.'),

        ('M09 – Opinions & Views','V9','Sub-variable decomposition','All rounds',
         'V9 (select_multiple attitudes question) appears as a single item in questionnaires but is split into sub-variables (v9_a, v9_b, v9_c, v9_e–v9_m) in do-files. Not all sub-vars appear in all rounds.',
         'V9 is single question in questionnaire across all rounds.',
         'v9 sub-variables vary: R1 has v9_a,v9_b,v9_c; R4/R5 have v9_a,v9_c,v9_e–v9_m.',
         'Medium','Ensure V9 sub-variable definitions are consistent across rounds. The expansion in R4–R5 may reflect new response options.'),

        # Cross-cutting
        ('All modules','Multiple','Select_multiple → separate binary vars','All rounds',
         'Questions coded as select_multiple in questionnaire (SH1b, A19, A21, F8, V9, H9, H11, H16) are systematically split into binary indicator variables in do-files (e.g., sh1b_1, sh1b_2... or f8_1, f8_2...). The number of binary vars sometimes differs across rounds.',
         'select_multiple items in questionnaires produce many _N suffix vars in data.',
         'R1 has sh1b_27/28/29; R3 do-file adds new sh1b codes; A19 sub-vars differ.',
         'Medium','Document complete code→variable mapping per round for all select_multiple items. Verify no code was silently dropped/added across rounds.'),

        ('All modules','dur_* variables','Zero-duration module flags','All rounds',
         'Some modules show duration=0 in certain rounds (dur_sh, dur_educ, dur_pp, dur_rr) indicating possible module skip or recording failure.',
         'Not directly in questionnaire sheets.',
         'Duration variables destringed in M00/passport sections across all rounds.',
         'Medium','Flag interviews with zero module duration for supervisor review. Check if module was genuinely skipped vs recording error.'),
    ]

    data_row = 4
    for i, (mod, var, issue_type, rounds_aff, desc, q_evid, do_evid, sev, action) in enumerate(issues):
        bg = SEV_BGS.get(sev, CLR_WHITE)
        ws.row_dimensions[data_row].height = 50

        row_data = [mod, var, issue_type, rounds_aff, desc, q_evid, do_evid, sev, action]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=ci, value=val)
            h = 'center' if ci == 8 else 'left'
            wrap = ci not in (1, 2, 3, 4, 8)
            style_cell(cell, size=9, bg=bg if ci != 8 else CLR_WHITE, h_align=h, v_align='top', wrap=wrap)

            if ci == 8:
                cell.font = Font(name=FONT_NAME, bold=True, size=9, color=SEV_COLORS.get(sev, '000000'))
                cell.fill = fill(bg)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if ci in (1, 3):
                cell.font = Font(name=FONT_NAME, bold=True, size=9)

        data_row += 1


# ── BUILD WORKBOOK ────────────────────────────────────────────────────────────
wb = Workbook()
print("Building cover sheet...")
build_cover(wb)
print("Building summary sheet...")
build_summary(wb)
print("Building changes sheet...")
build_changes_sheet(wb)
print("Building DQ issues sheet...")
build_dq_issues_sheet(wb)
print("Building module sheets...")
for mod in MODULES:
    print(f"  {mod}...")
    build_module_sheet(wb, mod)

out_path = _os.path.join(_OUTPUT, 'L2PHL_Questionnaire_Cross_Round_Report.xlsx')
wb.save(out_path)
print(f"\nSaved: {out_path}")
