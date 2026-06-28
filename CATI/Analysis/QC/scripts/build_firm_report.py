"""Generate the firm-actionable Excel QC tracker from issues.json."""
import os, json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from firm_report import firm_rows

_HERE = os.path.dirname(__file__)
_CACHE = os.path.join(_HERE, "..", "cache")
_OUTPUT = os.path.join(_HERE, "..", "output")

HEADERS = ["#", "Module", "Variable", "Issue", "Rounds affected", "Total flagged",
           "Root cause", "Owner", "Evidence — Kobo gate", "Evidence — Do-file",
           "Recommended fix", "Status", "Firm response / fixed?", "Date fixed"]
WIDTHS  = [4, 8, 10, 40, 16, 8, 26, 16, 40, 22, 40, 14, 26, 14]
WRAP    = {"Issue", "Evidence — Kobo gate", "Recommended fix"}
RESP    = {"Firm response / fixed?", "Date fixed"}
NAVY = "FF002244"; CREAM = "FFFFF8DC"

def build(records, today):
    wb = Workbook(); ws = wb.active; ws.title = "Firm QC Tracker"
    rows = firm_rows(records)
    ws.cell(1, 1, "L2PHL CATI — Firm Data Quality Tracker").font = Font(size=14, bold=True)
    ws.cell(2, 1, f"Generated {today} · {len(rows)} open firm issue(s) · please fill the two right-hand columns")
    hdr = 4
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(hdr, c, h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    for i, r in enumerate(rows, 1):
        vals = [i, r["module"], r["variable"], r["issue"], r["rounds"], r["total"],
                r["root_cause"], r["owner"], r["kobo_gate"], r["dofile"],
                r["fix"], r["status"], "", ""]
        rr = hdr + i
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, v)
            cell.alignment = Alignment(vertical="top", wrap_text=HEADERS[c-1] in WRAP)
            cell.border = border
            if HEADERS[c-1] in RESP:
                cell.fill = PatternFill("solid", fgColor=CREAM)
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(hdr + 1, 1)
    return wb

def main():
    issues = json.load(open(os.path.join(_CACHE, "issues.json")))
    today = datetime.date.today().strftime("%Y%m%d")
    wb = build(issues, today)
    os.makedirs(_OUTPUT, exist_ok=True)
    path = os.path.join(_OUTPUT, f"L2PHL_CATI_Firm_QC_Tracker_{today}.xlsx")
    wb.save(path)
    print(f"Firm QC Tracker: {len(firm_rows(issues))} open firm issue(s) -> {os.path.basename(path)}")

if __name__ == "__main__":
    main()
