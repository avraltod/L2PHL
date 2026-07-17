#!/usr/bin/env python3
"""
Parse all L2PHL CATI Kobo XLSForm files (R1-R9) and extract skip logic, mandatory, constraint rules.
Output: kobo_skip_logic.json with comprehensive variable-by-round mapping.
Handles nested groups properly by tracking the group stack.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict, OrderedDict

# File mapping
_KOBO_DIR = str(Path(__file__).resolve().parent.parent.parent.parent / "KOBO")
KOBO_FILES = {
    "1": f"{_KOBO_DIR}/L2PHL_CATI_R01.xlsx",
    "2": f"{_KOBO_DIR}/L2PHL_CATI_R02.xlsx",
    "3": f"{_KOBO_DIR}/L2PHL_CATI_R03.xlsx",
    "4": f"{_KOBO_DIR}/L2PHL_CATI_R04.xlsx",
    "5": f"{_KOBO_DIR}/L2PHL_CATI_R05.xlsx",
    "6": f"{_KOBO_DIR}/L2PHL_CATI_R06.xlsx",
    "7": f"{_KOBO_DIR}/L2PHL_CATI_R07.xlsx",
    "8": f"{_KOBO_DIR}/L2PHL_CATI_R08.xlsx",  # real R08 XLSForm (firm export)
    "9": f"{_KOBO_DIR}/L2PHL_CATI_R09.xlsx",  # real R09 XLSForm (firm export, 06Jul2026)
}

# Top-level group name → Module mapping
# Only top-level groups matter
GROUP_TO_MODULE = {
    "group_intro": "M00",
    "group_demogs": "M01",
    "group_educ": "M02",
    "group_shocks": "M03",
    "group_employment": "M04",
    "group_employment_new": "M04",
    "group_income": "M05",
    "group_income_new": "M05",
    "group_finance": "M06",
    "group_health": "M07",
    "group_food-nonfood": "M08",
    "group_opinions-views": "M09",
    "group_closing": "M99",
}

def normalize_variable_name(name):
    """Remove trailing underscore if present (R3+ convention)."""
    if name and isinstance(name, str):
        return name.rstrip("_")
    return name

def extract_label(row, ws):
    """Extract label from either 'label' or 'label::English (en)' column."""
    label = None
    if "label" in row:
        label = row["label"]
    elif "label::English (en)" in row:
        label = row["label::English (en)"]
    return label

def parse_kobo_file(filepath, round_num):
    """Parse a single Kobo XLSForm file and extract skip logic rules."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb["survey"]

    # Extract column headers
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[col_idx] = cell.value

    variables_by_module = defaultdict(list)
    group_stack = []  # Stack of (group_name, module) tuples
    question_order = []

    # Parse all rows
    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        row = {}
        for col_idx, cell in enumerate(row_cells, 1):
            if col_idx in headers:
                row[headers[col_idx]] = cell.value

        row_type = row.get("type") or ""
        row_type = row_type.strip() if row_type else ""
        row_name = row.get("name") or ""
        row_name = row_name.strip() if row_name else None

        # Handle group stack (capture each group/repeat's own relevant for inheritance)
        if row_type in ["begin_group", "begin_repeat"]:
            group_name = row_name
            # Check if this is a top-level module group
            module = GROUP_TO_MODULE.get(group_name, None)
            _grp_rel = row.get("relevant")
            _grp_rel = _grp_rel.strip() if isinstance(_grp_rel, str) and _grp_rel.strip() else None
            group_stack.append((group_name, module, _grp_rel))

        elif row_type in ["end_group", "end_repeat"]:
            if group_stack:
                group_stack.pop()

        # Skip structural rows
        if row_type in ["begin_group", "end_group", "begin_repeat", "end_repeat", "start_group"]:
            continue

        # Skip rows without names
        if not row_name:
            continue

        # Determine module from group stack (use deepest module in stack)
        module = None
        for group_name, mod, _grp_rel in reversed(group_stack):
            if mod:
                module = mod
                break

        if not module:
            module = "UNKNOWN"

        # Extract rule fields
        label = extract_label(row, {})
        var_type = row.get("type", "")
        relevant = row.get("relevant")
        required = row.get("required")
        constraint = row.get("constraint")
        constraint_message = row.get("constraint_message")

        # Skip notes/calculates without rules
        if var_type in ["note", "calculate"]:
            if not any([relevant, required, constraint]):
                continue

        # Skip metadata/device fields
        if var_type in ["metadata", "deviceid", "start", "end"]:
            continue

        # Skip if name looks like it's just a group marker (ends with _end)
        if row_name.endswith("_end") and var_type in ["text", "note"]:
            # Skip these structural markers
            continue

        # Normalize variable name
        norm_name = normalize_variable_name(row_name)

        # Effective skip-logic = own relevant AND all enclosing repeat/sub-group
        # relevants. Module-level groups (group_shocks, group_health, …) are
        # excluded — they carry the universal ${Z9}=1 consent gate, not a
        # question-specific skip. This recovers gates like SH2 (inside the
        # sh2_ repeat, relevant=${SH1}=1) that the variable's own row omits.
        _own_rel = relevant.strip() if isinstance(relevant, str) and relevant.strip() else None
        _inherited = [gr for (gn, gm, gr) in group_stack
                      if gr and gn not in GROUP_TO_MODULE]
        _parts = _inherited + ([_own_rel] if _own_rel else [])
        if len(_parts) > 1:
            _effective = " and ".join(f"({p})" for p in _parts)
        elif _parts:
            _effective = _parts[0]
        else:
            _effective = None

        # Create rule entry (includes type so Kobo panel can detect type changes per round)
        rule = {
            "type": var_type if var_type else None,
            "relevant": _effective,
            "own_relevant": _own_rel,
            "inherited_relevant": (" and ".join(f"({p})" for p in _inherited) if _inherited else None),
            "required": bool(required) if required else False,
            "constraint": constraint if constraint else None,
            "constraint_message": constraint_message if constraint_message else None,
        }

        # Store
        variables_by_module[module].append({
            "var_name": row_name,
            "normalized_name": norm_name,
            "type": var_type,
            "label": label,
            "rule": rule,
            "row_idx": row_idx,
        })

    return variables_by_module

def merge_rounds():
    """Parse all 5 rounds and merge into a single JSON structure."""

    all_variables = defaultdict(lambda: OrderedDict())

    for round_num, filepath in KOBO_FILES.items():
        print(f"Parsing R{round_num}...")
        try:
            variables_by_module = parse_kobo_file(filepath, round_num)

            for module, vars_list in variables_by_module.items():
                for var_info in vars_list:
                    norm_name = var_info["normalized_name"]

                    if norm_name not in all_variables[module]:
                        all_variables[module][norm_name] = {
                            "normalized_name": norm_name,
                            "original_names": {round_num: var_info["var_name"]},
                            "type": var_info["type"],
                            "label": var_info["label"],
                            "rules_by_round": {},
                        }
                    else:
                        if var_info["type"] and var_info["type"] != all_variables[module][norm_name]["type"]:
                            all_variables[module][norm_name]["type"] = var_info["type"]
                        if var_info["label"] and var_info["label"] != all_variables[module][norm_name]["label"]:
                            all_variables[module][norm_name]["label"] = var_info["label"]

                    all_variables[module][norm_name]["original_names"][round_num] = var_info["var_name"]
                    all_variables[module][norm_name]["rules_by_round"][round_num] = var_info["rule"]

        except Exception as e:
            print(f"Error parsing R{round_num}: {e}")
            raise

    # Reassign root-level "UNKNOWN" vars to their canonical module. A variable that
    # sits inside a module group in one round but moves to the survey root in later
    # rounds (e.g. M09 V1/V5/V11/V12 leave group_opinions-views at R2+) would otherwise
    # be lost from that module for the later rounds. Merge its per-round rules into the
    # module that owns the same variable name.
    unknown = all_variables.get("UNKNOWN", {})
    if unknown:
        name_to_mod = {}
        for _m, _vars in all_variables.items():
            if _m == "UNKNOWN":
                continue
            for _nm in _vars:
                name_to_mod.setdefault(_nm, _m)
        for _nm, _udata in unknown.items():
            _canon = name_to_mod.get(_nm)
            if not _canon:
                continue
            _tgt = all_variables[_canon][_nm]
            for _r, _rule in _udata["rules_by_round"].items():
                _tgt["rules_by_round"][_r] = _rule
            for _r, _onm in _udata["original_names"].items():
                _tgt["original_names"][_r] = _onm

    # Build output structure
    output = {}

    for module in sorted(all_variables.keys()):
        # Skip UNKNOWN module entirely
        if module == "UNKNOWN":
            continue

        module_vars = []

        for norm_name, var_data in all_variables[module].items():
            # Round list is derived from KOBO_FILES at the top of this file —
            # so adding R6/R7/.../R0n to KOBO_FILES auto-expands here too.
            rules_by_round = {}
            for round_num in sorted(KOBO_FILES.keys(), key=int):
                rules_by_round[round_num] = var_data["rules_by_round"].get(round_num, None)

            module_vars.append({
                "name": norm_name,
                "original_names": var_data["original_names"],
                "type": var_data["type"],
                "label": var_data["label"],
                "rules_by_round": rules_by_round,
            })

        output[module] = {
            "variables": module_vars,
        }

    return output

if __name__ == "__main__":
    print("Parsing all 5 L2PHL CATI Kobo XLSForm files...\n")

    result = merge_rounds()

    # Write output — to QC cache dir if running from pipeline, else to working dir
    _here = Path(__file__).resolve().parent
    _cache = _here.parent / 'cache'
    if _cache.exists():
        output_path = str(_cache / 'kobo_skip_logic.json')
    else:
        output_path = "/sessions/dazzling-exciting-goldberg/kobo_skip_logic.json"
    with open(output_path, "w") as f:
        json.dump(result, f, indent=2)

    print(f"\nOutput written to: {output_path}")

    # Print summary
    total_vars = sum(len(m["variables"]) for m in result.values())
    print(f"Total modules: {len(result)}")
    print(f"Total unique variables: {total_vars}")
    for module in sorted(result.keys()):
        print(f"  {module}: {len(result[module]['variables'])} variables")
