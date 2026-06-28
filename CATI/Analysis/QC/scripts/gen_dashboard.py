#!/usr/bin/env python3
import os as _os
_HERE   = _os.path.dirname(_os.path.abspath(__file__))   # scripts/
_QC     = _os.path.dirname(_HERE)                         # Analysis/QC/
_CACHE  = _os.path.join(_QC, 'cache')
_OUTPUT = _os.path.join(_QC, 'output')

"""Generate enhanced L2PHL DQ dashboard v3 — per-module, per-question, per-round"""
import json, os

with open(_os.path.join(_CACHE, 'dq_data.json')) as f:
    dq_raw = json.load(f)
with open(_os.path.join(_CACHE, 'module_tables.json')) as f:
    module_tables = json.load(f)
with open(_os.path.join(_CACHE, 'all_questions.json')) as f:
    all_qs = json.load(f)

# ── Kobo round-presence index ───────────────────────────────────────────
# Pre-scan all 5 Kobo XLSForms to build a lookup of which variables exist
# in which rounds. Used to populate in_R1-R5 on injected tracker rows so
# they show green ✓ instead of red —.
import openpyxl as _openpyxl
_KOBO_DIR = _os.path.join(_os.path.dirname(_QC), '..', 'KOBO')
_KOBO_ROUND_PRESENCE = {}   # {VAR_UPPER: set of round ints}
for _rn in range(1, 9):
    _kobo_file = _os.path.join(_KOBO_DIR, f'L2PHL_CATI_R0{_rn}.xlsx')
    if not _os.path.exists(_kobo_file):
        continue
    _wb = _openpyxl.load_workbook(_kobo_file, read_only=True, data_only=True)
    _ws = _wb['survey']
    _hdrs = [c.value for c in next(_ws.iter_rows(min_row=1, max_row=1))]
    _ni = _hdrs.index('name')
    for _row in _ws.iter_rows(min_row=2):
        _name = _row[_ni].value
        if _name:
            _key = str(_name).upper().rstrip('_')
            _KOBO_ROUND_PRESENCE.setdefault(_key, set()).add(_rn)
    _wb.close()

def _set_round_presence(synth_row, var_name):
    """Populate in_R1-R5 and first_round on a synthetic tracker row
    based on Kobo XLSForm presence."""
    _key = var_name.upper().rstrip('_')
    _rounds = _KOBO_ROUND_PRESENCE.get(_key, set())
    _first = None
    for _r in range(1, 9):
        _present = _r in _rounds
        synth_row[f'in_R{_r}'] = '✓' if _present else ''
        if _present and _first is None:
            _first = f'R{_r}'
    synth_row['first_round'] = _first or ''
    synth_row['status'] = 'All rounds' if len(_rounds) == 5 else (
        f'R{min(_rounds)}–R{max(_rounds)}' if _rounds else '')

# ── M01 sub-variable expansion ───────────────────────────────────────────
# The questionnaire has aggregate questions (D25, D26, M10, M13) but the
# pooled data splits them into sub-variables (_oth, _1/_2/_3). Expand so
# the tracker matches the DQ panels.
if 'M01' in module_tables:
    import copy as _copy

    # ── Kobo type corrections ────────────────────────────────────────────
    # Questionnaire xlsx uses different type names; override with Kobo types:
    #   select_one → Categorical, text → Text, integer → Integer
    _KOBO_TYPE_FIX = {
        'D5a': 'Categorical',         # select_one confirm_mem
        'D25': 'Categorical',         # select_one leave_reason
        'D25_oth': 'Text',            # text (other specify)
        'D26': 'Categorical',         # parent (select_one country/province/city)
        'D26_1': 'Categorical',       # select_one country
        'D26_2': 'Categorical',       # select_one province
        'D26_3': 'Categorical',       # select_one city
        'D27': 'Text',                # text (correct name)
        'D28': 'Integer',             # integer (correct age)
        'D33': 'Categorical',         # select_one gender
        'D29': 'Categorical',         # select_one yes-none
        'D30': 'Text',                # text (new member name)
        'D31': 'Categorical',         # select_one gender
        'D32': 'Integer',             # integer (new member age)
        'D6':  'Categorical',         # select_one relationship
        'M13': 'Categorical',         # select_one movein
        'M13_oth': 'Text',            # text (other specify)
        'M10': 'Categorical',         # parent (select_one country/province/city)
        'M10_1': 'Categorical',       # select_one country
        'M10_2': 'Categorical',       # select_one province
        'M10_3': 'Categorical',       # select_one city
    }
    # Build case-insensitive lookup for type fixes
    _KOBO_TYPE_FIX_UPPER = {k.upper(): v for k, v in _KOBO_TYPE_FIX.items()}
    for _row in module_tables['M01']:
        _fix = _KOBO_TYPE_FIX_UPPER.get(_row['variable'].upper())
        if _fix:
            _row['question_type'] = _fix

    # ── Sub-variable expansion ───────────────────────────────────────────
    _M01_EXPANSIONS = {
        'D25': [('D25_oth', 'OTHER REASON FOR LEAVING', 'Text')],
        'D26': [
            ('D26_1', 'COUNTRY WHERE MOVED', 'Categorical'),
            ('D26_2', 'PROVINCE WHERE MOVED', 'Categorical'),
            ('D26_3', 'CITY WHERE MOVED', 'Categorical'),
        ],
        'M13': [('M13_oth', 'OTHER REASON FOR MOVING IN', 'Text')],
        'M10': [
            ('M10_1', 'COUNTRY WHERE CAME FROM', 'Categorical'),
            ('M10_2', 'PROVINCE WHERE CAME FROM', 'Categorical'),
            ('M10_3', 'CITY WHERE CAME FROM', 'Categorical'),
        ],
    }
    # Collect existing variable names (uppercase) to avoid duplicates
    _existing_vars = {_row['variable'].upper() for _row in module_tables['M01']}
    _new_m01 = []
    for _row in module_tables['M01']:
        _v = _row['variable'].upper()
        if _v in ('D26', 'M10'):
            # Replace parent with sub-variables (only if not already present)
            for _nv, _nt, _tt in _M01_EXPANSIONS[_v]:
                if _nv.upper() in _existing_vars:
                    continue  # Already exists from questionnaire parsing
                _r2 = _copy.deepcopy(_row)
                _r2['variable'] = _nv
                _r2['question_title'] = _nt
                if _tt: _r2['question_type'] = _tt
                _set_round_presence(_r2, _nv)
                _new_m01.append(_r2)
        else:
            _new_m01.append(_row)
            if _v in _M01_EXPANSIONS:
                for _nv, _nt, _tt in _M01_EXPANSIONS[_v]:
                    if _nv.upper() in _existing_vars:
                        continue  # Already exists from questionnaire parsing
                    _r2 = _copy.deepcopy(_row)
                    _r2['variable'] = _nv
                    _r2['question_title'] = _nt
                    if _tt: _r2['question_type'] = _tt
                    _set_round_presence(_r2, _nv)
                    _new_m01.append(_r2)
    module_tables['M01'] = _new_m01

    # ── M01 case normalisation: align tracker names to Kobo case ─────────
    # Kobo uses D5a/D25_oth/M13_oth but questionnaire parsing yields D5A/D25_OTH/M13_OTH
    _M01_CASE_MAP = {
        'D5A': 'D5a',
        'D25_OTH': 'D25_oth',
        'M13_OTH': 'M13_oth',
    }
    for _row in module_tables['M01']:
        _norm = _M01_CASE_MAP.get(_row['variable'])
        if _norm:
            _row['variable'] = _norm

    # ── M01 authoritative variable list (17 vars, user-specified Kobo order) ──
    _M01_AUTHORITATIVE = [
        'hhid', 'fmid', 'hhsize',
        'D5a',          # expands to age, gender, isfmid sub-rows
        'D6',           # relationship
        'D25', 'D25_oth',
        'D26_1', 'D26_2', 'D26_3',
        'M13', 'M13_oth',
        'M10_1', 'M10_2', 'M10_3',
        'dur_rr',
    ]
    _M01_AUTH_SET = {v.upper() for v in _M01_AUTHORITATIVE}

    # ── Drop non-authoritative tracker vars ────────────────────────────
    module_tables['M01'] = [
        r for r in module_tables['M01'] if r['variable'].upper() in _M01_AUTH_SET
    ]

    # ── Inject missing vars as synthetic rows ──────────────────────────
    _m01_existing = {r['variable'].upper() for r in module_tables['M01']}
    _M01_INJECT = {
        'hhid':   ('HOUSEHOLD ID',          'Text'),
        'fmid':   ('FAMILY MEMBER ID',      'Text'),
        'hhsize': ('HOUSEHOLD SIZE',         'Integer'),
        'dur_rr': ('ROSTER DURATION (MIN)',  'Decimal'),
    }
    _template = module_tables['M01'][0] if module_tables['M01'] else {}
    for _var, (_title, _type) in _M01_INJECT.items():
        if _var.upper() not in _m01_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M01'].append(_synth)

    # ── Reorder tracker to authoritative order ─────────────────────────
    _by_var = {}
    for _row in module_tables['M01']:
        _by_var[_row['variable'].upper()] = _row
    _ordered_m01 = []
    for _v in _M01_AUTHORITATIVE:
        _r = _by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v  # normalise case
            _ordered_m01.append(_r)
    _seen = {v.upper() for v in _M01_AUTHORITATIVE}
    for _row in module_tables['M01']:
        if _row['variable'].upper() not in _seen:
            _ordered_m01.append(_row)
    module_tables['M01'] = _ordered_m01

_panel_path = _os.path.join(_CACHE, 'panel_data.json')
if _os.path.exists(_panel_path):
    with open(_panel_path) as f:
        panel_raw = json.load(f)
else:
    panel_raw = {}

_iview_path = _os.path.join(_CACHE, 'interviewer_data.json')
if _os.path.exists(_iview_path):
    with open(_iview_path) as f:
        iview_raw = json.load(f)
else:
    iview_raw = {}

_kobo_path = _os.path.join(_CACHE, 'kobo_skip_logic.json')
if _os.path.exists(_kobo_path):
    with open(_kobo_path) as f:
        kobo_raw = json.load(f)
else:
    kobo_raw = {}

_issues_path = _os.path.join(_CACHE, 'issues.json')
issues_raw = json.load(open(_issues_path)) if _os.path.exists(_issues_path) else []
_isum_path = _os.path.join(_CACHE, 'issue_summary.json')
isum_raw = json.load(open(_isum_path)) if _os.path.exists(_isum_path) else {}

# ── M01 Kobo variable reorder ───────────────────────────────────────────
# The Kobo XLSForm has D33 after demogs_end (physical form position).
# Reorder to match the Question-Level Cross-Round Tracker order, and
# exclude form-control variables (demogs_end, etc.).
if 'M01' in kobo_raw and 'variables' in kobo_raw['M01']:
    # Use M01 authoritative list for Kobo reorder (matches _M01_AUTHORITATIVE)
    _M01_KOBO_AUTH = [
        'hhid', 'fmid', 'hhsize',
        'D5a',          # isfmid / age / gender confirmation
        'D6',           # relationship
        'D25', 'D25_oth',
        'D26_1', 'D26_2', 'D26_3',
        'M13', 'M13_oth',
        'M10_1', 'M10_2', 'M10_3',
        'dur_rr',
    ]
    _KOBO_EXCLUDE = {'demogs_end'}  # form control variables, not data
    # Case-insensitive lookup, stripping trailing underscores (R3-R5 convention)
    _kobo_by_name = {}
    for _v in kobo_raw['M01']['variables']:
        _key = _v['name'].upper().rstrip('_')
        if _key not in _kobo_by_name:
            _kobo_by_name[_key] = _v
    _ordered = []
    _kobo_seen = set()
    for _auth_var in _M01_KOBO_AUTH:
        _key = _auth_var.upper().rstrip('_')
        if _key in _kobo_by_name and _key not in _kobo_seen:
            _ordered.append(_kobo_by_name[_key])
            _kobo_seen.add(_key)
        elif _key not in _kobo_seen:
            # Inject synthetic entry for vars not in Kobo (CRITICAL: rules_by_round)
            _ordered.append({
                'name': _auth_var,
                'label': _auth_var,
                'rules_by_round': {str(r): [] for r in range(1, 9)},
            })
            _kobo_seen.add(_key)
    # Append remaining Kobo vars not in authoritative list (safety net)
    for _v in kobo_raw['M01']['variables']:
        _key = _v['name'].upper().rstrip('_')
        if _key not in _kobo_seen and _v['name'] not in _KOBO_EXCLUDE:
            _ordered.append(_v)
    kobo_raw['M01']['variables'] = _ordered

# ── M00 post-processing ─────────────────────────────────────────────────
# Align Tracker, Heatmap, and Kobo panels to the SAME authoritative variable
# list (user's original M00 list minus Z21/PII).  Like M01, every panel must
# show the same set of variables in the same order.
#
# Authoritative order (25 vars):
#   Z0_first Z0_last int_id hhid D3 Z6 Z7 Z8 member_called call_status1
#   Z16 Z17 member_talkedto n_Z17 Z9 Z18 Z18_oth Z19 backgound_audio
#   Z20 Z1 Z2 Z3 Z4 Z5

_M00_AUTHORITATIVE = [
    'Z0_first', 'Z0_last', 'int_id', 'hhid', 'D3', 'Z6', 'Z7', 'Z8',
    'member_called', 'call_status1', 'Z16', 'Z17', 'member_talkedto',
    'n_Z17', 'Z9', 'Z18', 'Z18_oth', 'Z19', 'backgound_audio',
    'Z20', 'Z1', 'Z2', 'Z3', 'Z4', 'Z5',
]
_M00_AUTH_SET = {v.upper() for v in _M00_AUTHORITATIVE}

if 'M00' in module_tables:
    import copy as _copy

    # ── 1. Drop vars NOT in authoritative list ──────────────────────────
    _M00_DROP = {'Z21', 'D2', 'D22', 'FMID'}
    module_tables['M00'] = [
        r for r in module_tables['M00'] if r['variable'].upper() not in _M00_DROP
    ]

    # ── 2. Expand Z0 → Z0_first + Z0_last ──────────────────────────────
    _new_m00 = []
    for _row in module_tables['M00']:
        if _row['variable'].upper() == 'Z0':
            for _sub in ('Z0_first', 'Z0_last'):
                _r2 = _copy.deepcopy(_row)
                _r2['variable'] = _sub
                _r2['question_title'] = 'FIRST NAME OF INTERVIEWER' if _sub == 'Z0_first' else 'LAST NAME OF INTERVIEWER'
                _new_m00.append(_r2)
        else:
            _new_m00.append(_row)
    module_tables['M00'] = _new_m00

    # ── 3. Inject missing vars as synthetic rows ────────────────────────
    # These exist in Kobo XLSForm but not in questionnaire Excel
    _existing = {r['variable'].upper() for r in module_tables['M00']}
    _INJECT = {
        'int_id':          ('int_id',                    'Categorical'),
        'hhid':            ('hhid',                      'Text'),
        'member_called':   ('MEMBER CALLED',             'Categorical'),
        'call_status1':    ('CALL STATUS',               'Categorical'),
        'member_talkedto': ('MEMBER TALKED TO',          'Categorical'),
        'n_Z17':           ('COUNT OF Z17 SELECTIONS',   'Integer'),
        'backgound_audio': ('BACKGROUND AUDIO CHECK',    'Categorical'),
        'Z18_oth':         ('OTHER REFUSAL REASON',      'Text'),
    }
    # Use first existing row as template for round structure
    _template = module_tables['M00'][0] if module_tables['M00'] else {}
    for _var, (_title, _type) in _INJECT.items():
        if _var.upper() not in _existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M00'].append(_synth)

    # ── 4. Reorder tracker to match authoritative order ─────────────────
    _by_var = {}
    for _row in module_tables['M00']:
        _by_var[_row['variable'].upper()] = _row
    _ordered_m00 = []
    for _v in _M00_AUTHORITATIVE:
        _r = _by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v  # normalise case to match authoritative
            _ordered_m00.append(_r)
    # Append any remaining rows not in authoritative list (safety net)
    _seen = {v.upper() for v in _M00_AUTHORITATIVE}
    for _row in module_tables['M00']:
        if _row['variable'].upper() not in _seen:
            _ordered_m00.append(_row)
    module_tables['M00'] = _ordered_m00

    # ── 5. Enrich empty skip_r* from Kobo relevant conditions ───────────
    _M00_KOBO_SKIP = {
        'Z16':     '${call_status1}=1',
        'Z18':     '${Z9}=2',
        'Z18_OTH': '${Z18}=5',
        'Z19':     '${Z9}=1',
        'Z20':     '${Z9}=1',
        'Z1':      '${Z20}=2',
        'Z2':      '${Z20}=2',
        'Z3':      '${Z20}=2',
        'Z4':      '${Z20}=2',
        'Z5':      '${Z20}=2',
    }
    _ROUNDS = ['r1', 'r2', 'r3', 'r4', 'r5']
    for _row in module_tables['M00']:
        _kobo_skip = _M00_KOBO_SKIP.get(_row['variable'].upper())
        if _kobo_skip:
            for _r in _ROUNDS:
                _key = f'skip_{_r}'
                if _key in _row and not _row[_key]:
                    _row[_key] = _kobo_skip

# ── M00 Kobo panel: drop non-authoritative vars + reorder ──────────────
_M00_KOBO_DROP = {'Z21', 'HHSIZE_WRONG', 'REPLACEMENT_HHID', 'Z4_TXT'}
if 'M00' in kobo_raw and 'variables' in kobo_raw.get('M00', {}):
    kobo_raw['M00']['variables'] = [
        v for v in kobo_raw['M00']['variables']
        if v['name'].upper() not in _M00_KOBO_DROP
    ]
    # Reorder Kobo to match authoritative order (like M01 _M01_KOBO_ORDER)
    _kobo_by_name = {v['name'].upper(): v for v in kobo_raw['M00']['variables']}
    _kobo_ordered = []
    _kobo_seen = set()
    for _v in _M00_AUTHORITATIVE:
        _k = _v.upper()
        if _k in _kobo_by_name and _k not in _kobo_seen:
            _kobo_ordered.append(_kobo_by_name[_k])
            _kobo_seen.add(_k)
    # Append any remaining Kobo vars not in authoritative list
    for _v in kobo_raw['M00']['variables']:
        if _v['name'].upper() not in _kobo_seen:
            _kobo_ordered.append(_v)
    kobo_raw['M00']['variables'] = _kobo_ordered

# ── M02 post-processing ─────────────────────────────────────────────────
# Authoritative list: ed15, ed16, ed16_1..ed16_14, ed16_96, ed16_oth, dur_edu
# ed16_N are split dummies from select_multiple ED16 (dropout reason codes).
# All ed16_* conditional on ed15=2 (not currently studying).
# R1 Kobo had person-indexed structure (ED15_1..ED15_25); R2+ generic.
_M02_AUTHORITATIVE = [
    'ed15', 'ed16',
    'ed16_1', 'ed16_2', 'ed16_3', 'ed16_4', 'ed16_5',
    'ed16_6', 'ed16_7', 'ed16_8', 'ed16_9', 'ed16_10',
    'ed16_11', 'ed16_12', 'ed16_13', 'ed16_14', 'ed16_96',
    'ed16_oth',
    'dur_edu',
]
_M02_AUTH_SET = {v.upper() for v in _M02_AUTHORITATIVE}

if 'M02' in module_tables:
    import copy as _copy

    # ── 1. Drop non-authoritative vars from tracker ─────────────────────
    module_tables['M02'] = [
        r for r in module_tables['M02'] if r['variable'].upper() in _M02_AUTH_SET
    ]

    # ── 2. Inject missing vars ──────────────────────────────────────────
    _existing = {r['variable'].upper() for r in module_tables['M02']}
    # ED16 dropout reason codes: select_multiple split dummies
    _ED16_REASONS = {
        'ed16_1':  'Illness / health problems',
        'ed16_2':  'Financial constraints',
        'ed16_3':  'Family matters / responsibility',
        'ed16_4':  'Work / employment',
        'ed16_5':  'Lack of interest / motivation',
        'ed16_6':  'Distance / access to school',
        'ed16_7':  'Safety / security concerns',
        'ed16_8':  'Early marriage / pregnancy',
        'ed16_9':  'Academic difficulties',
        'ed16_10': 'Weather / natural disasters',
        'ed16_11': 'Disability',
        'ed16_12': 'Graduated / completed',
        'ed16_13': 'Conflict / displacement',
        'ed16_14': 'COVID-related',
        'ed16_96': 'Other reason (specify)',
    }
    _M02_INJECT = {
        'ed16_oth': ('OTHER REASON (SPECIFY)',     'Text'),
        'dur_edu':  ('EDUCATION DURATION (MIN)',    'Decimal'),
    }
    # Add all ed16_N dummies
    for _code, _label in _ED16_REASONS.items():
        _M02_INJECT[_code] = (_label, 'Binary')
    _template = module_tables['M02'][0] if module_tables['M02'] else {}
    for _var, (_title, _type) in _M02_INJECT.items():
        if _var.upper() not in _existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M02'].append(_synth)

    # ── 3. Reorder tracker to match authoritative order ─────────────────
    _by_var = {}
    for _row in module_tables['M02']:
        _by_var[_row['variable'].upper()] = _row
    _ordered_m02 = []
    for _v in _M02_AUTHORITATIVE:
        _r = _by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m02.append(_r)
    module_tables['M02'] = _ordered_m02

# ── M02 Kobo panel: collapse person-indexed vars + reorder ──────────────
# R1 Kobo uses ED15_1..ED15_25, ED16_1..ED16_25, ED16_1oth..ED16_25oth
# (person-indexed repeat group). R2-5 uses generic ed15/ed16/ed16_oth.
# Collapse to authoritative generic vars. ed16_N split dummies are
# Stata-only (not in Kobo XLSForm) — injected as synthetic entries.
if 'M02' in kobo_raw and 'variables' in kobo_raw.get('M02', {}):
    import re as _re
    _m02_kobo_collapsed = {}  # name → first matching variable dict
    for _v in kobo_raw['M02']['variables']:
        _name = _v['name']
        # Map person-indexed names to generic: ED15_1 → ed15, ED16_3oth → ed16_oth
        _generic = _name
        if _re.match(r'^ED16_\d+oth', _name, _re.I):
            _generic = 'ed16_oth'
        elif _re.match(r'^ED16_\d+oth_new', _name, _re.I):
            _generic = 'ed16_oth'
        elif _re.match(r'^ED16_\d+_new', _name, _re.I):
            _generic = 'ed16'
        elif _re.match(r'^ED16_\d+', _name, _re.I):
            _generic = 'ed16'
        elif _re.match(r'^ED15_\d+_new', _name, _re.I):
            _generic = 'ed15'
        elif _re.match(r'^ED15_\d+', _name, _re.I):
            _generic = 'ed15'
        _gen_upper = _generic.upper()
        # Only collapse to ed15, ed16, ed16_oth (Kobo-level vars)
        if _gen_upper in ('ED15', 'ED16', 'ED16_OTH') and _gen_upper not in _m02_kobo_collapsed:
            _collapsed = dict(_v)
            _collapsed['name'] = _generic
            _m02_kobo_collapsed[_gen_upper] = _collapsed

    # Build ordered list from authoritative order
    _m02_kobo_ordered = []
    _synth_rules_tpl = {str(r): {'type': None, 'relevant': None, 'required': False,
                                  'constraint': None, 'constraint_message': None}
                        for r in range(1, 9)}
    for _v in _M02_AUTHORITATIVE:
        _entry = _m02_kobo_collapsed.get(_v.upper())
        if _entry:
            _m02_kobo_ordered.append(_entry)
        else:
            # Inject synthetic entry (ed16_N dummies, dur_edu)
            _is_dummy = _v.startswith('ed16_') and _v != 'ed16_oth'
            _m02_kobo_ordered.append({
                'name': _v,
                'type': 'binary (split dummy)' if _is_dummy else 'decimal' if _v == 'dur_edu' else 'metadata',
                'label': _v.upper() + ' (Stata split dummy from select_multiple ED16)' if _is_dummy else _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': {k: dict(v) for k, v in _synth_rules_tpl.items()},
            })
    kobo_raw['M02']['variables'] = _m02_kobo_ordered

# ── M00 Heatmap: bucket-and-reorder to authoritative order ────────────
if 'M00' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m00_hm_buckets = {v.lower(): [] for v in _M00_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M00']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m00_hm_buckets:
            _m00_hm_buckets[_stata].append(_r)
            continue
        for _auth in _M00_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m00_hm_buckets[_auth.lower()].append(_r)
                break
    _m00_hm_ordered = []
    for _v in _M00_AUTHORITATIVE:
        _m00_hm_ordered.extend(_m00_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M00'] = _m00_hm_ordered

# ── M01 Heatmap: bucket-and-reorder to authoritative order ────────────
# Note: _M01_AUTHORITATIVE is defined earlier in the M01 tracker block
if 'M01' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m01_auth_for_hm = [
        'D5a', 'D25', 'D25_oth',
        'D26_1', 'D26_2', 'D26_3',
        'D27', 'D28', 'D33',
        'D6',
        'M13', 'M13_oth',
        'M10_1', 'M10_2', 'M10_3',
        'D29', 'D30', 'D31', 'D32',
    ]
    # Heatmap uses Stata names in parens: 'D5A (isfmid)' → extract 'isfmid'
    # But M01 DISPLAY_NAMES maps stata → questionnaire, so we need reverse lookup
    # The heatmap var labels are like 'D5A (isfmid)' — extract stata name from parens
    _m01_hm_buckets = {}
    # Build buckets keyed by DISPLAY_NAME questionnaire prefix (lowercase)
    for _v in _m01_auth_for_hm:
        _m01_hm_buckets[_v.lower()] = []
    for _r in dq_raw['heatmap_data']['M01']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        # Try exact match on stata name against known mappings
        _matched = False
        for _auth in _m01_auth_for_hm:
            if _stata == _auth.lower() or _stata.startswith(_auth.lower() + '_'):
                _m01_hm_buckets[_auth.lower()].append(_r)
                _matched = True
                break
        if not _matched:
            # Also try the var label prefix (before parentheses)
            _label_prefix = _var.split('(')[0].strip().lower() if '(' in _var else _var.lower()
            for _auth in _m01_auth_for_hm:
                if _label_prefix == _auth.lower() or _label_prefix.startswith(_auth.lower()):
                    _m01_hm_buckets[_auth.lower()].append(_r)
                    _matched = True
                    break
        if not _matched:
            # Append unmatched rows at end
            _m01_hm_buckets.setdefault('__unmatched__', []).append(_r)
    _m01_hm_ordered = []
    for _v in _m01_auth_for_hm:
        _m01_hm_ordered.extend(_m01_hm_buckets.get(_v.lower(), []))
    _m01_hm_ordered.extend(_m01_hm_buckets.get('__unmatched__', []))
    dq_raw['heatmap_data']['M01'] = _m01_hm_ordered

# ── M02 Heatmap: bucket-and-reorder to authoritative order ────────────
# Pipeline produces: ED15 (ed15), ED16 (ed16), ED16 (ed16) (multi), ED16_oth (ed16_oth)
# ed16_N split dummies and dur_edu are injected via MODULE_VAR_ORDER augmentation.
if 'M02' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m02_hm_buckets = {v.lower(): [] for v in _M02_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M02']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue  # Skip multi-select aggregate row
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m02_hm_buckets:
            _m02_hm_buckets[_stata].append(_r)
            continue
        # Try prefix match for any pipeline sub-variables
        for _auth in _M02_AUTHORITATIVE:
            if _stata.startswith(_auth.lower()):
                _m02_hm_buckets[_auth.lower()].append(_r)
                break
    _m02_hm_ordered = []
    for _v in _M02_AUTHORITATIVE:
        _m02_hm_ordered.extend(_m02_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M02'] = _m02_hm_ordered

# ── M03 Shocks + Natural Hazards standardisation ────────────────────────
# Authoritative list: 15 vars — matches HF pooled dataset
# (hf_l2phl_analysis@AP@20260322.do §M03, line 633 `order` command)
#
# ── 2026-04-03 audit notes ──────────────────────────────────────────────
# REMOVED from previous 21-var list:
#   sh1b_oth — lost during shock splitting (`cap drop sh1b sh1b*` in round do-files
#              wipes all sh1b dummies including _oth before the reshape merge)
#   sh2_oth  — lost during shock splitting/merge step (same `cap drop` pattern)
#   nh14, nh15, nh16, nh17 — explicitly dropped as processing artefacts in HF do-file
#              (line 630: `cap drop dur_sh nh14_1 nh15_1 nh16_1 nh17_1_1 nh14_2 nh17_1_2`)
#              These exist in R3 round-level .dta but not in the pooled panel.
#
# STRUCTURAL NOTE — variable meaning changes after reshape:
#   sh1b in Kobo: select_multiple string ("1 3 7") → split into binary dummies
#   sh1b in pooled data: renamed shock_id (shock type code for 1st instance)
#   sh1b_1–sh1b_4 in pooled: shock type code per instance (NOT binary dummies)
#   sh2_N_M in pooled: coping mechanism M for shock instance N (double-indexed)
#   nh7_M_N, nh10_M_N, nh3_M_N: select_multiple code M × hazard instance N
#
# R1 vs R2–R5 structural difference:
#   R1 creates 22 binary sh1b dummies (one per shock type) then reshapes
#   R2–R5 use numeric repeat-shock structure directly
#   HF pooling harmonizes both into the R2–R5 repeat-shock format
# ─────────────────────────────────────────────────────────────────────────
_M03_AUTHORITATIVE = [
    # Core shocks (R1-R5): sh1 gate → sh1b (shock types) → sh2 (coping) → sh3/sh4
    'sh1', 'sh1b', 'sh2',
    'sh3', 'sh4', 'el5', 'n5',
    # Natural hazard block (R3+ only): nh2 → nh7 → nh10 → nh3
    # After reshape wide, these become double-indexed: nh7_1_1, nh10_3_1, etc.
    'nh2',
    'nh7', 'nh7_oth',
    'nh10', 'nh10_oth',
    'nh3',
    # Internet block (R3 only)
    'n1', 'n3',
]
_M03_AUTH_SET = {v.upper() for v in _M03_AUTHORITATIVE}

# ── M03 Tracker: drop non-auth vars, inject missing, reorder ──────────
if 'M03' in module_tables:
    # 1. Drop non-authoritative vars (UTILITIES, N1, N3 already present but
    #    we keep them; drop only truly non-auth like UTILITIES)
    module_tables['M03'] = [
        r for r in module_tables['M03']
        if r['variable'].upper() in _M03_AUTH_SET
    ]

    # 2. Inject missing vars (nh7_oth / nh10_oth may not appear in tracker)
    #    sh1b_oth and sh2_oth removed 2026-04-03: dropped during shock splitting
    _m03_existing = {r['variable'].upper() for r in module_tables['M03']}
    _M03_INJECT = {
        'nh7_oth':  ('OTHER WARNING CHANNEL (SPECIFY)',   'Text'),
        'nh10_oth': ('OTHER ACTION TAKEN (SPECIFY)',      'Text'),
    }
    _template = module_tables['M03'][0] if module_tables['M03'] else {}
    for _var, (_title, _type) in _M03_INJECT.items():
        if _var.upper() not in _m03_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M03'].append(_synth)

    # 3. Reorder to authoritative order
    _m03_by_var = {}
    for _row in module_tables['M03']:
        _m03_by_var[_row['variable'].upper()] = _row
    _ordered_m03 = []
    for _v in _M03_AUTHORITATIVE:
        _r = _m03_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m03.append(_r)
    module_tables['M03'] = _ordered_m03

# ── M03 Kobo: drop non-auth vars, reorder ─────────────────────────────
if 'M03' in kobo_raw and 'variables' in kobo_raw.get('M03', {}):
    # Build lookup by name (case-insensitive)
    _m03_kobo_lookup = {}
    for _v in kobo_raw['M03']['variables']:
        _key = _v['name'].upper().rstrip('_')  # SH1b_ → SH1B, SH2_ → SH2
        # Handle trailing underscore Kobo naming (SH1b_ → sh1b, SH2_ → sh2)
        if _key not in _m03_kobo_lookup:
            _m03_kobo_lookup[_key] = _v

    # Also map specific known aliases
    _KOBO_ALIASES = {
        'SH1B': ['SH1B', 'SH1B_'],
        'SH2': ['SH2', 'SH2_'],
        'NH7': ['NH7', 'NH7_'],
        'NH10': ['NH10', 'NH10_'],
        'NH3': ['NH3', 'NH3_'],
        'N3': ['N3', 'N3_'],
    }

    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}

    _m03_kobo_ordered = []
    for _v in _M03_AUTHORITATIVE:
        _key = _v.upper()
        _entry = _m03_kobo_lookup.get(_key)
        # Try aliases if not found
        if not _entry and _key in _KOBO_ALIASES:
            for _alias in _KOBO_ALIASES[_key]:
                _entry = _m03_kobo_lookup.get(_alias)
                if _entry:
                    break
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m03_kobo_ordered.append(_collapsed)
        else:
            # Inject synthetic entry for vars not in Kobo
            _m03_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M03']['variables'] = _m03_kobo_ordered

# ── M03 Heatmap: filter to authoritative vars, remove (multi), reorder ──
if 'M03' in dq_raw.get('heatmap_data', {}):
    import re as _re
    # Build lookup: authoritative_key → list of heatmap rows (for SATA sub-vars)
    _m03_hm_buckets = {v.lower(): [] for v in _M03_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M03']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        # Exact match
        if _stata in _m03_hm_buckets:
            _m03_hm_buckets[_stata].append(_r)
            continue
        # Prefix match: nh7_1_1 → nh7, sh1b_1 → sh1b
        for _auth in _M03_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m03_hm_buckets[_auth.lower()].append(_r)
                break
    # Emit in authoritative order
    _m03_hm_ordered = []
    for _v in _M03_AUTHORITATIVE:
        _m03_hm_ordered.extend(_m03_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M03'] = _m03_hm_ordered

# ── M04 Employment standardisation ──────────────────────────────────────
# Authoritative list: 28 vars — R4-R5 routing (a24/a25/a26/a27) + core employment
#
# ── 2026-04-03 audit notes ──────────────────────────────────────────────
# a24, a25, a26, a27: R4–R5 panel continuity routing vars. Kept in HF
#   pooled data (2026-04-03) — missing for R1–R3. Gated with __round__ [4,5].
#
# a19 (benefits) and a21 (transport): R1–R3 only — NOT collected R4–R5.
#   HF pooled data saves split dummies (a19_1–a19_5, a21_1–a21_3) but these
#   are NOT in this authoritative list. By design: dashboard tracks questionnaire-
#   level variables (was the question answered?), not derived dummies.
#   Heatmap shows aggregate a19/a21 missing rates, not individual codes.
#
# Kobo naming shift: A19→A19_, A21→A21_ in R3+ (trailing underscore).
#   Dashboard normalizes to lowercase a19, a21.
#
# A10/A11 routing gates are approximate for R4–R5: Kobo references a24/a26/a27
#   (not in pooled data), so dashboard uses a1 as primary gate.
# ─────────────────────────────────────────────────────────────────────────
_M04_AUTHORITATIVE = [
    # Routing vars (R4–R5 only, Kobo order): panel continuity tracking
    'a24', 'a27', 'a25', 'a1', 'a26',
    # Core employment (R1–R5, varying skip logic)
    'a3', 'a3_oth', 'a4', 'a4_oth', 'a5', 'a5_oth',
    'a6', 'a7', 'a8', 'a9',
    'a10', 'a11', 'a16', 'a17', 'a18',
    # Benefits + transport (R1–R3 only; split dummies in HF but not here)
    'a19', 'a19_oth', 'a20', 'a21', 'a21_oth', 'a21_own',
    'a22', 'a23',
]
_M04_AUTH_SET = {v.upper() for v in _M04_AUTHORITATIVE}

# ── M04 Tracker: drop non-auth vars, inject missing, reorder ──────────
if 'M04' in module_tables:
    # 1. Drop non-authoritative (TRANSPORTATION, etc.)
    module_tables['M04'] = [
        r for r in module_tables['M04']
        if r['variable'].upper() in _M04_AUTH_SET
    ]

    # 2. Inject missing vars
    _m04_existing = {r['variable'].upper() for r in module_tables['M04']}
    _M04_INJECT = {
        'a3_oth':  ('OTHER REASON FOR NOT WORKING (SPECIFY)',  'Text'),
        'a4_oth':  ('OTHER OCCUPATION (SPECIFY)',              'Text'),
        'a5_oth':  ('OTHER INDUSTRY (SPECIFY)',                'Text'),
        'a19_oth': ('OTHER EMPLOYMENT BENEFIT (SPECIFY)',      'Text'),
        'a21_oth': ('OTHER MODE OF TRANSPORT (SPECIFY)',       'Text'),
        'a21_own': ('OWN VEHICLE TYPE (SPECIFY)',              'Text'),
    }
    _template = module_tables['M04'][0] if module_tables['M04'] else {}
    for _var, (_title, _type) in _M04_INJECT.items():
        if _var.upper() not in _m04_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M04'].append(_synth)

    # 3. Reorder to authoritative order
    _m04_by_var = {}
    for _row in module_tables['M04']:
        _m04_by_var[_row['variable'].upper()] = _row
    _ordered_m04 = []
    for _v in _M04_AUTHORITATIVE:
        _r = _m04_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m04.append(_r)
    module_tables['M04'] = _ordered_m04

# ── M04 Kobo: drop _new variants and non-auth vars, reorder ───────────
if 'M04' in kobo_raw and 'variables' in kobo_raw.get('M04', {}):
    _m04_kobo_lookup = {}
    for _v in kobo_raw['M04']['variables']:
        _key = _v['name'].upper().rstrip('_')
        if _key not in _m04_kobo_lookup:
            _m04_kobo_lookup[_key] = _v

    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}

    _m04_kobo_ordered = []
    for _v in _M04_AUTHORITATIVE:
        _key = _v.upper()
        _entry = _m04_kobo_lookup.get(_key)
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m04_kobo_ordered.append(_collapsed)
        else:
            _m04_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M04']['variables'] = _m04_kobo_ordered

# ── M04 Heatmap: filter to authoritative vars, remove (multi), reorder ──
if 'M04' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m04_hm_buckets = {v.lower(): [] for v in _M04_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M04']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m04_hm_buckets:
            _m04_hm_buckets[_stata].append(_r)
            continue
        for _auth in _M04_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m04_hm_buckets[_auth.lower()].append(_r)
                break
    _m04_hm_ordered = []
    for _v in _M04_AUTHORITATIVE:
        _m04_hm_ordered.extend(_m04_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M04'] = _m04_hm_ordered

    # ── Inject synthetic heatmap rows for authoritative vars missing from cache ──
    # a24-a27 exist in Kobo for ALL rounds (R1–R5) but are not yet in the pooled
    # .dta (user recently un-dropped them from HF do-file, Stata re-run pending).
    # Once the .dta includes them, build_dq.py will generate real rows and these
    # synthetics will be skipped automatically (the "if missing" check below).
    # NOTE: a24-a27 are NOT R4-R5 only — they appear in all rounds. Their role
    # changes (follow-up in R1-R3, primary routing in R4-R5) but collection is
    # continuous. Not gated (M04 dataset = selected roster members already).
    _M04_SYNTH = {
        'a24':     {'display': 'A24 (a24)',  '1': 0, '2': 0, '3': 0, '4': 0, '5': 0},
        'a27':     {'display': 'A27 (a27)',  '1': 0, '2': 0, '3': 0, '4': 0, '5': 0},
        'a25':     {'display': 'A25 (a25)',  '1': 0, '2': 0, '3': 0, '4': 0, '5': 0},
        'a26':     {'display': 'A26 (a26)',  '1': 0, '2': 0, '3': 0, '4': 0, '5': 0},
    }
    # Build final ordered list, injecting synthetics only for vars missing from cache
    _m04_final = []
    _m04_by_auth = {}  # auth_var → list of existing rows
    for _r in dq_raw['heatmap_data']['M04']:
        _var = _r.get('var', '')
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        for _auth in _M04_AUTHORITATIVE:
            if _stata == _auth.lower() or _stata.startswith(_auth.lower() + '_'):
                _m04_by_auth.setdefault(_auth.lower(), []).append(_r)
                break
    for _v in _M04_AUTHORITATIVE:
        _vl = _v.lower()
        if _vl in _m04_by_auth:
            _m04_final.extend(_m04_by_auth[_vl])
        elif _vl in _M04_SYNTH:
            _s = _M04_SYNTH[_vl]
            _m04_final.append({
                'var': _s['display'], 'rag': 'green',
                '1': _s['1'], '2': _s['2'], '3': _s['3'], '4': _s['4'], '5': _s['5'],
                'note': 'Synthetic — var not yet in .dta cache. Will auto-replace when Stata re-runs.',
            })
    dq_raw['heatmap_data']['M04'] = _m04_final

# ── M05 Income standardisation ──────────────────────────────────────────
# Authoritative list: 15 vars (Kobo order: IA2 → IA3_A-F → IA5 → IA6_A-F → IA7)
_M05_AUTHORITATIVE = [
    'ia2',
    'ia3_a', 'ia3_b', 'ia3_c', 'ia3_d', 'ia3_e', 'ia3_f',
    'ia5',
    'ia6_a', 'ia6_b', 'ia6_c', 'ia6_d', 'ia6_e', 'ia6_f',
    'ia7',
]
_M05_AUTH_SET = {v.upper() for v in _M05_AUTHORITATIVE}

if 'M05' in module_tables:
    import copy as _copy
    # Drop non-auth vars (IC*, ID*, aggregate IA3/IA6)
    module_tables['M05'] = [
        r for r in module_tables['M05'] if r['variable'].upper() in _M05_AUTH_SET
    ]
    # Inject missing sub-vars (ia3_a-f, ia6_a-f may need injection)
    _m05_existing = {r['variable'].upper() for r in module_tables['M05']}
    _M05_INJECT = {
        'ia3_a': ('CASH EARNINGS – REGULAR (MONTHLY)',   'Integer'),
        'ia3_b': ('IN-KIND EARNINGS – REGULAR (MONTHLY)','Integer'),
        'ia3_c': ('CASH EARNINGS – REGULAR (PIECE)',     'Integer'),
        'ia3_d': ('IN-KIND EARNINGS – REGULAR (PIECE)',  'Integer'),
        'ia3_e': ('CASH EARNINGS – REGULAR (TIPS)',      'Integer'),
        'ia3_f': ('IN-KIND EARNINGS – REGULAR (OTHER)',  'Integer'),
        'ia6_a': ('CASH EARNINGS – SEASONAL (MONTHLY)',  'Integer'),
        'ia6_b': ('IN-KIND EARNINGS – SEASONAL (MONTHLY)','Integer'),
        'ia6_c': ('CASH EARNINGS – SEASONAL (PIECE)',    'Integer'),
        'ia6_d': ('IN-KIND EARNINGS – SEASONAL (PIECE)', 'Integer'),
        'ia6_e': ('CASH EARNINGS – SEASONAL (TIPS)',     'Integer'),
        'ia6_f': ('IN-KIND EARNINGS – SEASONAL (OTHER)', 'Integer'),
    }
    _template = module_tables['M05'][0] if module_tables['M05'] else {}
    for _var, (_title, _type) in _M05_INJECT.items():
        if _var.upper() not in _m05_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M05'].append(_synth)
    # Reorder
    _m05_by_var = {}
    for _row in module_tables['M05']:
        _m05_by_var[_row['variable'].upper()] = _row
    _ordered_m05 = []
    for _v in _M05_AUTHORITATIVE:
        _r = _m05_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m05.append(_r)
    module_tables['M05'] = _ordered_m05

# M05 Kobo: drop calculations/notes, reorder
if 'M05' in kobo_raw and 'variables' in kobo_raw.get('M05', {}):
    _m05_kobo_lookup = {}
    for _v in kobo_raw['M05']['variables']:
        _key = _v['name'].upper().rstrip('_')
        if _key not in _m05_kobo_lookup:
            _m05_kobo_lookup[_key] = _v
    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}
    _m05_kobo_ordered = []
    for _v in _M05_AUTHORITATIVE:
        _key = _v.upper()
        _entry = _m05_kobo_lookup.get(_key)
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m05_kobo_ordered.append(_collapsed)
        else:
            _m05_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M05']['variables'] = _m05_kobo_ordered

# M05 Heatmap: filter + reorder
if 'M05' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m05_hm_buckets = {v.lower(): [] for v in _M05_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M05']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m05_hm_buckets:
            _m05_hm_buckets[_stata].append(_r)
            continue
        for _auth in _M05_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m05_hm_buckets[_auth.lower()].append(_r)
                break
    _m05_hm_ordered = []
    for _v in _M05_AUTHORITATIVE:
        _m05_hm_ordered.extend(_m05_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M05'] = _m05_hm_ordered

# ── M06 Finance standardisation ────────────────────────────────────────
# Authoritative list: 17 vars (R5 Kobo order: F17→F1→F18→F2→…→F16)
# F17/F18 are R5-only bank/mobile-money gates added before F1/F2.
_M06_AUTHORITATIVE = [
    'f17', 'f1', 'f18', 'f2', 'f3', 'f6', 'f7',
    'f8', 'f8_oth', 'f9', 'f9_oth', 'f10',
    'f13_a', 'f13_b', 'f14', 'f15', 'f16',
]
_M06_AUTH_SET = {v.upper() for v in _M06_AUTHORITATIVE}

if 'M06' in module_tables:
    import copy as _copy
    # Drop LOAN (not in Kobo) and F13 (split into F13_A/F13_B)
    module_tables['M06'] = [
        r for r in module_tables['M06'] if r['variable'].upper() in _M06_AUTH_SET
    ]
    _m06_existing = {r['variable'].upper() for r in module_tables['M06']}
    _M06_INJECT = {
        'f8_oth':  ('OTHER LOAN PURPOSE (SPECIFY)',         'Text'),
        'f9_oth':  ('OTHER LENDING INSTITUTION (SPECIFY)',  'Text'),
        'f13_a':   ('RECEIVED REMITTANCE – DOMESTIC',       'Categorical'),
        'f13_b':   ('RECEIVED REMITTANCE – INTERNATIONAL',  'Categorical'),
        'f17':     ('HAS BANK ACCOUNT (R5 ONLY)',           'Categorical'),
        'f18':     ('HAS MOBILE MONEY ACCOUNT (R5 ONLY)',   'Categorical'),
    }
    _template = module_tables['M06'][0] if module_tables['M06'] else {}
    for _var, (_title, _type) in _M06_INJECT.items():
        if _var.upper() not in _m06_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M06'].append(_synth)
    _m06_by_var = {}
    for _row in module_tables['M06']:
        _m06_by_var[_row['variable'].upper()] = _row
    _ordered_m06 = []
    for _v in _M06_AUTHORITATIVE:
        _r = _m06_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m06.append(_r)
    module_tables['M06'] = _ordered_m06

if 'M06' in kobo_raw and 'variables' in kobo_raw.get('M06', {}):
    _m06_kobo_lookup = {}
    for _v in kobo_raw['M06']['variables']:
        _key = _v['name'].upper().rstrip('_')
        if _key not in _m06_kobo_lookup:
            _m06_kobo_lookup[_key] = _v
    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}
    _m06_kobo_ordered = []
    for _v in _M06_AUTHORITATIVE:
        _key = _v.upper()
        _entry = _m06_kobo_lookup.get(_key)
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m06_kobo_ordered.append(_collapsed)
        else:
            _m06_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M06']['variables'] = _m06_kobo_ordered

if 'M06' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m06_hm_buckets = {v.lower(): [] for v in _M06_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M06']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m06_hm_buckets:
            _m06_hm_buckets[_stata].append(_r)
            continue
        for _auth in _M06_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m06_hm_buckets[_auth.lower()].append(_r)
                break
    _m06_hm_ordered = []
    for _v in _M06_AUTHORITATIVE:
        _m06_hm_ordered.extend(_m06_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M06'] = _m06_hm_ordered

# ── M07 Health standardisation ─────────────────────────────────────────
# Authoritative list: 24 vars (R5 Kobo order)
# R1-R4: only H2/H2A/H3/H3_oth. Full expansion in R5.
_M07_AUTHORITATIVE = [
    'h2', 'h2a', 'h3', 'h3_oth',
    'h4', 'h4_oth', 'h7', 'h8', 'h8_amt',
    'h9a', 'h9b', 'h9c',
    'h10', 'h11a', 'h11b', 'h11b_oth',
    'h12', 'h13', 'h13_oth', 'h14', 'h15', 'h16', 'h16_oth',
    'h17',
]
_M07_AUTH_SET = {v.upper() for v in _M07_AUTHORITATIVE}

if 'M07' in module_tables:
    import copy as _copy
    module_tables['M07'] = [
        r for r in module_tables['M07'] if r['variable'].upper() in _M07_AUTH_SET
    ]
    _m07_existing = {r['variable'].upper() for r in module_tables['M07']}
    _M07_INJECT = {
        'h3_oth':  ('OTHER REASON FOR NOT GETTING CARE',    'Text'),
        'h4_oth':  ('OTHER HEALTHCARE FACILITY (SPECIFY)',   'Text'),
        'h8_amt':  ('OUT-OF-POCKET AMOUNT',                  'Integer'),
        'h9a':     ('SERVICE: CONSULTATION',                 'Categorical'),
        'h9b':     ('SERVICE: LABORATORY/DIAGNOSTIC',        'Categorical'),
        'h9c':     ('SERVICE: MEDICINE/PRESCRIPTION',        'Categorical'),
        'h11b':    ('WHO PAID FOR SERVICE',                  'Categorical'),
        'h11b_oth':('OTHER PAYER (SPECIFY)',                 'Text'),
        'h13_oth': ('OTHER HOSPITAL FACILITY (SPECIFY)',     'Text'),
        'h16':     ('PAYMENT SOURCE FOR HOSPITALISATION',    'Categorical'),
        'h16_oth': ('OTHER PAYMENT SOURCE (SPECIFY)',        'Text'),
    }
    _template = module_tables['M07'][0] if module_tables['M07'] else {}
    for _var, (_title, _type) in _M07_INJECT.items():
        if _var.upper() not in _m07_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _title
            _synth['question_type'] = _type
            _set_round_presence(_synth, _var)
            module_tables['M07'].append(_synth)
    _m07_by_var = {}
    for _row in module_tables['M07']:
        _m07_by_var[_row['variable'].upper()] = _row
    _ordered_m07 = []
    for _v in _M07_AUTHORITATIVE:
        _r = _m07_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m07.append(_r)
    module_tables['M07'] = _ordered_m07

if 'M07' in kobo_raw and 'variables' in kobo_raw.get('M07', {}):
    _m07_kobo_lookup = {}
    for _v in kobo_raw['M07']['variables']:
        _key = _v['name'].upper().rstrip('_')
        if _key not in _m07_kobo_lookup:
            _m07_kobo_lookup[_key] = _v
    _KOBO_M07_ALIASES = {
        'H11B': ['H11B', 'H11B_'],
        'H16': ['H16', 'H16_'],
    }
    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}
    _m07_kobo_ordered = []
    for _v in _M07_AUTHORITATIVE:
        _key = _v.upper()
        _entry = _m07_kobo_lookup.get(_key)
        if not _entry and _key in _KOBO_M07_ALIASES:
            for _alias in _KOBO_M07_ALIASES[_key]:
                _entry = _m07_kobo_lookup.get(_alias)
                if _entry:
                    break
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m07_kobo_ordered.append(_collapsed)
        else:
            _m07_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M07']['variables'] = _m07_kobo_ordered

if 'M07' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m07_hm_buckets = {v.lower(): [] for v in _M07_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M07']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m07_hm_buckets:
            _m07_hm_buckets[_stata].append(_r)
            continue
        for _auth in _M07_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m07_hm_buckets[_auth.lower()].append(_r)
                break
    _m07_hm_ordered = []
    for _v in _M07_AUTHORITATIVE:
        _m07_hm_ordered.extend(_m07_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M07'] = _m07_hm_ordered

# ── M08 Food/FIES standardisation ──────────────────────────────────────
# Authoritative list: 5 vars (Kobo order: F08_A through F08_E)
_M08_AUTHORITATIVE = ['f08_a', 'f08_b', 'f08_c', 'f08_d', 'f08_e']
_M08_AUTH_SET = {v.upper() for v in _M08_AUTHORITATIVE}

if 'M08' in module_tables:
    import copy as _copy
    # Tracker has single F08 — expand to individual FIES items
    module_tables['M08'] = [
        r for r in module_tables['M08'] if r['variable'].upper() in _M08_AUTH_SET
    ]
    _m08_existing = {r['variable'].upper() for r in module_tables['M08']}
    _FIES_LABELS = {
        'f08_a': 'WORRIED ABOUT NOT HAVING ENOUGH FOOD',
        'f08_b': 'UNABLE TO EAT HEALTHY/NUTRITIOUS FOOD',
        'f08_c': 'ATE ONLY A FEW KINDS OF FOOD',
        'f08_d': 'HAD TO SKIP A MEAL',
        'f08_e': 'DID NOT EAT FOR A WHOLE DAY',
    }
    _template_row = module_tables['M08'][0] if module_tables['M08'] else {}
    # If template is from the F08 aggregate, use it as base
    if not _template_row:
        _template_row = {'variable': '', 'question_title': '', 'question_type': 'Categorical'}
    for _var in _M08_AUTHORITATIVE:
        if _var.upper() not in _m08_existing:
            _synth = {k: '' for k in _template_row} if _template_row else {}
            _synth['variable'] = _var
            _synth['question_title'] = _FIES_LABELS.get(_var, _var.upper())
            _synth['question_type'] = 'Categorical'
            _set_round_presence(_synth, _var)
            module_tables['M08'].append(_synth)
    _m08_by_var = {}
    for _row in module_tables['M08']:
        _m08_by_var[_row['variable'].upper()] = _row
    _ordered_m08 = []
    for _v in _M08_AUTHORITATIVE:
        _r = _m08_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m08.append(_r)
    module_tables['M08'] = _ordered_m08

if 'M08' in kobo_raw and 'variables' in kobo_raw.get('M08', {}):
    _m08_kobo_lookup = {}
    for _v in kobo_raw['M08']['variables']:
        _key = _v['name'].upper()
        if _key not in _m08_kobo_lookup:
            _m08_kobo_lookup[_key] = _v
    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}
    _m08_kobo_ordered = []
    for _v in _M08_AUTHORITATIVE:
        _entry = _m08_kobo_lookup.get(_v.upper())
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m08_kobo_ordered.append(_collapsed)
        else:
            _m08_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M08']['variables'] = _m08_kobo_ordered

if 'M08' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m08_hm_buckets = {v.lower(): [] for v in _M08_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M08']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m08_hm_buckets:
            _m08_hm_buckets[_stata].append(_r)
    _m08_hm_ordered = []
    for _v in _M08_AUTHORITATIVE:
        _m08_hm_ordered.extend(_m08_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M08'] = _m08_hm_ordered

# ── M09 Views standardisation ──────────────────────────────────────────
# Authoritative list: 15 vars (R1 Kobo order)
# V1/V5: asked all rounds. V9_A-M: randomised agreement statements.
# V9_B: dropped R4+. V11/V12: R1 only.
_M09_AUTHORITATIVE = [
    'v1', 'v5',
    'v9_a', 'v9_b', 'v9_c', 'v9_e', 'v9_f', 'v9_g',
    'v9_i', 'v9_j', 'v9_k', 'v9_l', 'v9_m',
    'v11', 'v12',
]
_M09_AUTH_SET = {v.upper() for v in _M09_AUTHORITATIVE}

if 'M09' in module_tables:
    import copy as _copy
    # Tracker has V1, V5, V11, V12, V9 — need to expand V9 to sub-items
    module_tables['M09'] = [
        r for r in module_tables['M09'] if r['variable'].upper() in _M09_AUTH_SET
    ]
    _m09_existing = {r['variable'].upper() for r in module_tables['M09']}
    _V9_LABELS = {
        'v9_a': 'AGREE: LIFE IN GENERAL IS SATISFYING',
        'v9_b': 'AGREE: AI WILL TAKE AWAY JOBS',
        'v9_c': 'AGREE: CLIMATE CHANGE WILL AFFECT LIVELIHOOD',
        'v9_e': 'AGREE: GENDER EQUALITY IN EDUCATION',
        'v9_f': 'AGREE: CHILDREN SHOULD GO TO SCHOOL',
        'v9_g': 'AGREE: GOVERNMENT HANDLES ECONOMY WELL',
        'v9_i': 'AGREE: VACCINES ARE SAFE AND EFFECTIVE',
        'v9_j': 'AGREE: FILIPINOS WORK HARD',
        'v9_k': 'AGREE: MY COMMUNITY IS SAFE',
        'v9_l': 'AGREE: EASY TO ACCESS HEALTH SERVICES',
        'v9_m': 'AGREE: DIGITAL TECHNOLOGY IMPROVES LIFE',
    }
    _template = module_tables['M09'][0] if module_tables['M09'] else {}
    for _var in _M09_AUTHORITATIVE:
        if _var.upper() not in _m09_existing:
            _synth = {k: '' for k in _template}
            _synth['variable'] = _var
            _synth['question_title'] = _V9_LABELS.get(_var, _var.upper())
            _synth['question_type'] = 'Categorical'
            _set_round_presence(_synth, _var)
            module_tables['M09'].append(_synth)
    _m09_by_var = {}
    for _row in module_tables['M09']:
        _m09_by_var[_row['variable'].upper()] = _row
    _ordered_m09 = []
    for _v in _M09_AUTHORITATIVE:
        _r = _m09_by_var.get(_v.upper())
        if _r:
            _r['variable'] = _v
            _ordered_m09.append(_r)
    module_tables['M09'] = _ordered_m09

if 'M09' in kobo_raw and 'variables' in kobo_raw.get('M09', {}):
    _m09_kobo_lookup = {}
    for _v in kobo_raw['M09']['variables']:
        _key = _v['name'].upper()
        if _key not in _m09_kobo_lookup:
            _m09_kobo_lookup[_key] = _v
    # Drop hidden_select_v9 (form control, not data)
    _M09_KOBO_DROP = {'HIDDEN_SELECT_V9'}
    _synth_rules = {str(r): {'relevant': None, 'required': False,
                              'constraint': None, 'constraint_message': None}
                    for r in range(1, 9)}
    _m09_kobo_ordered = []
    for _v in _M09_AUTHORITATIVE:
        _entry = _m09_kobo_lookup.get(_v.upper())
        if _entry:
            _collapsed = dict(_entry)
            _collapsed['name'] = _v
            _m09_kobo_ordered.append(_collapsed)
        else:
            _m09_kobo_ordered.append({
                'name': _v, 'type': 'metadata', 'label': _v.upper(),
                'relevant': '', 'constraint': '', 'choice_filter': '',
                'rounds': {str(r): True for r in range(1, 9)},
                'rules_by_round': _synth_rules,
            })
    kobo_raw['M09']['variables'] = _m09_kobo_ordered

if 'M09' in dq_raw.get('heatmap_data', {}):
    import re as _re
    _m09_hm_buckets = {v.lower(): [] for v in _M09_AUTHORITATIVE}
    for _r in dq_raw['heatmap_data']['M09']:
        _var = _r.get('var', '')
        if '(multi)' in _var:
            continue
        _m = _re.search(r'\(([^)]+)\)', _var)
        _stata = (_m.group(1) if _m else _var).lower()
        if _stata in _m09_hm_buckets:
            _m09_hm_buckets[_stata].append(_r)
            continue
        for _auth in _M09_AUTHORITATIVE:
            if _stata.startswith(_auth.lower() + '_'):
                _m09_hm_buckets[_auth.lower()].append(_r)
                break
    _m09_hm_ordered = []
    for _v in _M09_AUTHORITATIVE:
        _m09_hm_ordered.extend(_m09_hm_buckets.get(_v.lower(), []))
    dq_raw['heatmap_data']['M09'] = _m09_hm_ordered

DQ   = json.dumps(dq_raw,       separators=(',',':'))
MT   = json.dumps(module_tables, separators=(',',':'))
AQ   = json.dumps(all_qs,        separators=(',',':'))
PAN  = json.dumps(panel_raw,     separators=(',',':'))
IVIEW= json.dumps(iview_raw,     separators=(',',':'))
KOBO = json.dumps(kobo_raw,      separators=(',',':'))
ISSUES = json.dumps(issues_raw,  separators=(',',':'))
ISUM   = json.dumps(isum_raw,     separators=(',',':'))

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>L2PHL Data Quality Dashboard v3</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Helvetica,sans-serif;background:#f0f2f5;color:#222;display:flex;min-height:100vh}

/* ── SIDEBAR ── */
#sidebar{width:235px;min-width:235px;background:#1a2332;color:#cdd5e0;display:flex;flex-direction:column;position:fixed;top:0;left:0;height:100vh;overflow-y:auto;z-index:100;scrollbar-width:thin;scrollbar-color:#2d3f55 #1a2332}
#sidebar .logo{padding:14px 16px;border-bottom:1px solid #2d3f55}
#sidebar .logo strong{display:block;font-size:14px;color:#fff;margin-bottom:1px}
#sidebar .logo span{color:#7d9ab8;font-size:10.5px}
#sidebar .nav-section{padding:10px 16px 3px;font-size:9.5px;text-transform:uppercase;letter-spacing:.08em;color:#506070;font-weight:600}
#sidebar a{display:flex;align-items:center;gap:7px;padding:7px 14px;color:#cdd5e0;text-decoration:none;font-size:12.5px;border-left:3px solid transparent;transition:.12s}
#sidebar a:hover{background:#243347;color:#fff}
#sidebar a.active{background:#1d3150;color:#4db8ff;border-left-color:#4db8ff}
.dot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.nav-count{margin-left:auto;background:#2d3f55;color:#aab8c8;font-size:10px;padding:1px 6px;border-radius:10px}
.nav-count.red{background:#c0392b;color:#fff}
.nav-count.yellow{background:#e67e22;color:#fff}

/* ── MAIN ── */
#main{margin-left:235px;flex:1;padding:22px;min-width:0;overflow-x:hidden}
.page{display:none}.page.active{display:block}
h1{font-size:21px;font-weight:700;margin-bottom:4px}
.subtitle{color:#666;font-size:12.5px;margin-bottom:18px}

/* ── CARDS ── */
.card{background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:18px;margin-bottom:18px;overflow-x:auto}
.card h2{font-size:14.5px;font-weight:600;margin-bottom:12px;color:#333;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.badge{font-size:10.5px;padding:2px 7px;border-radius:10px;font-weight:600}
.badge-red{background:#fde;color:#c0392b}.badge-yellow{background:#fff3cd;color:#856404}
.badge-green{background:#d4edda;color:#155724}.badge-blue{background:#cce5ff;color:#004085}
.badge-grey{background:#f0f0f0;color:#555}.badge-purple{background:#f0e6ff;color:#6c3483}

/* ── STATS ROW ── */
.stats-row{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px}
.stat-box{background:#fff;border-radius:8px;padding:10px 16px;min-width:110px;box-shadow:0 1px 3px rgba(0,0,0,.07);border:1px solid #e8ebef;text-align:center}
.stat-box .num{font-size:26px;font-weight:700;line-height:1.1}
.stat-box .lbl{font-size:10.5px;color:#666;margin-top:2px}
.stat-box.red .num{color:#e74c3c}.stat-box.yellow .num{color:#e67e22}
.stat-box.green .num{color:#27ae60}.stat-box.blue .num{color:#2980b9}
.stat-box.purple .num{color:#8e44ad}

/* ── NOTE BOXES ── */
.note-box{padding:9px 13px;border-radius:6px;font-size:12px;margin-bottom:12px;line-height:1.5}
.note-info{background:#e8f4fd;border-left:4px solid #3498db;color:#1a5276}
.note-warn{background:#fff8e1;border-left:4px solid #f39c12;color:#7d4e00}
.note-ok{background:#eafaf1;border-left:4px solid #2ecc71;color:#1a5c32}
.note-flag{background:#fff0f0;border-left:4px solid #e74c3c;color:#7b0000}
.note-purple{background:#f5eef8;border-left:4px solid #8e44ad;color:#4a235a}

/* ── VIOL ROWS ── */
.viol-row{display:flex;align-items:flex-start;gap:10px;padding:10px 12px;border-radius:7px;margin-bottom:8px}
.viol-row.high{background:#fff5f5;border-left:4px solid #e74c3c}
.viol-row.medium{background:#fffdf0;border-left:4px solid #f39c12}
.viol-row.ok{background:#f5fff8;border-left:4px solid #2ecc71}
.viol-row.clean{background:#f5fff8;border-left:4px solid #2ecc71}
.viol-icon{font-size:18px;flex-shrink:0;margin-top:1px}
.viol-text strong{display:block;font-size:12.5px;margin-bottom:2px}
.viol-text .viol-path{font-size:11px;color:#777;font-family:monospace;margin:2px 0}
.viol-text .viol-note{font-size:11.5px;color:#555;margin-top:3px;line-height:1.4}
.viol-pills{display:flex;gap:4px;flex-wrap:wrap;margin-top:5px}
.vpill{border-radius:10px;padding:2px 8px;font-size:10.5px;font-weight:600;border:1px solid}
.vpill-red{background:#fde;border-color:#f5c6cb;color:#c0392b}
.vpill-yellow{background:#fff3cd;border-color:#ffc107;color:#856404}
.vpill-green{background:#d4edda;border-color:#c3e6cb;color:#155724}
.vpill-grey{background:#f0f0f0;border-color:#ddd;color:#666}

/* ── CHART ── */
.chart-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:14px}
.chart-box{background:#fff;border-radius:8px;padding:12px;box-shadow:0 1px 3px rgba(0,0,0,.07);border:1px solid #e8ebef;overflow:hidden}
.chart-box canvas{max-width:100%}
.chart-box.flagged{border-left:4px solid #e74c3c}.chart-box.warn{border-left:4px solid #f39c12}
.ch-title{font-size:12.5px;font-weight:600;margin-bottom:2px;color:#222}
.ch-sub{font-size:11px;color:#888;margin-bottom:7px;line-height:1.35}
.ch-note{font-size:10.5px;color:#555;background:#f7f9fc;padding:4px 7px;border-radius:4px;margin-top:5px;line-height:1.4}

/* ── HEATMAP ── */
.heatmap-wrap{overflow-x:auto}
.heatmap{border-collapse:collapse;font-size:11.5px;width:100%}
.heatmap th,.heatmap td{padding:5px 9px;text-align:center;border-bottom:1px solid #f2f2f2}
.heatmap th{background:#f7f9fc;font-weight:600;color:#555;font-size:10.5px;text-transform:uppercase;position:sticky;top:0}
.heatmap td.vn{text-align:left;font-family:monospace;font-size:11px;white-space:nowrap;max-width:260px;overflow:hidden;text-overflow:ellipsis}
.hm-cell{border-radius:3px;font-weight:500;min-width:48px}
.gate-tag{font-family:sans-serif;font-size:9px;color:#7c6f64;background:#f5f0e8;border-radius:3px;padding:1px 5px;margin-left:4px;font-weight:500;letter-spacing:.2px}
.skip-badge{display:inline-block;font-size:8px;color:#fff;background:#e67e22;border-radius:3px;padding:0 4px;margin-left:2px;font-weight:600;vertical-align:middle;line-height:14px;cursor:help}

/* ── MODULE TABS ── */
.mod-tabs{display:flex;gap:3px;flex-wrap:wrap;margin-bottom:12px}
.mtab{padding:5px 11px;border-radius:5px;background:#f0f2f5;border:1px solid #dde;cursor:pointer;font-size:11.5px;font-weight:500;color:#444;transition:.12s}
.mtab.active,.mtab:hover{background:#1d3150;color:#fff;border-color:#1d3150}

/* ── MODULE GRID ── */
.mod-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px}
.mod-card{border-radius:8px;padding:12px;border:1px solid #e5e8ec;position:relative;overflow:hidden;cursor:pointer;transition:.12s}
.mod-card:hover{box-shadow:0 2px 8px rgba(0,0,0,.12)}
.mod-card.red{border-left:4px solid #e74c3c;background:#fff9f9}
.mod-card.yellow{border-left:4px solid #f39c12;background:#fffef5}
.mod-card.green{border-left:4px solid #2ecc71;background:#f5fff8}
.mod-card .mname{font-weight:700;font-size:12.5px;margin-bottom:5px;color:#222}
.mod-card .mstat{font-size:11.5px;color:#555;margin:1px 0;line-height:1.4}
.mod-card .mstat.warn{color:#c0392b;font-weight:600}
.rag-chip{position:absolute;top:8px;right:8px;font-size:9.5px;padding:2px 6px;border-radius:10px;font-weight:700}
.istrip{display:flex;gap:3px;margin:4px 0}
.idot{width:13px;height:13px;border-radius:3px;font-size:8px;text-align:center;line-height:13px;color:#fff}
.idot.red{background:#e74c3c}.idot.yellow{background:#f1c40f;color:#5b4a00}.idot.green{background:#cfe8d6;color:#cfe8d6}.idot.closed{background:#d6d6d6;color:#888}
.vbadge{display:inline-block;font-size:9px;font-weight:700;border-radius:3px;padding:1px 5px;color:#fff}
.vbadge.A1,.vbadge.A2{background:#c0392b}.vbadge.B{background:#e67e22}.vbadge.C{background:#8e44ad}.vbadge.D{background:#95a5a6}.vbadge.REVIEW{background:#34495e}
.schip{display:inline-block;font-size:9px;border-radius:8px;padding:1px 6px;background:#ecf0f1;color:#34495e;margin-left:4px}
.evbox{background:#f7f9fb;border-left:3px solid #4db8ff;padding:8px 11px;margin-top:5px;font-size:11px;font-family:monospace;line-height:1.5;white-space:pre-wrap}

/* ── LEGEND ── */
.legend{display:flex;gap:12px;flex-wrap:wrap;font-size:11.5px;margin-bottom:10px}
.legend-item{display:flex;align-items:center;gap:5px}
.legend-dot{width:11px;height:11px;border-radius:2px}

/* ═════════════════════════════════════════
   PER-QUESTION TRACKER STYLES
   ═════════════════════════════════════════ */

/* Round presence matrix */
.qtrack-table{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:6px}
.qtrack-table th{background:#f0f4fa;padding:5px 8px;font-size:10.5px;font-weight:600;color:#555;text-transform:uppercase;border-bottom:2px solid #dde;text-align:center;position:sticky;top:0;z-index:2}
.qtrack-table th.left{text-align:left}
.qtrack-table td{padding:6px 8px;border-bottom:1px solid #eef;vertical-align:middle}
.qtrack-table td.var-name{font-family:monospace;font-weight:700;font-size:12px;color:#1a2332;white-space:nowrap}
.qtrack-table td.q-title{font-size:11.5px;color:#333;max-width:220px}
.qtrack-table td.q-text{font-size:11px;color:#666;max-width:280px;font-style:italic}
.qtrack-table tr:hover td{background:#f5f8ff}

/* Presence pill */
.pres{display:inline-block;width:28px;height:22px;border-radius:4px;font-size:10px;font-weight:700;line-height:22px;text-align:center;cursor:default}
.pres.yes{background:#d4edda;color:#155724}
.pres.no{background:#f8d7da;color:#721c24}
.pres.na{background:#f0f0f0;color:#999}

/* Change tags */
.chg-tag{display:inline-flex;align-items:center;gap:3px;border-radius:10px;padding:2px 7px;font-size:10px;font-weight:600;margin:1px;white-space:nowrap}
.chg-new{background:#d4edda;color:#155724;border:1px solid #c3e6cb}
.chg-drop{background:#f8d7da;color:#721c24;border:1px solid #f5c6cb}
.chg-title{background:#fff3cd;color:#856404;border:1px solid #ffc107}
.chg-skip{background:#cce5ff;color:#004085;border:1px solid #bee5eb}
.chg-opts{background:#e8daef;color:#4a235a;border:1px solid #d2b4de}
.chg-code{background:#f5eef8;color:#6c3483;border:1px solid #d7bde2}
.chg-check{background:#e8f4fd;color:#1a5276;border:1px solid #aed6f1}

/* DQ issue inline */
.dq-inline{display:inline-flex;align-items:center;gap:3px;border-radius:4px;padding:2px 7px;font-size:10px;font-weight:600;margin:1px}
.dq-skip{background:#fff0f0;color:#c0392b;border:1px solid #f5c6cb}
.dq-mand{background:#fff8e1;color:#7d4e00;border:1px solid #ffc107}
.dq-oor{background:#f5eef8;color:#6c3483;border:1px solid #d7bde2}

/* Round filter pills */
.round-filters{display:flex;gap:5px;flex-wrap:wrap;margin-bottom:12px;align-items:center}
.rfil{padding:4px 11px;border-radius:14px;background:#f0f2f5;border:1px solid #dde;cursor:pointer;font-size:11.5px;font-weight:500;color:#444;transition:.12s}
.rfil.active{background:#2980b9;color:#fff;border-color:#2980b9}
.rfil-label{font-size:11px;color:#888;margin-right:3px}

/* Expandable detail panel */
.q-detail{display:none;background:#f7f9ff;border-radius:6px;padding:10px 14px;margin:-4px 0 6px;border-left:3px solid #3498db;font-size:11.5px;line-height:1.55}
.q-detail.open{display:block}
.q-detail dl{display:grid;grid-template-columns:auto 1fr;gap:2px 10px}
.q-detail dt{font-weight:600;color:#555;font-size:10.5px;text-transform:uppercase;white-space:nowrap}
.q-detail dd{color:#333;margin:0}
.q-detail .skip-rules{margin-top:6px}
.q-detail .skip-rule-row{display:flex;gap:6px;margin:2px 0;align-items:flex-start}
.q-detail .skip-round{font-size:9.5px;font-weight:700;background:#1d3150;color:#fff;border-radius:10px;padding:1px 6px;min-width:28px;text-align:center;flex-shrink:0;margin-top:1px}
.q-detail .skip-text{font-size:11px;color:#333;font-family:monospace}

/* Module page header */
.mod-page-header{display:flex;align-items:center;gap:12px;margin-bottom:14px;flex-wrap:wrap}
.mod-page-header h1{margin-bottom:0}
.mod-round-summary{display:flex;gap:6px;flex-wrap:wrap}
.mrs-pill{display:flex;align-items:center;gap:4px;border-radius:12px;padding:3px 10px;font-size:11px;font-weight:600;border:1px solid}
.mrs-ok{background:#d4edda;color:#155724;border-color:#c3e6cb}
.mrs-warn{background:#fff3cd;color:#856404;border-color:#ffc107}
.mrs-flag{background:#f8d7da;color:#721c24;border-color:#f5c6cb}
.mrs-na{background:#f0f0f0;color:#999;border-color:#ddd}

/* Question type badge */
.qtype{font-size:9.5px;padding:1px 6px;border-radius:8px;font-weight:600;background:#e8f4fd;color:#004085;border:1px solid #bee5eb}

/* DQ count badge on question row */
.dq-cnt{display:inline-flex;align-items:center;justify-content:center;width:18px;height:18px;border-radius:50%;font-size:9px;font-weight:700;vertical-align:middle;margin-left:3px}
.dq-cnt.red{background:#e74c3c;color:#fff}
.dq-cnt.yellow{background:#f39c12;color:#fff}

/* Sticky round header in question table */
.sticky-round{position:sticky;top:0;z-index:3}

/* Summary bar at top of module page */
.mod-dq-bar{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}
.mdb-item{flex:1;min-width:100px;background:#fff;border-radius:7px;padding:8px 12px;border:1px solid #e8ebef;box-shadow:0 1px 3px rgba(0,0,0,.05)}
.mdb-item .mdb-num{font-size:22px;font-weight:700;line-height:1}
.mdb-item .mdb-lbl{font-size:10px;color:#666;margin-top:2px}
.mdb-item.red .mdb-num{color:#e74c3c}
.mdb-item.yellow .mdb-num{color:#e67e22}
.mdb-item.green .mdb-num{color:#27ae60}
.mdb-item.blue .mdb-num{color:#2980b9}
.mdb-item.purple .mdb-num{color:#8e44ad}

/* Change summary header per round transition */
.round-change-header{background:#f0f4fa;border-radius:6px;padding:6px 12px;margin:10px 0 6px;font-size:12px;font-weight:600;color:#1a2332;display:flex;align-items:center;gap:8px}
.rch-arrow{color:#aaa;font-size:14px}

/* Search / filter */
.q-search{padding:5px 10px;border:1px solid #dde;border-radius:6px;font-size:12px;width:220px;outline:none}
.q-search:focus{border-color:#3498db}

hr{border:none;border-top:1px solid #eee;margin:14px 0}
.source-tag{font-size:9.5px;background:#e8f4fd;color:#1a5276;padding:1px 5px;border-radius:8px;font-weight:500;vertical-align:middle;margin-left:5px}
.toggle-row{cursor:pointer;user-select:none}
.toggle-row:hover td{background:#eef3ff !important}
.toggle-btn{font-size:11px;color:#3498db;border:none;background:none;cursor:pointer;padding:0;vertical-align:middle}

/* ── KOBO SKIP LOGIC MAP ── */
.kobo-table{width:100%;border-collapse:collapse;font-size:11.5px;margin-bottom:6px}
.kobo-table th{background:#f0f4fa;padding:6px 8px;font-size:10px;font-weight:600;color:#555;text-transform:uppercase;border-bottom:2px solid #dde;text-align:center;position:sticky;top:0;z-index:2}
.kobo-table th.left{text-align:left}
.kobo-table td{padding:5px 8px;border-bottom:1px solid #eef;vertical-align:top;font-size:11px}
.kobo-table tr:hover td{background:#f5f8ff}
.kobo-table td.kv-name{font-family:'SFMono-Regular',Consolas,monospace;font-weight:700;font-size:11.5px;color:#1a2332;white-space:nowrap}
.kobo-table td.kv-type{font-size:10px;color:#666;white-space:nowrap}
.kobo-rule{font-family:'SFMono-Regular',Consolas,monospace;font-size:10px;line-height:1.5;word-break:break-all;color:#333}
.kobo-rule.skip{color:#1a5276}
.kobo-rule.req{color:#155724;font-weight:600}
.kobo-rule.constraint{color:#6c3483}
.kobo-na{color:#bbb;font-style:italic;font-size:10px}
.kobo-nodata{color:#ccc;text-align:center;font-size:10px}
.kobo-changed{background:#fff8e1 !important}
.kobo-type-changed{background:#f3e5f5 !important}
.kobo-diff{background:#ffeaa7;border-radius:2px;padding:0 1px;font-weight:700}
.kobo-filter-bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px;align-items:center}
.kobo-filter-btn{padding:3px 10px;border-radius:12px;background:#f0f2f5;border:1px solid #dde;cursor:pointer;font-size:11px;font-weight:500;color:#444;transition:.12s}
.kobo-filter-btn.active{background:#2980b9;color:#fff;border-color:#2980b9}
.kobo-req-dot{display:inline-block;width:7px;height:7px;border-radius:50%;margin-right:3px}
.kobo-req-dot.yes{background:#27ae60}
.kobo-req-dot.no{background:#e0e0e0}
</style>
</head>
<body>
<nav id="sidebar">
  <div class="logo">
    <strong>L2PHL · Data Quality</strong>
    <span>Per-module · Per-question · Per-round</span>
  </div>

  <div class="nav-section">Overview</div>
  <a href="#" onclick="return showPage('overview')" id="nav-overview" class="active">
    <span class="dot" style="background:#4db8ff"></span>Dashboard Overview
  </a>

  <div class="nav-section">Modules</div>
  <div id="mod-nav-links"></div>

  <div class="nav-section">Panel & QC</div>
  <a href="#" onclick="return showPage('panel')" id="nav-panel">
    <span class="dot" style="background:#8e44ad"></span>Panel Tracking
  </a>
  <div class="nav-section">Operator Performance</div>
  <a href="#" onclick="return showPage('operators')" id="nav-operators">
    <span class="dot" style="background:#c0392b"></span>Operator QC
    <span class="nav-count red" id="nc-ops-red">—</span>
  </a>
  <div class="nav-section">Questionnaire Changes</div>
  <a href="#" onclick="return showPage('changes')" id="nav-changes">
    <span class="dot" style="background:#f1c40f"></span>All Changes by Round
  </a>
  <div class="nav-section">Issue Intelligence</div>
  <a href="#" onclick="return showPage('issues')" id="nav-issues">
    <span class="dot" style="background:#e74c3c"></span>Issues &amp; Root Cause
  </a>
</nav>

<div id="main">
<!-- ═══════ ISSUE INTELLIGENCE ═══════ -->
<div id="page-issues" class="page">
<h1>Issue Intelligence</h1>
<p class="subtitle">Every flag root-caused to a layer · A1 questionnaire · A2 field · B firm do-file · C our check · D structural</p>
<div style="margin:8px 0;font-size:12px">
  <label><input type="checkbox" id="iss-firm-only" onchange="renderIssues()"> Firm report only (A1/A2/B, open)</label>
  &nbsp;&nbsp;<label><input type="checkbox" id="iss-review-only" onchange="renderIssues()"> Review queue only</label>
</div>
<div id="issues-body"></div>
</div>
<!-- ═══════ OVERVIEW ═══════ -->
<div id="page-overview" class="page active">
<h1>Data Quality Overview</h1>
<p class="subtitle">L2PHL "Listening to the Philippines" Panel Survey · <span id="overview-round-range">Rounds 1–N</span> · Questionnaire-grounded checks</p>
<div class="note-info note-box">
  <strong>New in v3:</strong> Each module now has a deep-dive page showing every question, its presence across rounds, questionnaire changes (new/dropped/renamed/skip-logic), and data quality issues linked to the question.
  Click any module card or the module links in the sidebar to explore.
</div>
<div id="ov-stats" class="stats-row"></div>
<div class="card" style="background:#1a2332;border:1px solid #2a3a4e;margin-bottom:10px">
  <h2 style="font-size:14px;margin-bottom:8px">🚦 RAG Status Criteria</h2>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
    <div>
      <div style="font-size:12px;font-weight:600;color:#8af;margin-bottom:6px">Module-level status (cards below)</div>
      <table style="width:100%;font-size:11.5px;border-collapse:collapse">
        <tr style="border-bottom:1px solid #2a3a4e">
          <td style="padding:4px 6px">🔴 <span style="color:#e74c3c;font-weight:600">Red</span></td>
          <td style="padding:4px 6px;color:#ccc">Skip violations &gt; 100 <em>or</em> mandatory missing &gt; 100 <em>or</em> worst variable ≥ 30% missing</td>
        </tr>
        <tr style="border-bottom:1px solid #2a3a4e">
          <td style="padding:4px 6px">🟡 <span style="color:#f39c12;font-weight:600">Yellow</span></td>
          <td style="padding:4px 6px;color:#ccc">Any skip/mandatory/OOR violations <em>or</em> worst variable ≥ 10% missing</td>
        </tr>
        <tr>
          <td style="padding:4px 6px">🟢 <span style="color:#2ecc71;font-weight:600">Green</span></td>
          <td style="padding:4px 6px;color:#ccc">Zero violations and all variables &lt; 10% missing</td>
        </tr>
      </table>
    </div>
    <div>
      <div style="font-size:12px;font-weight:600;color:#8af;margin-bottom:6px">Variable-level status (heatmap cells)</div>
      <table style="width:100%;font-size:11.5px;border-collapse:collapse">
        <tr style="border-bottom:1px solid #2a3a4e">
          <td style="padding:4px 6px">🔴 <span style="color:#e74c3c;font-weight:600">Red</span></td>
          <td style="padding:4px 6px;color:#ccc">≥ 15% missing in any round</td>
        </tr>
        <tr style="border-bottom:1px solid #2a3a4e">
          <td style="padding:4px 6px">🟡 <span style="color:#f39c12;font-weight:600">Yellow</span></td>
          <td style="padding:4px 6px;color:#ccc">≥ 5% missing in any round</td>
        </tr>
        <tr>
          <td style="padding:4px 6px">🟢 <span style="color:#2ecc71;font-weight:600">Green</span></td>
          <td style="padding:4px 6px;color:#ccc">&lt; 5% missing in all rounds</td>
        </tr>
      </table>
    </div>
  </div>
  <p style="font-size:11px;color:#889;margin-top:8px;margin-bottom:0">Module max-missing uses only unconditional variables (not gated sub-items with small denominators). Skip &amp; mandatory counts are summed across all rounds.</p>
</div>
<div class="card">
  <h2>Module-Level Quality Summary <span class="badge badge-blue">Click a card to view details</span></h2>
  <div style="margin:2px 0 10px;font-size:11px;color:#555">
    Per-round strip:
    <span class="idot red" style="display:inline-block;vertical-align:middle">!</span> open firm issue
    &nbsp;<span class="idot yellow" style="display:inline-block;vertical-align:middle"></span> open
    &nbsp;<span class="idot closed" style="display:inline-block;vertical-align:middle">·</span> closed
    &nbsp;<span class="idot green" style="display:inline-block;vertical-align:middle;border:1px solid #cfe8d6"></span> clean
  </div>
  <div id="mod-grid" class="mod-grid"></div>
</div>
<div class="card">
  <h2>Sample Counts per Round</h2>
  <div style="height:260px"><canvas id="sampleChart"></canvas></div>
</div>
<div class="card">
  <h2>Questionnaire Change Summary by Round</h2>
  <div id="change-summary-bar"></div>
</div>
</div>






<div id="page-panel" class="page">
<h1>Panel Structure &amp; Household Tracking</h1>
<p class="subtitle">Participation across rounds · Attrition bias · PSU coverage vs targets · In / out households</p>

<!-- Panel Health Scorecard -->
<div id="panel-scorecard" style="display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:14px"></div>

<!-- Funnel + Retention Charts -->
<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px">
  <div class="card" style="margin-bottom:0">
    <h2 style="font-size:13px">Sample Funnel — Cumulative Attrition</h2>
    <p style="font-size:10.5px;color:#888;margin-bottom:6px">R1 baseline HHs tracked through all rounds (blue = retained from R1, orange = new entries)</p>
    <div style="position:relative;height:220px"><canvas id="panelFunnelChart"></canvas></div>
  </div>
  <div class="card" style="margin-bottom:0">
    <h2 style="font-size:13px">Retention Rate &amp; Effective Sample</h2>
    <p style="font-size:10.5px;color:#888;margin-bottom:6px">% of previous-round HHs retained. Dashed line = call interval violation rate.</p>
    <div style="position:relative;height:220px"><canvas id="panelRetentionChart"></canvas></div>
  </div>
</div>

<div id="panel-bias-banner"></div>
<div class="card" style="margin-top:10px">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    🔬 Attrition Bias Analysis <span id="panel-bias-verdict-badge"></span>
    <span class="section-summary" id="panel-bias-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500"></span>
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:12px">Round-to-round selection bias: for each transition (R1→R2, R2→R3, …) are households that <em>stay</em> systematically different from those that <em>drop out</em>? Characteristics compared using data from the <em>previous</em> round.</p>
  <div id="panel-bias-table"></div>
  <div class="chart-grid" style="margin-top:14px">
    <div class="chart-box">
      <div class="ch-title">Regional Retention Rate by Transition <span id="bias-reg-chart-label" style="font-weight:normal;font-size:11px;color:#888">(R1→R2)</span></div>
      <div class="ch-sub">Click a transition tab above to update — sorted lowest to highest</div>
      <div style="height:340px"><canvas id="panelBiasRegChart"></canvas></div>
    </div>
    <div class="chart-box">
      <div class="ch-title">Sample Composition Drift</div>
      <div class="ch-sub">Each region's share of total sample: R1 vs latest complete round (pp change)</div>
      <div style="height:340px"><canvas id="panelBiasDriftChart"></canvas></div>
    </div>
  </div>
  <div style="margin-top:14px">
    <div class="ch-title" style="margin-bottom:6px">Retention Rate Heatmap — Region × Transition</div>
    <p style="font-size:11.5px;color:#666;margin-bottom:8px">% of previous-round HHs retained per transition, by region. Red &lt;50%, orange 50–65%, yellow 65–75%, green ≥75%.</p>
    <div id="panel-trans-heatmap"></div>
  </div>
  </div>
</div>
<div id="panel-stats" class="stats-row" style="margin-top:10px"></div>
<div id="panel-attrition-note" class="note-box note-info" style="margin-bottom:14px"></div>
<div class="chart-grid">
  <div class="chart-box"><div class="ch-title">Households per Round</div><div class="ch-sub">Retained from R1 vs new entries</div><div style="height:220px"><canvas id="panelAttrChart"></canvas></div></div>
  <div class="chart-box"><div class="ch-title">Participation Pattern Distribution</div><div class="ch-sub">Top patterns (1=present, 0=absent per round)</div><div style="height:220px"><canvas id="panelPatChart"></canvas></div></div>
</div>
<div class="card" style="margin-top:6px">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    🏠 Household Panel Tracker <span id="hh-matrix-badge" class="badge badge-blue"></span>
    <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500"><span style="color:#888">Full HH participation matrix</span></span>
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:10px">Every household across all rounds — when they joined, when they left, and their full participation history. Green = interviewed, Red = absent.</p>
  <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:10px">
    <input id="hh-search" type="text" placeholder="Search by HHID or PSU…"
      style="padding:6px 10px;border:1px solid #ccc;border-radius:4px;font-size:12.5px;width:200px">
    <select id="hh-region-filter" style="padding:6px 8px;border:1px solid #ccc;border-radius:4px;font-size:12.5px">
      <option value="">All Regions</option>
    </select>
    <div id="hh-status-filter" style="display:flex;gap:4px;flex-wrap:wrap"></div>
    <div id="hh-urban-filter" style="display:flex;gap:4px"></div>
    <button onclick="window._hhDownloadCsv()" style="margin-left:auto;padding:5px 12px;background:#1a2332;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">⬇ Download CSV</button>
  </div>
  <div id="hh-matrix-table"></div>
  <div id="hh-matrix-pagination" style="margin-top:8px;display:flex;gap:4px;align-items:center;flex-wrap:wrap"></div>
  </div>
</div>
<div class="card" style="margin-top:6px">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    📍 PSU Coverage vs Targets <span class="badge badge-blue">Urban: 6 HH/PSU &nbsp;|&nbsp; Rural: 7 HH/PSU</span>
  </h2>
  <div class="collapsible-body">
  <div id="panel-psu-status"></div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    🗺️ Household Counts by Region, Urban/Rural &amp; Round
  </h2>
  <div class="collapsible-body">
  <div id="panel-region-table"></div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    🔄 Round-by-Round In / Out Summary
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:10px">Tracked relative to R1 baseline. "New" = households not seen in R1.</p>
  <div id="panel-inout-table"></div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    🔁 Leavers vs. New Entries — Are Replacements Representative? <span id="lvn-verdict-badge" class="badge"></span>
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:12px">For each transition, compares households that <strong>left</strong> (using their last-observed characteristics) against households that <strong>entered</strong> (using first-observed characteristics). If replacements differ from leavers, the survey composition is shifting.</p>
  <div id="panel-lvn-summary"></div>
  <div class="chart-grid" style="margin-top:14px">
    <div class="chart-box">
      <div class="ch-title">% Urban: Leavers vs New Entries</div>
      <div class="ch-sub">Are new entries more/less urban than those who left?</div>
      <div style="height:220px"><canvas id="lvnUrbanChart"></canvas></div>
    </div>
    <div class="chart-box">
      <div class="ch-title">Mean HH Size: Leavers vs New Entries</div>
      <div class="ch-sub">Are new households larger/smaller?</div>
      <div style="height:220px"><canvas id="lvnHhsizeChart"></canvas></div>
    </div>
  </div>
  <div style="margin-top:14px">
    <div class="ch-title" style="margin-bottom:6px">Regional Composition: Leavers vs New Entries</div>
    <p style="font-size:11.5px;color:#666;margin-bottom:8px">Each region's share of leavers vs share of new entries per transition. Large differences indicate regional replacement bias.</p>
    <div id="panel-lvn-region"></div>
  </div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    👥 Attrition Composition — Who Stays, Who Leaves, Who Is New?
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:12px">Urban/Rural split of retained, dropped, and new-entry households per round transition.</p>
  <div id="panel-attrition-profile"></div>
  <div class="chart-grid" style="margin-top:14px">
    <div class="chart-box"><div class="ch-title">Urban/Rural Mix by Group</div><div class="ch-sub">% Urban among retained vs dropped vs new (each transition)</div><div style="height:240px"><canvas id="panelAttrProfileChart"></canvas></div></div>
    <div class="chart-box"><div class="ch-title">Volume by Group per Transition</div><div class="ch-sub">Absolute HH counts</div><div style="height:240px"><canvas id="panelAttrVolumeChart"></canvas></div></div>
  </div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    ⚠️ PSU Problem Tracker <span class="badge badge-red">Under-Target Only</span>
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:8px">PSUs that are below their household target in at least one round. Sorted by most rounds under target. Over-target is not flagged.</p>
  <div id="panel-psu-filter" style="margin-bottom:10px"></div>
  <div id="panel-psu-problems"></div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    🔗 Within-PSU Refusal Clustering Risk <span class="badge badge-red">Field Action Required</span>
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:8px">Cross-reference of under-target PSUs with round-to-round attrition rates by region. Regions flagged HIGH show both elevated attrition AND PSU coverage problems — two failure modes that compound each other and cannot both be fixed by weighting alone.</p>
  <div id="panel-refusal-risk"></div>
  </div>
</div>
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    📅 Call Interval Tracker <span id="panel-call-badge" class="badge badge-red"></span>
  </h2>
  <div class="collapsible-body">
  <p style="font-size:12px;color:#666;margin-bottom:10px">Days between consecutive interviews for the same household. Minimum required interval is 20 days. Intervals of 20&#x2013;29 days are permissible with supervisor approval; anything under 20 days is flagged as a violation.</p>
  <div id="panel-call-summary"></div>
  <div style="margin-top:14px">
    <div class="ch-title" style="margin-bottom:6px">Households Called Too Early (&lt;20 days since last interview)</div>
    <div id="panel-call-round-tabs" style="margin-bottom:8px"></div>
    <div id="panel-call-violations"></div>
  </div>
  </div>
</div>
</div>

<!-- ═══════ OPERATOR PERFORMANCE ═══════ -->
<div id="page-operators" class="page">
<h1>Operator Performance QC</h1>
<p style="font-size:12.5px;color:#555;margin-bottom:14px">
  Performance metrics per interviewer (int_id), derived from M00 passport and cross-module checks.
  Flags are independent indicators — a red operator should be reviewed, not automatically penalised.
  <strong>Short-circuit rates</strong> show the % of interviews where a skip-heavy gate answer was recorded
  (M03 SH1=2 "no shock", M04 A1=2 "not employed", M06 F1=2 "no bank account") compared with the fleet average.
  Unusually high rates may indicate that the operator is guiding respondents toward shorter-survey answers.
</p>

<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px" id="ops-rag-summary"></div>

<!-- Fleet Analytics Charts -->
<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    📊 Fleet Analytics Charts
    <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500"><span style="color:#888">Volume, flags, duration, scatter</span></span>
  </h2>
  <div class="collapsible-body">
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px">
    <div class="card" style="margin-bottom:0">
      <h2 style="font-size:13px">Interviews & Operators by Round</h2>
      <div style="position:relative;height:200px"><canvas id="chart-fleet-volume"></canvas></div>
    </div>
    <div class="card" style="margin-bottom:0">
      <h2 style="font-size:13px">Flag Distribution</h2>
      <div style="position:relative;height:200px"><canvas id="chart-fleet-flags"></canvas></div>
    </div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px">
    <div class="card" style="margin-bottom:0">
      <h2 style="font-size:13px">Duration vs Skip Violations</h2>
      <p style="font-size:10.5px;color:#888;margin-bottom:6px">Each bubble = one operator. Size = interview count. Color = RAG status.</p>
      <div style="position:relative;height:220px"><canvas id="chart-fleet-scatter"></canvas></div>
    </div>
    <div class="card" style="margin-bottom:0">
      <h2 style="font-size:13px">Module Duration — Fleet Distribution</h2>
      <p style="font-size:10.5px;color:#888;margin-bottom:6px">Mean duration per module across all operators (fleet P10–P90 range).</p>
      <div style="position:relative;height:220px"><canvas id="chart-fleet-moddur"></canvas></div>
    </div>
  </div>
  </div>
</div>

<div class="card">
  <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
    <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
    📋 Operator Table & Detail
    <span class="section-summary" id="ops-table-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500"></span>
  </h2>
  <div class="collapsible-body">
  <div style="display:flex;gap:10px;align-items:center;margin-bottom:10px;flex-wrap:wrap">
    <span style="font-size:11.5px;color:#888">Filter:</span>
    <div id="ops-filter-btns" style="display:flex;gap:6px;flex-wrap:wrap"></div>
    <input id="ops-search" type="text" placeholder="Search operator ID…"
      style="margin-left:8px;padding:4px 10px;border:1px solid #ccc;border-radius:4px;font-size:12px;width:160px"
      oninput="renderOpsTable()">
  </div>

  <div id="ops-table-wrap" style="overflow-x:auto"></div>

  <!-- Operator detail drawer -->
  <div id="ops-detail" style="display:none;margin-top:18px;background:#f4f7fb;border:1px solid #d0d8e4;border-radius:6px;padding:16px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
      <h2 style="margin:0;font-size:15px" id="ops-detail-title">Operator Detail</h2>
      <button onclick="document.getElementById('ops-detail').style.display='none'"
        style="background:none;border:none;font-size:18px;cursor:pointer;color:#888">✕</button>
    </div>
    <div id="ops-detail-body"></div>
  </div>
  </div>
</div>
</div>

<!-- ═══════ ALL CHANGES BY ROUND ═══════ -->
<div id="page-changes" class="page">
<h1>All Questionnaire Changes by Round</h1>
<p class="subtitle">Every new question, dropped question, wording change, skip-logic change, and code-list change across <span id="all-changes-round-range">R2–RN</span></p>
<div class="note-purple note-box">
  Changes are detected by comparing questionnaire workbooks for each consecutive round pair. The "presence" column shows which rounds include each variable.
</div>
<div id="changes-filters" class="round-filters" style="margin-bottom:14px"></div>
<div id="changes-content"></div>
</div>

<!-- ═══════ MODULE DEEP-DIVE PAGES (generated) ═══════ -->
<div id="module-pages-container"></div>

</div><!-- /main -->

<script>
// ── DATA ──────────────────────────────────────────────────────────────────────
const DQ   = """ + DQ   + """;
const MT   = """ + MT   + """;
const AQ   = """ + AQ   + """;
const PAN  = """ + PAN  + """;
const IVIEW= """ + IVIEW+ """;
const KOBO = """ + KOBO + """;
const ISSUES = """ + ISSUES + """;
const ISUM   = """ + ISUM + """;

const ROUNDS = [1,2,3,4,5,6,7,8];
const RLABELS = {1:'R1 (Nov)',2:'R2 (Dec)',3:'R3 (Jan)',4:'R4 (Feb)',5:'R5 (Mar)',6:'R6 (Apr)',7:'R7 (May)',8:'R8 (Jun)'};
const R_COLORS = ['#3498db','#2ecc71','#e67e22','#e74c3c','#8e44ad'];
const MODULES = ['M00','M01','M02','M03','M04','M05','M06','M07','M08','M09'];
const MOD_NAMES = {
  M00:'Passport',M01:'Roster',M02:'Education',
  M03:'Shocks',M04:'Employment',M05:'Income',M06:'Finance',
  M07:'Health',M08:'Food',M09:'Views'
};

// ── MODULE QUESTIONNAIRE NOTES ───────────────────────────────────────────────
// Surfaced on each module tab — documents significant Kobo XLSForm changes
// across rounds that affect data interpretation and skip logic.
// COMPREHENSIVE: covers wording changes, type changes, response option
// changes, added/dropped questions (with Kobo types), and variable renames.
const MODULE_NOTES = {
  'M00': {
    title: 'Questionnaire Evolution Notes — M00 Passport (8 Kobo Differences)',
    items: [
      {rounds:'R2→R3', tag:'Z9 wording change', text:'Z9 (consent script) wording changed: R1-R2 references household "visited in person"; R3+ changed to "called" to match CATI mode. Substantive framing difference — consent language shifted from in-person to phone context.'},
      {rounds:'R3→R4', tag:'X16 wording change', text:'X16 (interviewer remarks) label wording modified R3→R4. Minor phrasing adjustment in enumerator instructions. No change to response structure or data coding.'},
      {rounds:'R3+', tag:'X13 hint format', text:'X13_PHONE hint format changed: "63(10-digits)" → "63+(10-digits)" with added plus sign for international dialing format. X19A (select_one yes-no), X19B (select_one yes-none) added R3+ as alternative contact confirmation fields.'},
      {rounds:'R5', tag:'Z4 options + Z4_TXT', text:'Z4 (barangay) response options updated R5 — option list reflects field team changes. Z4_TXT (text type) added R5 only as free-text interviewer ID field.'},
      {rounds:'R1→R2', tag:'Choice list renames', text:'X17 (phone type) choice list renamed: x17→prepaid_postpaid. Code values 1=Prepaid/2=Postpaid unchanged — cosmetic Kobo cleanup only.'},
      {rounds:'R3+', tag:'Tracking vars added', text:'Pre-fill infrastructure added R3+: round_lastint (text), date_lastint (date), sample_rd (select_one). n_Z17 (calculate) added R3+. All are administrative routing variables.'},
      {rounds:'R2+', tag:'Panel linkage added', text:'oldz1-oldz5 (text, prior-round location), old_psgc (text, prior geo code) added R2+ for panel tracking. next_call (text) scheduling variable added R2+. Administrative, not questionnaire content.'},
      {rounds:'All', tag:'Synthetic vars', text:'Dashboard injects int_id, hhid, round, psu, region, urban as synthetic tracking variables not in Kobo XLSForm. Z0 split into Z0_first/Z0_last for name fields.'}
    ]
  },
  'M01': {
    title: 'Questionnaire Evolution Notes — M01 Roster (10 Kobo Differences)',
    items: [
      {rounds:'R1→R2→R3', tag:'D5a type change ⚠️', text:'D5a (member confirmation): R1 = select_one confirm_mem (4 options). R2 = select_multiple confirm_mem (5 options — code 5 "sex" added). R3+ = reverted to select_one (labels changed to "Yes, but..." prefix). Check R2 data carefully for multi-select encoding artifacts.'},
      {rounds:'R3+', tag:'D5a wording change', text:'D5a label text: R1-R2 asks to confirm "name and age"; R3+ expanded to confirm "name, age, and gender". Substantive addition reflecting data quality concern about gender verification in panel.'},
      {rounds:'R3+', tag:'D5a options changed ⚠️', text:'D5a confirm_mem options: R1 = 4 codes (1-4). R2 = 5 codes (added code 5 for sex mismatch). R3+ = labels reworded to "Yes, but name is different" / "Yes, but age is different" / "Yes, but sex is different" prefix format. Code values stable 1-5 from R2+, but label meaning shifted.'},
      {rounds:'R3+', tag:'M13 type change ⚠️', text:'M13 (move-in reason): R1-R2 = select_multiple M13 (multi-select). R3+ = select_one movein (single-select). Option 40 added R3. Major type change — R1-R2 data allows multiple reasons; R3+ only one reason per person. Pooling requires harmonization.'},
      {rounds:'R3+', tag:'D5a_/M13_ renames ⚠️', text:'D5a renamed to D5a_ (trailing underscore) from R3+. M13 renamed to M13_ from R3+. Skip logic syntax changed: selected(${D5a},X) → ${D5a_}=X. Dashboard normalizes both. Merging raw data across rounds requires explicit rename.'},
      {rounds:'R1→R2', tag:'D25 choice list', text:'D25 (leave reason) choice list renamed: D25→leave_reason. Code values unchanged — Kobo form cleanup only.'},
      {rounds:'R2+', tag:'D3/D33 added', text:'D3 (select_one surveylang) added R2+ — captures survey language. D33 (select_one gender) added R2+ — explicit gender tracking for births/deaths. Both are substantive new questions, not administrative.'},
      {rounds:'R2+', tag:'Panel member IDs', text:'oldisfmid1-24 (24 calculate variables) added R2+ for member tracking across rounds. d30index_new (calculate) added R3+ for new member indexing. old_d23 (calculate) added R3+ for prior-round HH composition comparison. newmembers_ct (calculate) R3+.'},
      {rounds:'R3+', tag:'D26/M10 sub-expansion', text:'D26 (moved-to destination) and M10 (migrated-from origin) sub-items expanded in dashboard heatmap as _1 (country), _2 (province), _3 (city). Case normalization applied: D5a (not D5A), D25_oth (not D25_OTH), M13_oth (not M13_OTH). Dashboard case-insensitive lookup handles both.'},
      {rounds:'All', tag:'18 auth vars', text:'18 authoritative variables tracked: hhid, fmid, hhsize, age, gender, relationship, isfmid, member_leftreason/_oth, country/prov/city_moved, moved_in_reason/_oth, country/province/city_migrated_from, dur_rr. Most impactful changes: D5a type oscillation (select_one to select_multiple to select_one) and M13 type change (select_multiple to select_one R3+).'}
    ]
  },
  'M02': {
    title: 'Questionnaire Evolution Notes — M02 Education (4 Kobo Differences)',
    items: [
      {rounds:'R1→R2', tag:'Structure collapse ⚠️', text:'R1 uses person-indexed Kobo variables: ED15_1 through ED15_5, ED16_1 through ED16_25 (plus _oth variants, ~75 vars total, all select_one type). R2+ collapses to single generic ED15/ED16 (both select_one). CATI mode asks one person per call, not all 5. ~72 person-indexed vars dropped R2+.'},
      {rounds:'R3+', tag:'ED16_1_ rename', text:'ED16_1 renamed to ED16_1_ (trailing underscore) from R3+. Same systematic rename pattern as D5a_, M13_, etc. Only affects R1 holdover variable name in pooled data.'},
      {rounds:'R2+', tag:'ED16 choice list', text:'ED16 choice list renamed: ED16→dropout (R2+). R1 had ED16_1-ED16_25 person-indexed choices; R2+ uses consolidated generic naming. Code values unchanged — 1=Illness, 2=Financial, etc. remain stable.'},
      {rounds:'All', tag:'Skip logic note ⚠️', text:'All ed16_* indicators (ed16_1 through ed16_14, ed16_96, ed16_oth) are conditional on ed15=2 (not currently studying). Kobo: ED16 relevant=${ED15}=2. In the pooled .dta, ED16 is select_multiple with codes 1-14 plus 96 (other). Stata export splits into binary dummies ed16_1 through ed16_96. These dummies inherit the same ed15=2 skip — missing values for ed15=1 (currently studying) are structural, not errors.'},
      {rounds:'All', tag:'19 auth vars', text:'19 authoritative variables tracked: ed15, ed16, ed16_1 through ed16_14 (split dummies), ed16_96, ed16_oth, dur_edu. No expenditure questions in CATI M02 — education spending (EL5) is CAPI only. The R1→R2 person-indexed structure collapse is the only major Kobo difference.'}
    ]
  },
  'M03': {
    title: 'Questionnaire Evolution Notes — M03 Shocks (12 Kobo Differences)',
    items: [
      {rounds:'R1→R2', tag:'EL5 type change', text:'EL5 (electricity hours unavailable, integer) changed to decimal in R2+, allowing fractional hours. This is an electricity/utility variable housed in the Shocks module group — not related to education despite the "EL" prefix.'},
      {rounds:'R3+', tag:'SH1b wording change ⚠️', text:'SH1b (shock types experienced) recall period changed: R1-R2 asks about "past 30 days"; R3+ asks about "past 3 months". Major comparability break — R3+ captures a wider recall window. Prevalence rates expected to be higher R3+.'},
      {rounds:'R3+', tag:'SH1b hint change ⚠️', text:'SH1b interviewer hint changed: R1-R2 says "IF MENTIONED" (interviewer marks what respondent volunteers); R3+ says "DO NOT READ OPTIONS" (explicit protocol). Effectively same intent but stronger wording R3+.'},
      {rounds:'R2→R3', tag:'SH1b options changed ⚠️', text:'SH1b shock options (select_multiple shocks): code 96 ("Other") removed in R2 (cannot capture unlisted shocks in R2 only). Code 96 restored R3+. R2 data may undercount unusual shocks. All other codes (1-15) stable across rounds.'},
      {rounds:'R3+', tag:'SH1b_oth wording', text:'SH1b_oth (other shock specify) label simplified R3: removed redundant phrasing. Cosmetic wording cleanup — no change to data capture.'},
      {rounds:'R3+', tag:'SH3 wording + choice ⚠️', text:'SH3 (received assistance) wording: R3 adds "DRINKING WATER" in caps. Choice list changed: yes-no → yes-none (R3+). SAME code values 1=Yes, 2=No/None — verified identical. Safe to pool without recoding.'},
      {rounds:'R1→R2', tag:'SH4 type change ⚠️', text:'SH4 (days affected) changed from integer to decimal (R2+). Constraint loosened: >=1 → >=0.1 (allows fractional days). Missing codes unchanged (-99/-95). R1 integer-only data merges cleanly but check for rounding artifacts.'},
      {rounds:'R3+', tag:'SH1b_/SH2_ renames ⚠️', text:'SH1b renamed to SH1b_ from R3+ (trailing underscore). SH2 (select_multiple copingmech) renamed to SH2_ from R3+. SH2_96 and SH2_96_oth (other coping specify) present R1-R2 only, dropped R3+. SH1b_naturalhazard (calculate) added R3+.'},
      {rounds:'R1→R2', tag:'Choice list renames', text:'Systematic choice list renaming: SH1b→shocks, S2→copingmech (for SH2). Code values unchanged — cosmetic Kobo cleanup only.'},
      {rounds:'R3 only', tag:'NH block added ⚠️', text:'13 natural hazard variables collected R3 ONLY (Typhoon Kristine window): N1 (select_one), N3 (select_multiple), NH2 (select_one), NH3 (integer), NH7 (select_multiple), NH10 (select_one), NH14-NH17 (mixed types, plus _oth text variants). Not in any other round. NH vars are select_multiple INSIDE repeat group: after split+reshape, nh7_3_2 = warning channel 3 for hazard instance 2.'},
      {rounds:'R1-R2', tag:'Validation vars dropped', text:'validate_SH1b (calculate) present R1-R2 only, removed R3+. n_sH1b (note, shock count) present R1-R2 only. n_hhsize (note) R1 only. All administrative — no substantive data loss.'},
      {rounds:'R3+', tag:'Calculation refs updated', text:'SH1b_ans and SH1b_ansname calculations update variable references from SH1b to SH1b_ in R3+. Functionally equivalent but automated validation must track both reference forms.'}
    ]
  },
  'M04': {
    title: 'Questionnaire Evolution Notes — M04 Employment (14 Kobo Differences)',
    items: [
      {rounds:'R1→R2', tag:'A4 type change ⚠️', text:'A4/A4_new (occupation) changed from text (free-text entry) to select_one occupation in R2+. Major structural change: open-ended → coded dropdown. R1 occupation data requires manual post-coding to match R2+ categories. Cannot pool R1 text with R2+ coded responses without harmonization.'},
      {rounds:'R1→R2', tag:'Choice list renames ⚠️', text:'25 choice lists systematically renamed R2: A3→reason_nojob, A5→industry, A6→class_worker, A7→subs_farming, A9→digi_platform, A16→contract, A17→contract_duration, A20→whom_worked, A24→change_employ, A26→employ_status2, A27→employ_status. All code values unchanged — Kobo cleanup only.'},
      {rounds:'R4+', tag:'A6 options changed ⚠️', text:'A6 (class of worker) select_one class_worker: code 9 ("None of the above") removed R4+. R1-R3 has 9 options; R4-R5 has 8 options. Workers who would have selected "None" in R4+ forced into other categories. Affects classification comparability.'},
      {rounds:'R3+', tag:'A8 choice change', text:'A8 (gig work) select_one choice list changed: yes-no-dk → yes-none-dk (R3+). Code values: 1=Yes, 2=No/None, -99=DK — verified identical across lists. Same "no"→"none" pattern as SH3 in M03. Safe to pool.'},
      {rounds:'R3+', tag:'A19/A21 renames + drop ⚠️', text:'A19 (select_multiple employ_benefits) renamed A19_ from R3+. A21 (select_multiple mode_transpo) renamed A21_ from R3+. CRITICAL: A19 and A21 NOT collected R4-R5 — gated out entirely. Split dummies (a19_1-5, a21_1-3) in HF data but not in heatmap. Benefits and transport data stops at R3.'},
      {rounds:'R1→R2', tag:'Constraints changed ⚠️', text:'A22 (integer, hours worked) constraint loosened: >=1 → >=0 (allows zero hours, R2+). A23 (integer, earnings) constraint: >=1→>=0, and missing code changed from -99 to -900 (R2+). Critical for data validation — different missing codes across rounds means -99 in R2+ is a real value, not missing.'},
      {rounds:'R3+', tag:'A25 hint wording', text:'A25_oldA8A9 hint text added R3: "Gig work involves short-term contracts, freelance work..." R4+ adds red bold HTML styling (<font color="red"><b>...</b></font>) to the hint. No question logic change — interviewer guidance only.'},
      {rounds:'R4-R5', tag:'Role shift ⚠️', text:'a24 (select_one change_employ), a25 (select_one), a26 (select_one employ_status2), a27 (select_one employ_status) present ALL rounds but role shifts: R1-R3 these are follow-up confirmation; R4-R5 they become primary employment routing gates. a19/a21 NOT collected R4-R5.'},
      {rounds:'R3+', tag:'Income calc change ⚠️', text:'IA3_TOTAL_INCOME and IA3_TOTAL_INCOME_NEW (calculate) changed R3+: simple sum (${IA3_A}+...) → conditional sum (if(>0, val, 0)+...). Negative income values treated as zero in R3+. Affects total income aggregation for analysis.'},
      {rounds:'R3+', tag:'Pre-fill vars added', text:'fmidA24, fmidA27, fmidfieldA24, fmidfieldA27 (all calculate) added R3+ for employment status pre-fill. R5 adds fmid_employment (calculate, consolidated gate). income_fmidA24/A27 (calculate) added R3+ for cross-module income routing.'},
      {rounds:'R4+', tag:'fmidA8 label change', text:'fmidA8label (calculate) values updated R4+: from "Yes"/"No" to "Yes,with gig work"/"No gig work". Cosmetic label change reflecting increased emphasis on gig economy classification.'},
      {rounds:'R2+', tag:'Tracking vars', text:'employ_fmidage (calculate), employ_new_fmid (calculate), fmid_list (calculate) added R2+. employ_addmembers (select_one) and n_employ (note) present R1 only — replaced by fmid-based tracking R2+.'},
      {rounds:'All', tag:'28 auth vars', text:'28 authoritative variables tracked. Most impactful: A4 text→select_one (R2), A6 option dropped (R4), A19/A21 collection stopped (R4), A23 missing code -99→-900 (R2), income calc logic change (R3). Five distinct data breaks across rounds.'},
      {rounds:'All', tag:'Response option summary', text:'Stable options: A3 reason_nojob (8 codes+other), A5 industry (22 codes+other), A16 contract (3 codes), A17 contract_duration (4 codes), A27 employ_status (5 codes). Changed: A6 class_worker lost code 9 R4+. A8 yes-no-dk→yes-none-dk R3+ (same codes).'}
    ]
  },
  'M05': {
    title: 'Questionnaire Evolution Notes — M05 Income (12 Kobo Differences)',
    items: [
      {rounds:'R1-R2 → R3+', tag:'Recall period ⚠️', text:'IA2/IA5 wording: R1 = "Sep-Oct 2025" (2-month). R2 = "Nov-Dec 2025" (2-month). R3+ = "in the past month" (rolling 1-month). Income levels NOT directly comparable: R1-R2 capture 2 months of income, R3+ captures 1 month. Must adjust for recall window when pooling.'},
      {rounds:'R3+', tag:'IA6_C-F wording', text:'IA6_C through IA6_F (income sub-categories, all integer type) label wording simplified R3 — shortened question text for phone interview efficiency. No change to response structure or codes.'},
      {rounds:'R3+', tag:'Income calcs changed ⚠️', text:'8+ calculation variables changed R3+: IA5_TOTAL_INCOME, TOTAL_INCOME, TOTAL_CASH_EARNINGS, TOTAL_EARNINGS_INKIND (plus _NEW, _GIG, _NEW_GIG variants). All changed from simple sum to conditional sum: if(>0, val, 0). Negative income values zeroed out R3+. Affects total income aggregation.'},
      {rounds:'R3+', tag:'IA3/IA5 bypass ⚠️', text:'R1-R2: IA3 (select_one) requires IA2=1. R3+: IA3 can be filled via pre-fill bypass (employed members routed directly). IA5 (integer): gated on fmidA1=2 in R1-R2 (~27-31% missing); unconditional R3+ (near 0% missing). IA2=2 → IA3 filled is NOT a skip violation R3+.'},
      {rounds:'R3 only', tag:'IC/ID block added ⚠️', text:'Entire IC section (~20 vars: IC1_A-C select_one, IC3-IC9 integer/select_one + sub-items) and ID section (~6 vars: ID1 select_one, ID2_A-B integer + grouping) collected R3 ONLY. "Other income source" questions were a one-round experiment. All show 100% missing R1-R2 and R4-R5 — not an error.'},
      {rounds:'R5', tag:'hhid type change ⚠️', text:'hhid variable type changed from text to select_one SN in R5 Kobo form. Household ID captured via dropdown instead of free text. Critical for data merge validation — verify R5 hhid encoding matches R1-R4 text format.'},
      {rounds:'R3+', tag:'IA routing evolution', text:'IA2 gate evolves: R1-R2 on fmidA1=2 only. R3-R4 add fmidA24=2/fmidA27=2. R5 consolidates to fmid_employment=2. Pre-fill variables exist in raw Kobo but dropped in Stata cleaning. group_othersourceincome (begin_group) R3 only — tied to IC/ID experimental block.'},
      {rounds:'R3+', tag:'Pre-fill vars added', text:'income_fmidage (calculate), income_fmidA24/A27 (calculate), income_round_lastint (calculate) added R3+. R5 adds income_fmid_employment and income_fmidfieldemploy (consolidated gates). income_addmembers (select_one) R1-R2 only — replaced by fmid-based routing.'},
      {rounds:'R3+', tag:'Note vars added', text:'n_IA5_new and n_IA6 (note, interviewer instruction) added R3+. n_income (note) R1 only.'},
      {rounds:'R4+', tag:'Zero-income validation', text:'TOTAL_INCOME_zero and TOTAL_INCOME_new_zero (calculate) validation fields added R4+. Flag zero-income households for enumerator verification. h10_H11_hhid (calculate) linkage added R5 for cross-module reference.'},
      {rounds:'All', tag:'IA7 gate stable', text:'IA7 (integer) gated on M04 A9 (gig work) via indexed-repeat — consistent all rounds. ~92-100% missing is structural (non-gig workers), not an error.'},
      {rounds:'All', tag:'Response options stable', text:'Core income response structure stable: IA3_A-F (select_one yes-no/yes-none), IA6_A-F (integer amounts). Choice list renames R2 (cosmetic only). The 3 data breaks are: recall period (R3), calc logic (R3), and IA5 gating (R3).'}
    ]
  },
  'M06': {
    title: 'Questionnaire Evolution Notes — M06 Finance (8 Kobo Differences)',
    items: [
      {rounds:'R3+', tag:'F1-F16 wording change', text:'All core finance questions: label wording "PAST 30 DAYS" changed to "PAST MONTH" in R3+. No substantive recall period change — cosmetic wording harmonization. Both mean the same recall window.'},
      {rounds:'R5', tag:'F6 threshold change ⚠️⚠️', text:'F6 (emergency fund capacity) threshold CHANGED: R1-R4 asks about ability to come up with PHP 300,000; R5 asks about PHP 50,000. CRITICAL data break — R5 emergency capacity rate is NOT comparable to R1-R4. A 6x lower threshold will mechanically increase the "yes" rate. Must be flagged in all trend analysis.'},
      {rounds:'R5', tag:'F17/F18 added ⚠️', text:'F17 (select_one yes-nonev2, formal bank account gate) and F18 (select_one yes-nonev2, mobile money account gate) added R5 only. R5 Kobo reorders questions: F17→F1→F18→F2→...→F16. F17/F18 are NOT in pooled .dta — R5 skip checks referencing these are dead code in earlier rounds.'},
      {rounds:'R3+', tag:'F8→F8_ rename ⚠️', text:'F8 (select_multiple loan_purpose) renamed to F8_ from R3+ (trailing underscore). Same systematic pattern as D5a_, M13_, SH1b_, SH2_, A19_, A21_. Dashboard normalizes via case-insensitive lookup stripping underscores.'},
      {rounds:'R1→R2', tag:'Choice list renames', text:'F16 choice list renamed: f16→fraud. F9: F9→institution. F8: F8→loan_purpose. All code values unchanged — cosmetic Kobo form cleanup only. No response option additions or removals.'},
      {rounds:'R4+', tag:'FIES gating vars', text:'hidden_select_f08 (select_one) added R4+ for conditional FIES section display. f08_hhid (calculate) household linkage added R4+. Administrative routing variables, not substantive question changes.'},
      {rounds:'All', tag:'Response options stable', text:'Core choice lists stable: F1 (yes-no, 2 codes), F9 institution (8 codes+other), F16 fraud (6 codes+other), F8 loan_purpose (10 codes+other). No option additions/removals across any round except the F6 threshold change.'},
      {rounds:'All', tag:'Stable core', text:'Core finance questions (F1-F15) structurally unchanged across all 5 rounds. F16 (fraud awareness) consistent. Module is one of the most stable — but F6 threshold break in R5 is the single most impactful change in any module for trend comparability.'}
    ]
  },
  'M07': {
    title: 'Questionnaire Evolution Notes — M07 Health (9 Kobo Differences)',
    items: [
      {rounds:'R1-R4 vs R5', tag:'Major expansion ⚠️', text:'R1-R4: only 4 core health variables — H2 (select_one healthcare_service), H2A (select_one yes-no), H3 (select_one reason_nohealthcare), H3_oth (text). R5: ~25 new variables added covering facility type, medicine, out-of-pocket costs, co-morbidities, hospitalization, and health roster. Most health vars show "—" for R1-R4 — structural, not error.'},
      {rounds:'R5', tag:'H2 wording change ⚠️', text:'H2 (healthcare access) recall period changed: R1-R4 asks about "last 30 days"; R5 asks about "past 3 months". Wider recall window in R5 will mechanically increase healthcare utilization rates. Not comparable to R1-R4 without adjustment.'},
      {rounds:'R5', tag:'New vars — types ⚠️', text:'R5 additions with Kobo types: H4 (select_one facility, ~8 options), H7 (select_one medicine, yes-no), H8 (select_one type_payment, 3 options: 1=Cash, 2=Insurance, 3=Both), H8_amt (integer, OOP amount), H9A-C (select_one, co-morbidity severity), H10 (select_one, hospitalized), H11 (select_one, hospital admission), H11B (select_multiple who_paid, 10 options including PhilHealth/family/savings/loans).'},
      {rounds:'R5', tag:'Hospital cascade added', text:'H12 (integer, hospital days), H13 (select_one hosp_type, 4 options: 1=Public, 2=Private, 3=Military, 4=Other), H14 (integer, hospital bill), H15 (select_one, bill payment method), H16 (text, hospital specify). Full hospitalization module is R5-only.'},
      {rounds:'R5', tag:'H9 multilingual', text:'H9_ansname added in 5 language variants: _en (English), _tl (Tagalog), _ceb (Cebuano), _hil (Hiligaynon), _war (Waray). Enables language-appropriate health condition display. H9A-C code conditions into severity categories.'},
      {rounds:'R5', tag:'H17 roster + routing', text:'H17 (health member repeat) with old/new/grpnew sub-variants added R5. H10_H11_hhid (calculate) and H10_H11 (begin_group) added for hospitalization tracking. R5 Kobo has ~19 skip rules vs 3 in R1-R4. Key chains: H2A←H2; H3←H2A=2; H4/H7/H8←H2=2/3; H12→H13→H14→H15→H16.'},
      {rounds:'R1→R2', tag:'Choice list renames', text:'H2 choice list: H2→healthcare_service. H3: H3→reason_nohealthcare. Both renamed R2+. Code values unchanged — same systematic Kobo cleanup pattern.'},
      {rounds:'R2+', tag:'A4_oth cross-ref', text:'A4_oth and A4_new_oth (text, occupation "other specify") added R2+ in M07 Kobo group. Cross-module reference to M04 occupation — used for health-employment linkage analysis.'},
      {rounds:'All', tag:'Response options summary', text:'H2 healthcare_service options stable R1-R5 (5 codes). H3 reason_nohealthcare stable (8 codes+other). R5-only lists: who_paid (10 codes), hosp_type (4 codes), type_payment (3 codes), facility (~8 codes). All new R5 lists fully documented in Kobo choices sheet.'}
    ]
  },
  'M08': {
    title: 'Questionnaire Evolution Notes — M08 Food (3 Kobo Differences)',
    items: [
      {rounds:'R4+', tag:'Structure change ⚠️', text:'group_food-nonfood changed from begin_group to begin_repeat in R4+. This structural change affects how food/non-food expenditure data is captured in Kobo — nested repeat vs flat group. May affect variable export structure and downstream merge logic.'},
      {rounds:'R4-R5', tag:'Gating vars added', text:'group_f08 (begin_group) wrapper added R5. hidden_select_f08 (select_one) gating variable added R4+ for conditional FIES display. f08_hhid (calculate) household linkage added R4+. All administrative routing variables.'},
      {rounds:'All', tag:'Core FIES unchanged', text:'Core F08_A through F08_E FIES items (all select_one, 2 codes: 1=Yes, 2=No) unchanged across all 5 rounds. Response options identical. Most stable substantive module in the survey — only administrative gating changes R4+.'}
    ]
  },
  'M09': {
    title: 'Questionnaire Evolution Notes — M09 Views (9 Kobo Differences)',
    items: [
      {rounds:'R3+', tag:'V9 hint protocol ⚠️⚠️', text:'V9_C through V9_M (10 Likert items, all select_one Agreement/5-point scale) hint changed: R1-R2 = "Do not read scale/choices" (self-anchoring); R3+ = "Read scale/choices" (interviewer-read). MAJOR protocol shift affecting response distributions. R1-R2 vs R3+ Likert responses are not directly comparable — expect higher use of extreme categories when scale is read aloud.'},
      {rounds:'R3+', tag:'V9_F wording changed', text:'V9_F (select_one Agreement) wording changed TWICE: R1-R2 asks about "education" concerns; R3 reworded to "resources for school"; R4-R5 may have further minor adjustments. Substantive framing shift — "education" vs "resources for school" primes different response patterns.'},
      {rounds:'R2→R4', tag:'V9_J numbering shift', text:'V9_J (select_one Agreement) Kobo row position shifted: R2→R3→R4 each have slightly different placement in the V9 sub-item list. Question content may be stable but display order to enumerator changed. Check for order effects in response patterns.'},
      {rounds:'R4+', tag:'V11/V12 dropped ⚠️', text:'V11 (select_one yes-no, economic change perception) and V12 (select_one Support/5-point scale, life improvement expectations) present R1-R3, removed R4+. Originally baseline questions extended through R3. Dropped to shorten R4-R5 instrument. Time series ends at R3 for these items.'},
      {rounds:'R4+', tag:'V9_B dropped ⚠️', text:'V9_B (select_one Agreement, one of 13 Likert sub-items) present R1-R3 only, dropped R4+. Likely complexity reduction to shorten interview duration. V9 parent question (select_one) appears R5 only as summary wrapper.'},
      {rounds:'R1→R2', tag:'V5 choice list rename', text:'V5 (economic outlook) choice list renamed: QLI→eco_change (R2+). 5-point scale: 1=Much worse, 2=Somewhat worse, 3=About the same, 4=Somewhat better, 5=Much better. Code values unchanged — cosmetic Kobo cleanup only.'},
      {rounds:'R4+', tag:'V9 gating restructure', text:'hidden_select_v9 (select_one) gating variable added R4+. v9_hhid (calculate) household linkage added R4+. R1-R4 collect individual V9_A-M sub-items directly; R5 has V9 (select_one) parent wrapper before sub-items.'},
      {rounds:'R2+', tag:'Tracking vars', text:'n_opinions_views (note) added R2+. n_op&v (note) present R1 only (original naming). vhhid (calculate) R5 only.'},
      {rounds:'All', tag:'Stable core + options', text:'V9_A, V9_D, V9_H (core sentiment including AI worry) present all 5 rounds, no structural changes. V5 eco_change (5 codes) and V6 life satisfaction (5 codes: 1=Very dissatisfied to 5=Very satisfied) stable. Agreement scale (5 codes: 1=Strongly disagree to 5=Strongly agree) and Support scale (5 codes) unchanged. Key data breaks: hint protocol (R3) and V11/V12/V9_B drop (R4).'}
    ]
  }
};

// ── MODULE VARIABLE ORDER (Kobo XLSForm question order) ──────────────────────
// Each module maps to an ordered array of {kobo, stata} pairs.
// Used for consistent variable display across Tracker, Heatmap, OOR, Kobo panels.
const MODULE_VAR_ORDER = {
  'M02': [
    // ── User-specified Kobo order (19 vars, dur_edu last) ──
    // ed16_N are Stata split dummies from select_multiple ED16 (dropout reasons).
    // All ed16_* conditional on ed15=2 (not currently studying).
    {kobo:'ED15',stata:'ed15'},
    {kobo:'ED16',stata:'ed16'},
    {kobo:'ed16_1',stata:'ed16_1'},{kobo:'ed16_2',stata:'ed16_2'},
    {kobo:'ed16_3',stata:'ed16_3'},{kobo:'ed16_4',stata:'ed16_4'},
    {kobo:'ed16_5',stata:'ed16_5'},{kobo:'ed16_6',stata:'ed16_6'},
    {kobo:'ed16_7',stata:'ed16_7'},{kobo:'ed16_8',stata:'ed16_8'},
    {kobo:'ed16_9',stata:'ed16_9'},{kobo:'ed16_10',stata:'ed16_10'},
    {kobo:'ed16_11',stata:'ed16_11'},{kobo:'ed16_12',stata:'ed16_12'},
    {kobo:'ed16_13',stata:'ed16_13'},{kobo:'ed16_14',stata:'ed16_14'},
    {kobo:'ed16_96',stata:'ed16_96'},
    {kobo:'ED16_oth',stata:'ed16_oth'},
    {kobo:'dur_edu',stata:'dur_edu'}
  ],
  'M01': [
    // ── User-specified Kobo order (18 vars, dur_rr last) ──
    {kobo:'hhid',stata:'hhid'},
    {kobo:'fmid',stata:'fmid'},
    {kobo:'hhsize',stata:'hhsize'},
    {kobo:'D5A (age)',stata:'age'},
    {kobo:'D5A (gender)',stata:'gender'},
    {kobo:'D6',stata:'relationship'},
    {kobo:'D5A (isfmid)',stata:'isfmid'},
    {kobo:'D25',stata:'member_leftreason'},
    {kobo:'D25_oth',stata:'member_leftreason_oth'},
    {kobo:'D26_1',stata:'country_moved'},
    {kobo:'D26_2',stata:'prov_moved'},
    {kobo:'D26_3',stata:'city_moved'},
    {kobo:'M13',stata:'moved_in_reason'},
    {kobo:'M13_oth',stata:'moved_in_reason_oth'},
    {kobo:'M10_1',stata:'country_migrated_from'},
    {kobo:'M10_2',stata:'province_migrated_from'},
    {kobo:'M10_3',stata:'city_migrated_from'},
    {kobo:'dur_rr',stata:'dur_rr'}
  ],
  'M00': [
    // ── Kobo question order (R05 XLSForm) ──
    {kobo:'Z0_first',stata:'z0_first'},{kobo:'Z0_last',stata:'z0_last'},
    {kobo:'int_id',stata:'int_id'},{kobo:'hhid',stata:'hhid'},
    {kobo:'replacement_hhid',stata:'replacement_hhid'},
    {kobo:'D3',stata:'survey_lang'},
    {kobo:'Z6',stata:'date_of_interview'},{kobo:'Z7',stata:'time_of_interview'},
    {kobo:'Z8',stata:'call_attemp'},
    {kobo:'member_called',stata:'member_called'},
    {kobo:'call_status1',stata:'call_status1'},
    {kobo:'Z16',stata:'correct_resp'},
    {kobo:'Z17',stata:'call_result'},
    {kobo:'member_talkedto',stata:'resp_fmid'},
    {kobo:'Z9',stata:'agreement'},
    {kobo:'Z18',stata:'refusal_reason'},{kobo:'Z18_oth',stata:'refusal_reason_oth'},
    {kobo:'Z19',stata:'interview_record'},
    {kobo:'Z20',stata:'address_unchanged'},{kobo:'Z21',stata:'new_address_str'},
    {kobo:'Z1',stata:'region'},{kobo:'Z2',stata:'province'},
    {kobo:'Z3',stata:'city'},{kobo:'Z4',stata:'barangay'},
    {kobo:'Z4_txt',stata:'z4_txt'},{kobo:'Z5',stata:'locale'},
    // ── Pipeline-computed vars (not Kobo questions) ──
    {kobo:'hhsize',stata:'hhsize'},
    {kobo:'dur_pp',stata:'dur_pp'},{kobo:'dur_rr',stata:'dur_rr'},
    {kobo:'dur_educ',stata:'dur_educ'},{kobo:'dur_sh',stata:'dur_sh'},
    {kobo:'dur_emp',stata:'dur_emp'},{kobo:'dur_inc',stata:'dur_inc'},
    {kobo:'dur_fin',stata:'dur_fin'},{kobo:'dur_hlt',stata:'dur_hlt'},
    {kobo:'dur_f_nf',stata:'dur_f_nf'},{kobo:'dur_vw',stata:'dur_vw'},
    {kobo:'dur_tot',stata:'dur_tot'}
  ]
};

// Module → own duration variable (for OOR filtering)
const MODULE_DUR_VAR = {
  'M00':'dur_pp','M01':'dur_rr','M02':'dur_educ','M03':'dur_sh','M04':'dur_emp',
  'M05':'dur_inc','M06':'dur_fin','M07':'dur_hlt','M08':'dur_f_nf','M09':'dur_vw'
};

// ── Helper functions for variable ordering & labeling ────────────────────────
function _buildVarMaps(mod){
  const order = MODULE_VAR_ORDER[mod];
  if(!order) return null;
  const m = {};
  order.forEach((v,i)=>{
    const kl = v.kobo.toLowerCase(), sl = v.stata.toLowerCase();
    const label = kl===sl ? v.kobo : v.kobo+' ('+v.stata+')';
    const entry = {idx:i, kobo:v.kobo, stata:v.stata, label};
    m[kl] = entry; if(kl!==sl) m[sl] = entry;
  });
  return m;
}
function _varLabel(v, vmap){ if(!vmap) return v; const e=vmap[(v||'').toLowerCase()]; return e?e.label:v; }
function _varSort(v, vmap){ if(!vmap) return 999; const e=vmap[(v||'').toLowerCase()]; return e?e.idx:999; }

// ── ROUTING ──────────────────────────────────────────────────────────────────
let currentPage = 'overview';
function renderIssues(){
  const firmOnly = document.getElementById('iss-firm-only') && document.getElementById('iss-firm-only').checked;
  const revOnly  = document.getElementById('iss-review-only') && document.getElementById('iss-review-only').checked;
  const OWN = {A1:'firm-questionnaire',A2:'firm-field',B:'firm-dofile',C:'us',D:'expected',REVIEW:'unassigned'};
  let rows = ISSUES.slice();
  if(firmOnly) rows = rows.filter(r=>['A1','A2','B'].includes(r.verdict) && ['new','acknowledged','fix-pending','reopened'].includes(r.status));
  if(revOnly)  rows = rows.filter(r=>r.review);
  const byMod = {};
  rows.forEach(r=>{ (byMod[r.module]=byMod[r.module]||[]).push(r); });
  const counts = Object.values(ISUM||{}).reduce((a,s)=>{a.open+=s.open||0;a.closed+=s.closed||0;return a;},{open:0,closed:0});
  const vc = {}; ISSUES.forEach(r=>{ vc[r.verdict] = (vc[r.verdict]||0)+1; });
  const vcLine = Object.keys(vc).sort().map(v=>`${v}:${vc[v]}`).join(' · ');
  let html = `<div class="mstat">Showing ${rows.length} issue(s) · ${counts.open} open / ${counts.closed} closed total</div>`
           + `<div class="mstat">By verdict: ${vcLine}</div>`;
  Object.keys(byMod).sort().forEach(m=>{
    html += `<h2 style="margin-top:16px">${m} – ${MOD_NAMES[m]||''}</h2>`;
    byMod[m].forEach((r,i)=>{
      const ev = r.evidence||{}; const k=ev.kobo||{}; const d=ev.dofile||{}; const da=ev.data||{};
      const rel = Object.entries(k.relevant_by_round||{}).slice(-1).map(([rd,x])=>`R${rd}: ${x||'(none)'}`).join('');
      const miss = (k.gate_refs_missing||[]).length ? '  ·  gate refs absent from data: '+k.gate_refs_missing.join(', ') : '';
      const note = r.notes ? '  ·  Note: '+r.notes : '';
      const cnts = Object.entries(r.counts_by_round||{}).map(([rd,n])=>`R${rd}:${n}`).join('  ');
      const did = `iss-${m}-${i}`;
      html += `<div style="border:1px solid #e3e3e3;border-radius:5px;padding:7px 10px;margin:5px 0">
        <div style="cursor:pointer" onclick="var e=document.getElementById('${did}');e.style.display=e.style.display==='none'?'block':'none'">
          <span class="vbadge ${r.verdict}">${r.verdict}</span>
          <span class="schip">${r.status}</span>
          <span class="schip">${OWN[r.verdict]||''}</span>
          <strong style="font-size:12px">&nbsp;${r.variable}</strong>
          <span style="font-size:11px;color:#666">&nbsp;${(r.label||'').slice(0,70)}</span>
          <span style="float:right;font-size:10.5px;color:#888">${cnts}</span>
        </div>
        <div id="${did}" style="display:none"><div class="evbox">Data    · ${da.total||0} total · kind ${da.kind||''}
Kobo    · ${rel||'(var not in Kobo)'}${miss}
Do-file · ${d.ever_touched?'touched by a round do-file':'not touched by any do-file'}
Verdict · ${r.verdict} via ${r.rule_fired} (confidence ${r.confidence})${note}</div></div>
      </div>`;
    });
  });
  document.getElementById('issues-body').innerHTML = html || '<p>No issues.</p>';
}

function showPage(id){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('#sidebar a').forEach(a=>a.classList.remove('active'));
  const pg = document.getElementById('page-'+id);
  if(pg) pg.classList.add('active');
  const nav = document.getElementById('nav-'+id);
  if(nav) nav.classList.add('active');
  if(id==='issues') renderIssues();
  currentPage = id;
  return false;
}

// ── HELPERS ──────────────────────────────────────────────────────────────────
function hmColor(v){
  if(v===null||v===undefined) return '#eee';
  if(v<=0) return '#1a9641';
  if(v<=20) return '#a6d96a';
  if(v<=50) return '#ffffbf';
  if(v<=80) return '#fdae61';
  return '#d73027';
}
function hmText(v){
  if(v===null||v===undefined) return '—';
  return v.toFixed(0)+'%';
}
function makeBar(id,labels,datasets,extraOpts={}){
  const el=document.getElementById(id); if(!el)return;
  new Chart(el.getContext('2d'),{type:'bar',data:{labels,datasets},options:{
    responsive:true,maintainAspectRatio:false,
    plugins:{legend:{display:false,labels:{font:{size:10},boxWidth:8}},...(extraOpts.plugins||{})},
    scales:{x:{ticks:{font:{size:10}}},y:{ticks:{font:{size:10}}},...(extraOpts.scales||{})},
    ...extraOpts
  }});
}
function makeLine(id,labels,datasets,extraOpts={}){
  const el=document.getElementById(id); if(!el)return;
  new Chart(el.getContext('2d'),{type:'line',data:{labels,datasets},options:{
    responsive:true,maintainAspectRatio:false,
    plugins:{legend:{display:true,labels:{font:{size:10},boxWidth:8}},...(extraOpts.plugins||{})},
    scales:{x:{ticks:{font:{size:10}}},y:{ticks:{font:{size:10}}},...(extraOpts.scales||{})},
    ...extraOpts
  }});
}

// ── SIDEBAR MODULE LINKS ──────────────────────────────────────────────────────
function buildModNavLinks(){
  const c = document.getElementById('mod-nav-links');
  const rag_map = {};
  MODULES.forEach(m=>{rag_map[m]=DQ.module_summary[m]?.rag||'green'});
  const rag_colors = {red:'#e74c3c',yellow:'#f39c12',green:'#2ecc71'};
  c.innerHTML = MODULES.map(m=>`
    <a href="#" onclick="return showPage('mod-${m}')" id="nav-mod-${m}">
      <span class="dot" style="background:${rag_colors[rag_map[m]||'green']}"></span>
      ${m} ${MOD_NAMES[m]}
    </a>`).join('');
}

// ── OVERVIEW ─────────────────────────────────────────────────────────────────
function buildOverview(){
  buildModNavLinks();

  // Subtitle round range — updates automatically with ROUNDS
  const _rrSpan = document.getElementById('overview-round-range');
  if(_rrSpan) _rrSpan.textContent = `Rounds 1–${ROUNDS.length}`;

  // Top stats
  const totalSkip = DQ.skip_issues.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0),0);
  const totalMand = DQ.mandatory_issues.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0),0);
  const totalOOR  = DQ.oor_issues.reduce((s,x)=>s+Object.values(x.counts||{}).reduce((a,v)=>a+(v||0),0),0);

  // Skip/mand nav counts removed since these nav items no longer exist

  document.getElementById('ov-stats').innerHTML = `
    <div class="stat-box ${totalSkip>0?'red':'green'}"><div class="num">${totalSkip}</div><div class="lbl">Skip violations</div></div>
    <div class="stat-box ${totalMand>0?'yellow':'green'}"><div class="num">${totalMand}</div><div class="lbl">Mandatory missing</div></div>
    <div class="stat-box ${totalOOR>0?'yellow':'green'}"><div class="num">${totalOOR}</div><div class="lbl">Out-of-range values</div></div>
    <div class="stat-box blue"><div class="num">${MODULES.length}</div><div class="lbl">Modules tracked</div></div>
    <div class="stat-box blue"><div class="num">${ROUNDS.length}</div><div class="lbl">Rounds (R1–R${ROUNDS.length})</div></div>
    <div class="stat-box purple"><div class="num">${countAllQChanges()}</div><div class="lbl">Questionnaire changes</div></div>
  `;

  // Module grid
  const mg = document.getElementById('mod-grid');
  mg.innerHTML = MODULES.map(m=>{
    const s = DQ.module_summary[m]||{};
    const iss = ISUM[m] || {strip:{}, headline:'green', open:0, closed:0, by_owner:{}};
    const _miss = s.max_missing_pct || 0;
    const _missSig = _miss >= 30 ? 'red' : (_miss >= 10 ? 'yellow' : 'green');
    const _ord = {green:0, yellow:1, red:2};
    const _ih = iss.headline || 'green';
    const rag = _ord[_ih] >= _ord[_missSig] ? _ih : _missSig;   // worst of open-issue status + missing%
    const istrip = [1,2,3,4,5,6,7,8].map(r=>{
      const st = iss.strip[String(r)] || 'green';
      const ch = st==='red' ? '!' : (st==='closed' ? '·' : '');
      return `<span class="idot ${st}" title="R${r}: ${st}">${ch}</span>`;
    }).join('');
    const ownerBits = Object.entries(iss.by_owner||{}).map(([o,n])=>`${n} ${o.replace('firm-','')}`).join(' · ');
    const ragLabel = {red:'⚠ Issues',yellow:'⚡ Watch',green:'✓ OK'};
    const ragChipBg = {red:'#fde',yellow:'#fff3cd',green:'#d4edda'};
    const ragChipColor = {red:'#c0392b',yellow:'#856404',green:'#155724'};

    // Count questionnaire changes for this module
    const rows = MT[m]||[];
    const newQ = rows.filter(r=>r.status&&r.status.startsWith('New')).length;
    const droppedQ = rows.filter(r=>r.status&&r.status.startsWith('Dropped')).length;
    const changedQ = rows.filter(r=>r.title_changes||r.skip_changes||r.option_changes).length;

    // Build "why" triggers list (open-issue + missing-data driven)
    const triggers = [];
    if(iss.open > 0) triggers.push(`${iss.open} open issue${iss.open>1?'s':''}`);
    if(_miss >= 10) triggers.push(`worst variable ${_miss.toFixed(0)}% missing`);
    const triggerHtml = triggers.length
      ? `<div style="margin-top:4px;padding:4px 6px;background:${rag==='red'?'rgba(231,76,60,0.12)':'rgba(243,156,18,0.12)'};border-radius:4px;font-size:10.5px;color:${rag==='red'?'#e74c3c':'#f39c12'}"><strong>Why ${rag}:</strong> ${triggers.join(' · ')}</div>`
      : '';

    return `<div class="mod-card ${rag}" onclick="showPage('mod-${m}')">
      <div class="rag-chip" style="background:${ragChipBg[rag]};color:${ragChipColor[rag]}">${ragLabel[rag]}</div>
      <div class="mname">${m} – ${MOD_NAMES[m]}</div>
      <div class="istrip" title="Per-round status (open issues only)">${istrip}</div>
      <div class="mstat">Issues: ${iss.open} open${ownerBits?` (${ownerBits})`:''} · ${iss.closed} closed</div>
      <div class="mstat ${s.n_skip_violations>0?'warn':''}">Skip violations: ${s.n_skip_violations||0}</div>
      <div class="mstat ${s.n_mandatory_missing>0?'warn':''}">Mandatory missing: ${s.n_mandatory_missing||0}</div>
      <div class="mstat">Max missing: ${(s.max_missing_pct||0).toFixed(1)}% · Avg: ${(s.avg_missing_pct||0).toFixed(1)}%</div>
      ${newQ?`<div class="mstat warn">+${newQ} new question${newQ>1?'s':''}</div>`:''}
      ${droppedQ?`<div class="mstat warn">−${droppedQ} dropped</div>`:''}
      ${changedQ?`<div class="mstat">~${changedQ} changed</div>`:''}
      ${triggerHtml}
    </div>`;
  }).join('');

  // Sample chart
  setTimeout(()=>{
    const sc = DQ.sample_counts;
    makeBar('sampleChart', MODULES,
      ROUNDS.map((r,i)=>({
        label: RLABELS[r],
        data: MODULES.map(m=>sc[m]?sc[m][r]??0:0),
        backgroundColor: R_COLORS[i], borderRadius:3
      })),
      {plugins:{legend:{display:true,labels:{font:{size:10},boxWidth:8}}},
       scales:{x:{ticks:{font:{size:10},maxRotation:0}},y:{ticks:{font:{size:10}}}}}
    );
  },50);

  buildChangeSummaryBar();
}

function countAllQChanges(){
  let n=0;
  MODULES.forEach(m=>{
    (MT[m]||[]).forEach(row=>{
      if(row.status&&(row.status.startsWith('New')||row.status.startsWith('Dropped'))) n++;
      if(row.title_changes) n++;
      if(row.skip_changes) n++;
      if(row.option_changes) n++;
    });
  });
  return n;
}

function buildChangeSummaryBar(){
  const container = document.getElementById('change-summary-bar');
  // Count changes by round
  const rounds = ['R2','R3','R4','R5','R6','R7','R8'];
  const prev   = {R2:'R1',R3:'R2',R4:'R3',R5:'R4',R6:'R5',R7:'R6',R8:'R7'};
  let html = '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px">';
  rounds.forEach(rnd=>{
    let newQ=0,dropped=0,changed=0;
    MODULES.forEach(m=>{
      (MT[m]||[]).forEach(row=>{
        const inC = row[`in_${rnd}`]==='✓';
        const inP = row[`in_${prev[rnd]}`]==='✓';
        const rIdx = ['R1','R2','R3','R4','R5','R6','R7','R8'].indexOf(rnd);
        const isFirst = ['R1','R2','R3','R4','R5','R6','R7','R8'].slice(0,rIdx).every(pr=>row[`in_${pr}`]!=='✓');
        if(inC&&isFirst) newQ++;
        if(!inC&&inP) dropped++;
        if(inC&&inP&&(row.title_changes||row.skip_changes||row.option_changes)) changed++;
      });
    });
    html += `<div class="card" style="margin:0;padding:14px">
      <div style="font-weight:700;font-size:14px;color:#1a2332;margin-bottom:8px">${rnd} vs ${prev[rnd]}</div>
      ${newQ?`<div class="chg-tag chg-new" style="margin-bottom:4px">+${newQ} new question${newQ>1?'s':''}</div><br>`:''}
      ${dropped?`<div class="chg-tag chg-drop" style="margin-bottom:4px">−${dropped} dropped</div><br>`:''}
      ${changed?`<div class="chg-tag chg-title" style="margin-bottom:4px">~${changed} changed</div>`:''}
      ${!newQ&&!dropped&&!changed?'<span style="font-size:12px;color:#888">No structural changes</span>':''}
    </div>`;
  });
  html += '</div>';
  container.innerHTML = html;
}

// ── SKIP VIOLATIONS ──────────────────────────────────────────────────────────
function buildSkip(){
  const issues = DQ.skip_issues;
  const withViolations = issues.filter(x=>Object.values(x.counts_by_round).some(v=>v>0));
  const total = issues.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0),0);

  document.getElementById('skip-stats').innerHTML = `
    <div class="stat-box ${total>0?'red':'green'}"><div class="num">${total}</div><div class="lbl">Total violations</div></div>
    <div class="stat-box ${withViolations.length>0?'yellow':'green'}"><div class="num">${withViolations.length}</div><div class="lbl">Checks with violations</div></div>
    <div class="stat-box blue"><div class="num">${issues.length}</div><div class="lbl">Total checks</div></div>
  `;

  const list = document.getElementById('skip-list');
  list.innerHTML = issues.map(x=>{
    const total = Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
    const sev = x.severity||'medium';
    const icon = total===0?'🟢':sev==='high'?'🔴':'🟡';
    const rowCls = total===0?'clean':sev;
    const pills = ROUNDS.map(r=>{
      const raw = x.counts_by_round[r]??x.counts_by_round[String(r)]??null;
      const hasData = raw!=null;
      const v = raw||0;
      const p = x.pct_by_round?.[r]??x.pct_by_round?.[String(r)]??0;
      const cls = !hasData?'vpill-grey':v>0?'vpill-red':'vpill-green';
      return `<span class="vpill ${cls}">R${r}: ${hasData?`${v} (${(p||0).toFixed(1)}%)`:'N/A'}</span>`;
    }).join('');
    return `<div class="viol-row ${rowCls}">
      <div class="viol-icon">${icon}</div>
      <div class="viol-text">
        <strong>${x.module}: ${x.rule}</strong>
        <div class="viol-path">${x.variable}</div>
        <div class="viol-note">${x.note||''}</div>
        <div class="viol-pills">${pills}</div>
      </div>
    </div>`;
  }).join('');

  setTimeout(()=>{
    const charts = document.getElementById('skip-charts');
    withViolations.forEach((x,i)=>{
      const cid = `sc_${i}`;
      charts.innerHTML += `<div class="chart-box flagged">
        <div class="ch-title">${x.module}: ${x.variable.split('→')[0].trim()}</div>
        <div class="ch-sub">${x.rule}</div>
        <div style="height:160px"><canvas id="${cid}"></canvas></div>
      </div>`;
      setTimeout(()=>makeBar(cid, ROUNDS.map(r=>`R${r}`), [{
        label:'Violations', data:ROUNDS.map(r=>x.counts_by_round[r]??x.counts_by_round[String(r)]??0),
        backgroundColor: ROUNDS.map(r=>(x.counts_by_round[r]??x.counts_by_round[String(r)]??0)>0?'#e74c3c':'#2ecc71'), borderRadius:4
      }]),30);
    });
  },50);
}

// ── MANDATORY MISSING ────────────────────────────────────────────────────────
function buildMandatory(){
  const issues = DQ.mandatory_issues;
  const withFails = issues.filter(x=>Object.values(x.counts_by_round).some(v=>v>0));
  const total = issues.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0),0);

  document.getElementById('mand-stats').innerHTML = `
    <div class="stat-box ${total>0?'yellow':'green'}"><div class="num">${total}</div><div class="lbl">Total failures</div></div>
    <div class="stat-box blue"><div class="num">${issues.length}</div><div class="lbl">Total checks</div></div>
  `;

  document.getElementById('mand-list').innerHTML = issues.map(x=>{
    const total = Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
    if(total===0) return '';  // All green → hide
    const sev = x.severity||'medium';
    const icon = sev==='high'?'🔴':'🟡';
    const rowCls = sev;
    const pills = ROUNDS.map(r=>{
      const raw = x.counts_by_round[r]??x.counts_by_round[String(r)]??null;
      const hasData = raw!=null;
      const v = raw||0;
      const cls = !hasData?'vpill-grey':v>0?'vpill-red':'vpill-green';
      return `<span class="vpill ${cls}">R${r}: ${hasData?v:'N/A'}</span>`;
    }).join('');
    return `<div class="viol-row ${rowCls}">
      <div class="viol-icon">${icon}</div>
      <div class="viol-text">
        <strong>${x.module}: ${x.rule}</strong>
        <div class="viol-path">${x.variable}</div>
        <div class="viol-note">${x.note||''}</div>
        <div class="viol-pills">${pills}</div>
      </div>
    </div>`;
  }).join('');
}

// ── OOR ──────────────────────────────────────────────────────────────────────
function buildOOR(){
  const issues = DQ.oor_issues;
  const total = issues.reduce((s,x)=>s+Object.values(x.counts||{}).reduce((a,v)=>a+(v||0),0),0);
  document.getElementById('oor-stats').innerHTML = `
    <div class="stat-box ${total>0?'yellow':'green'}"><div class="num">${total}</div><div class="lbl">Total out-of-range</div></div>
    <div class="stat-box blue"><div class="num">${issues.length}</div><div class="lbl">Variables checked</div></div>
  `;
  document.getElementById('oor-list').innerHTML = issues.map(x=>{
    const total=Object.values(x.counts||{}).reduce((a,v)=>a+(v||0),0);
    const sev=x.severity||'medium';
    const icon = total===0?'🟢':sev==='high'?'🔴':'🟡';
    const rowCls = total===0?'clean':sev;
    const pills = ROUNDS.map(r=>{
      const raw = x.counts?.[r]??x.counts?.[String(r)]??null;
      const hasData = raw!=null;
      const v = raw||0;
      const cls = !hasData?'vpill-grey':v>0?'vpill-red':'vpill-green';
      return `<span class="vpill ${cls}">R${r}: ${hasData?v:'N/A'}</span>`;
    }).join('');
    return `<div class="viol-row ${rowCls}">
      <div class="viol-icon">${icon}</div>
      <div class="viol-text">
        <strong>${x.module}: ${x.label||x.variable}</strong>
        <div class="viol-path">${x.variable} — ${x.rule}</div>
        <div class="viol-note">${x.note||''}</div>
        <div class="viol-pills">${pills}</div>
      </div>
    </div>`;
  }).join('');

  setTimeout(()=>{
    const charts = document.getElementById('oor-charts');
    issues.filter(x=>Object.values(x.counts||{}).some(v=>v>0)).forEach((x,i)=>{
      const cid=`oor_${i}`;
      charts.innerHTML += `<div class="chart-box warn">
        <div class="ch-title">${x.module}: ${x.label||x.variable}</div>
        <div class="ch-sub">${x.rule}</div>
        <div style="height:150px"><canvas id="${cid}"></canvas></div>
      </div>`;
      setTimeout(()=>makeBar(cid,ROUNDS.map(r=>`R${r}`),[{
        label:'Out-of-range',data:ROUNDS.map(r=>x.counts?.[r]??x.counts?.[String(r)]??0),
        backgroundColor:ROUNDS.map(r=>(x.counts?.[r]??x.counts?.[String(r)]??0)>0?'#f39c12':'#2ecc71'),borderRadius:4
      }]),30);
    });
  },50);
}

// ── PANEL TRACKING ───────────────────────────────────────────────────────────
// ── OPERATOR PERFORMANCE ─────────────────────────────────────────────────────
let _opsFilter = 'all';

function buildOperators(){
  if(!IVIEW || !IVIEW.operators) return;
  const fleet = IVIEW.fleet || {};
  const ops   = IVIEW.operators || [];

  // Nav badge
  const nRed = fleet.n_red || 0;
  const badge = document.getElementById('nc-ops-red');
  if(badge){ badge.textContent=nRed; badge.className=`nav-count ${nRed>0?'red':''}`; }

  // RAG summary boxes
  const sumEl = document.getElementById('ops-rag-summary');
  if(sumEl){
    const RAG_CLR = {red:'#e74c3c', amber:'#e67e22', green:'#27ae60'};
    sumEl.innerHTML = `
      <div class="stat-box red"><div class="num">${fleet.n_red||0}</div><div class="lbl">🔴 Needs review</div></div>
      <div class="stat-box yellow"><div class="num">${fleet.n_amber||0}</div><div class="lbl">🟡 Watch list</div></div>
      <div class="stat-box green"><div class="num">${fleet.n_green||0}</div><div class="lbl">🟢 OK</div></div>
      <div class="stat-box"><div class="num">${fleet.n_operators||0}</div><div class="lbl">Total operators</div></div>
      <div class="stat-box"><div class="num">${fleet.median_dur||'—'}</div><div class="lbl">Fleet median duration (min)</div></div>`;
  }
  // Populate collapse summary for operator table
  const otSumEl = document.getElementById('ops-table-summary');
  if(otSumEl){
    otSumEl.innerHTML=`<span style="color:#888">${fleet.n_operators||0} operators</span>
      <span style="color:#e74c3c">${fleet.n_red||0} red</span>
      <span style="color:#e67e22">${fleet.n_amber||0} amber</span>
      <span style="color:#27ae60">${fleet.n_green||0} green</span>`;
  }

  // Filter buttons
  const filtEl = document.getElementById('ops-filter-btns');
  if(filtEl){
    filtEl.innerHTML = ['all','red','amber','green'].map(f=>{
      const lbl = f==='all' ? `All (${ops.length})` :
                  f==='red' ? `🔴 Review (${fleet.n_red||0})` :
                  f==='amber' ? `🟡 Watch (${fleet.n_amber||0})` :
                  `🟢 OK (${fleet.n_green||0})`;
      const clr = f==='red'?'#e74c3c':f==='amber'?'#e67e22':f==='green'?'#27ae60':'#2d3f55';
      return `<button onclick="_opsFilter='${f}';renderOpsTable()"
        id="ops-flt-${f}"
        style="padding:3px 11px;border-radius:3px;font-size:11.5px;font-weight:600;cursor:pointer;
               border:1.5px solid ${clr};background:#f8f9fa;color:${clr}">${lbl}</button>`;
    }).join('');
  }

  renderOpsTable();
  renderFleetCharts();
}

function renderFleetCharts(){
  if(!IVIEW || !IVIEW.fleet) return;
  const fleet = IVIEW.fleet;

  // ── 1. Interviews & Operators by Round (grouped bar) ──
  const rs = fleet.round_stats || {};
  const rLabels = Object.keys(rs).sort().map(r=>`R${r}`);
  const rKeys   = Object.keys(rs).sort();
  if(rLabels.length>0){
    const ctx1 = document.getElementById('chart-fleet-volume');
    if(ctx1) new Chart(ctx1, {
      type:'bar',
      data:{
        labels: rLabels,
        datasets:[
          { label:'Interviews', data:rKeys.map(r=>rs[r].n_interviews), backgroundColor:'#3498db', yAxisID:'y' },
          { label:'Operators', data:rKeys.map(r=>rs[r].n_operators), backgroundColor:'#e67e22', yAxisID:'y1' },
          { label:'Skip violations', data:rKeys.map(r=>rs[r].total_skip_violations), backgroundColor:'#e74c3c', yAxisID:'y1' }
        ]
      },
      options:{
        responsive:true, maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
        scales:{
          y:{position:'left',title:{display:true,text:'Interviews',font:{size:10}},ticks:{font:{size:10}}},
          y1:{position:'right',title:{display:true,text:'Count',font:{size:10}},ticks:{font:{size:10}},grid:{drawOnChartArea:false}}
        }
      }
    });
  }

  // ── 2. Flag Distribution (horizontal bar) ──
  const flagDist = fleet.flag_distribution || {};
  const flagLabels = Object.keys(flagDist);
  const flagVals = Object.values(flagDist);
  if(flagLabels.length>0){
    const ctx2 = document.getElementById('chart-fleet-flags');
    if(ctx2) new Chart(ctx2, {
      type:'bar',
      data:{
        labels: flagLabels,
        datasets:[{
          data: flagVals,
          backgroundColor: flagLabels.map(l=>{
            if(l.includes('Skip')) return '#e74c3c';
            if(l.includes('attrition')) return '#e67e22';
            if(l.includes('missing')) return '#9b59b6';
            if(l.includes('replacement')) return '#f39c12';
            if(l.includes('Short') || l.includes('fast')) return '#2980b9';
            return '#95a5a6';
          })
        }]
      },
      options:{
        indexAxis:'y', responsive:true, maintainAspectRatio:false,
        plugins:{legend:{display:false}},
        scales:{x:{ticks:{font:{size:10}}},y:{ticks:{font:{size:10}}}}
      }
    });
  }

  // ── 3. Duration vs Skip Violations (bubble scatter) ──
  const scatter = fleet.scatter || [];
  if(scatter.length>0){
    const ragColor = {red:'rgba(231,76,60,0.6)', amber:'rgba(230,126,34,0.6)', green:'rgba(39,174,96,0.6)'};
    const datasets = ['red','amber','green'].map(rag=>({
      label: rag.charAt(0).toUpperCase()+rag.slice(1),
      data: scatter.filter(s=>s.rag===rag && s.dur).map(s=>({
        x: s.dur,
        y: s.skips,
        r: Math.max(3, Math.min(15, Math.sqrt(s.n)*0.8))
      })),
      backgroundColor: ragColor[rag],
      borderColor: ragColor[rag].replace('0.6','1'),
      borderWidth: 1
    }));
    const ctx3 = document.getElementById('chart-fleet-scatter');
    if(ctx3) new Chart(ctx3, {
      type:'bubble',
      data:{datasets},
      options:{
        responsive:true, maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
        scales:{
          x:{title:{display:true,text:'Median duration (min)',font:{size:10}},ticks:{font:{size:10}}},
          y:{title:{display:true,text:'Skip violations',font:{size:10}},ticks:{font:{size:10}}}
        }
      }
    });
  }

  // ── 4. Module Duration Fleet Distribution (range bar) ──
  const modDur = fleet.module_dur_fleet || {};
  const modKeys = Object.keys(modDur);
  if(modKeys.length>0){
    const labels = modKeys.map(k=>modDur[k].label);
    const means  = modKeys.map(k=>modDur[k].fleet_mean);
    const p10s   = modKeys.map(k=>modDur[k].fleet_p10);
    const p90s   = modKeys.map(k=>modDur[k].fleet_p90);
    const ctx4 = document.getElementById('chart-fleet-moddur');
    if(ctx4) new Chart(ctx4, {
      type:'bar',
      data:{
        labels,
        datasets:[
          { label:'Fleet mean', data:means, backgroundColor:'#3498db', borderRadius:3 },
          { label:'P10', data:p10s, backgroundColor:'rgba(52,152,219,0.2)', borderColor:'#3498db', borderWidth:1, borderDash:[3,3], type:'line', pointRadius:3, fill:false },
          { label:'P90', data:p90s, backgroundColor:'rgba(231,76,60,0.2)', borderColor:'#e74c3c', borderWidth:1, borderDash:[3,3], type:'line', pointRadius:3, fill:false }
        ]
      },
      options:{
        responsive:true, maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
        scales:{
          y:{title:{display:true,text:'Minutes',font:{size:10}},ticks:{font:{size:10}}}
        }
      }
    });
  }
}

function renderOpsTable(){
  if(!IVIEW || !IVIEW.operators) return;
  const fleet = IVIEW.fleet || {};
  const sc_fleet = fleet.fleet_shortcircuit || {};

  const search = (document.getElementById('ops-search')||{value:''}).value.trim().toLowerCase();

  // Update button styles
  ['all','red','amber','green'].forEach(f=>{
    const btn = document.getElementById(`ops-flt-${f}`);
    if(!btn) return;
    const clr = f==='red'?'#e74c3c':f==='amber'?'#e67e22':f==='green'?'#27ae60':'#2d3f55';
    btn.style.background = _opsFilter===f ? clr : '#f8f9fa';
    btn.style.color = _opsFilter===f ? '#fff' : clr;
  });

  const filtered = IVIEW.operators.filter(op=>{
    if(_opsFilter !== 'all' && op.rag !== _opsFilter) return false;
    if(search && !String(op.int_id).includes(search)) return false;
    return true;
  });

  const wrapEl = document.getElementById('ops-table-wrap');
  if(!wrapEl) return;
  if(filtered.length===0){
    wrapEl.innerHTML='<p style="color:#888;padding:12px">No operators match the current filter.</p>';
    return;
  }

  const RAG_ICON = {red:'🔴', amber:'🟡', green:'🟢'};
  const RAG_BG   = {red:'#fde8e8', amber:'#fff8e1', green:'#e8f8ee'};

  let html=`<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:900px">
    <thead><tr style="background:#1a2332;color:#fff;font-size:11.5px">
      <th style="padding:7px 10px;text-align:center">Status</th>
      <th style="padding:7px 10px">Operator ID</th>
      <th style="padding:7px 10px;text-align:center">Interviews</th>
      <th style="padding:7px 10px;text-align:center">Rounds</th>
      <th style="padding:7px 10px;text-align:center">Median dur (min)</th>
      <th style="padding:7px 10px;text-align:center">vs fleet</th>
      <th style="padding:7px 10px;text-align:center">% &lt;5min</th>
      <th style="padding:7px 10px;text-align:center">Skip viols</th>
      <th style="padding:7px 10px;text-align:center">Replacements</th>
      <th style="padding:7px 10px;text-align:center">Attrition %</th>
      <th style="padding:7px 10px;text-align:center">Late-night</th>
      <th style="padding:7px 10px;text-align:left">Flags</th>
      <th style="padding:7px 10px;text-align:center">Detail</th>
    </tr></thead><tbody>`;

  filtered.forEach((op,i)=>{
    const bg = i%2===0 ? '#f8f9fa' : '#fff';
    const dur = op.duration || {};
    const ratio = dur.ratio_to_median;
    const ratioBg = ratio!==null && ratio<0.75 ? '#e74c3c' : ratio!==null && ratio<0.85 ? '#e67e22' : 'transparent';
    const ratioTxt = ratio!==null && ratio<0.85 ? '#fff' : '#444';
    const ratioStr = ratio!=null ? `${(ratio*100).toFixed(0)}%` : '—';

    const skipTot = (op.skip_violations||{})._total || 0;
    const skipClr = skipTot>10 ? '#e74c3c' : skipTot>3 ? '#e67e22' : '#27ae60';

    const repPct = (op.replacements||{}).pct || 0;
    const repClr = repPct>10 ? '#e74c3c' : repPct>5 ? '#e67e22' : '#444';

    const attr = ((op.attrition||{})._overall)||{};
    const attrPct = attr.pct || 0;
    const attrN   = attr.n_total || 0;
    const attrClr = attrPct>20 ? '#e74c3c' : attrPct>12 ? '#e67e22' : '#444';
    const attrStr = attrN>=5 ? `${attrPct.toFixed(1)}%` : '—';

    const tf = op.time_flags || {};
    const late = tf.n_late_night || 0;
    const lateClr = late>3 ? '#e74c3c' : late>0 ? '#e67e22' : '#444';

    const flags = op.flags || [];
    const flagHtml = flags.length===0
      ? '<span style="color:#27ae60;font-size:11px">✓ No flags</span>'
      : flags.map(f=>`<span style="display:inline-block;background:#fde8e8;color:#c0392b;border-radius:3px;padding:1px 5px;font-size:10.5px;margin:1px">${f}</span>`).join('');

    const rounds = (op.rounds||[]).map(r=>`R${r}`).join(' ');

    html+=`<tr style="background:${bg}">
      <td style="padding:6px 10px;text-align:center;font-size:16px">${RAG_ICON[op.rag]||''}</td>
      <td style="padding:6px 10px;font-weight:700;color:#1a2332">${op.int_id}</td>
      <td style="padding:6px 10px;text-align:center">${op.n_interviews}</td>
      <td style="padding:6px 10px;text-align:center;font-size:11px;color:#666">${rounds}</td>
      <td style="padding:6px 10px;text-align:center;font-weight:600">${dur.median??'—'}</td>
      <td style="padding:6px 10px;text-align:center">
        <span style="background:${ratioBg};color:${ratioTxt};border-radius:3px;padding:2px 6px;font-size:11px;font-weight:600">${ratioStr}</span>
      </td>
      <td style="padding:6px 10px;text-align:center;color:${dur.pct_fast>10?'#e74c3c':dur.pct_fast>5?'#e67e22':'#444'};font-weight:${dur.pct_fast>5?700:400}">${dur.pct_fast??0}%</td>
      <td style="padding:6px 10px;text-align:center;color:${skipClr};font-weight:700">${skipTot}</td>
      <td style="padding:6px 10px;text-align:center;color:${repClr}">${repPct.toFixed(1)}%</td>
      <td style="padding:6px 10px;text-align:center;color:${attrClr}">${attrStr}</td>
      <td style="padding:6px 10px;text-align:center;color:${lateClr}">${late>0?late:'—'}</td>
      <td style="padding:6px 10px;max-width:260px">${flagHtml}</td>
      <td style="padding:6px 10px;text-align:center">
        <button onclick="showOpDetail(${op.int_id})"
          style="padding:2px 8px;background:#2d3f55;color:#fff;border:none;border-radius:3px;cursor:pointer;font-size:11px">▶ View</button>
      </td>
    </tr>`;
  });

  html+=`</tbody></table>`;
  wrapEl.innerHTML=html;
}

function showOpDetail(intId){
  if(!IVIEW) return;
  const op = IVIEW.operators.find(o=>o.int_id===intId);
  if(!op) return;
  const fleet = IVIEW.fleet || {};
  const sc_fleet = fleet.fleet_shortcircuit || {};
  const dur = op.duration || {};
  const attr = (op.attrition||{})._overall || {};
  const tf   = op.time_flags || {};
  const sc   = op.shortcircuit || {};

  document.getElementById('ops-detail-title').textContent =
    `Operator ${intId}  —  ${op.n_interviews} interviews  (${(op.rounds||[]).map(r=>'R'+r).join(', ')})`;

  const RAG_BG = {red:'#fde8e8', amber:'#fff8e1', green:'#e8f8ee'};
  const RAG_CLR = {red:'#c0392b', amber:'#d35400', green:'#1e8449'};
  const RAG_LBL = {red:'⚠ Needs Review', amber:'⚡ Watch List', green:'✓ OK'};

  // Duration breakdown
  let durRows = '';
  const modDur = op.module_dur || {};
  Object.values(modDur).forEach(md=>{
    if(!md.mean) return;
    const zAbs = Math.abs(md.z||0);
    const zClr = zAbs>2 ? '#e74c3c' : zAbs>1.5 ? '#e67e22' : '#444';
    durRows+=`<tr>
      <td style="padding:4px 10px">${md.label}</td>
      <td style="padding:4px 10px;text-align:center">${md.mean} min</td>
      <td style="padding:4px 10px;text-align:center;color:${zClr};font-weight:${zAbs>1.5?700:400}">${md.z!=null?md.z:'—'}</td>
    </tr>`;
  });

  // Skip violations detail
  const skips = op.skip_violations || {};
  const skipDetail = Object.entries(skips).filter(([k])=>k!=='_total')
    .map(([mod,cnt])=>`<span style="background:#fde8e8;color:#c0392b;border-radius:3px;padding:2px 8px;margin:2px;display:inline-block;font-size:11.5px">${mod}: ${cnt}</span>`)
    .join('') || '<span style="color:#27ae60">None</span>';

  // Short-circuit rates
  const scLabels = {sh1_skip:'M03 No-shock (SH1=2)', a1_skip:'M04 Not-employed (A1=2)', f1_skip:'M06 No-account (F1=2)'};
  let scRows = '';
  Object.entries(scLabels).forEach(([key, label])=>{
    const opPct  = (sc[key]||{}).pct;
    const fltPct = sc_fleet[key];
    if(opPct==null) return;
    const diff = opPct - (fltPct||0);
    const clr = diff>15 ? '#e74c3c' : diff>8 ? '#e67e22' : '#27ae60';
    const diffStr = diff>=0 ? `+${diff.toFixed(1)}pp` : `${diff.toFixed(1)}pp`;
    scRows+=`<tr>
      <td style="padding:4px 10px">${label}</td>
      <td style="padding:4px 10px;text-align:center;font-weight:600">${opPct.toFixed(1)}%</td>
      <td style="padding:4px 10px;text-align:center">${fltPct!=null?fltPct.toFixed(1)+'%':'—'}</td>
      <td style="padding:4px 10px;text-align:center;color:${clr};font-weight:${Math.abs(diff)>8?700:400}">${diffStr}</td>
    </tr>`;
  });

  // Attrition by transition
  let attrRows = '';
  Object.entries(op.attrition||{}).filter(([k])=>k!=='_overall').forEach(([key,d])=>{
    const clr = d.pct>20?'#e74c3c':d.pct>12?'#e67e22':'#444';
    attrRows+=`<tr>
      <td style="padding:4px 10px">${key}</td>
      <td style="padding:4px 10px;text-align:center">${d.n_total} HHs</td>
      <td style="padding:4px 10px;text-align:center;color:${clr};font-weight:${d.pct>12?700:400}">${d.pct.toFixed(1)}%</td>
    </tr>`;
  });

  const body = `
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:16px">
      <div style="background:${RAG_BG[op.rag]};border-radius:6px;padding:10px;text-align:center">
        <div style="font-size:22px">${{red:'🔴',amber:'🟡',green:'🟢'}[op.rag]}</div>
        <div style="color:${RAG_CLR[op.rag]};font-weight:700">${RAG_LBL[op.rag]}</div>
      </div>
      <div style="background:#f0f4f8;border-radius:6px;padding:10px">
        <div style="font-size:11px;color:#888">Duration (median / fleet ratio)</div>
        <div style="font-size:20px;font-weight:700">${dur.median??'—'} min <span style="font-size:13px;color:#666">(${dur.ratio_to_median!=null?(dur.ratio_to_median*100).toFixed(0)+'%':'—'} of fleet)</span></div>
        <div style="font-size:11px;color:#888;margin-top:4px">P10: ${dur.p10??'—'} | P90: ${dur.p90??'—'} | CV: ${dur.cv??'—'}</div>
      </div>
      <div style="background:#f0f4f8;border-radius:6px;padding:10px">
        <div style="font-size:11px;color:#888">Fast interviews / Replacements</div>
        <div style="font-size:18px;font-weight:700">${dur.pct_fast??0}% &lt;5min &nbsp;|&nbsp; ${(op.replacements||{}).pct??0}% replaced</div>
        <div style="font-size:11px;color:#888;margin-top:4px">Late-night: ${tf.n_late_night||0} | Call attempts (mean): ${(op.call_attempts||{}).mean??'—'}</div>
      </div>
    </div>
    ${op.flags&&op.flags.length>0 ? `<div style="margin-bottom:14px"><strong style="font-size:12.5px">Active flags:</strong><br>${op.flags.map(f=>`<span style="display:inline-block;background:#fde8e8;color:#c0392b;border-radius:3px;padding:2px 8px;margin:2px;font-size:11.5px">⚠ ${f}</span>`).join('')}</div>` : ''}
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px">
      <div>
        <strong style="font-size:12px">Module duration z-scores</strong>
        <p style="font-size:10.5px;color:#888;margin:2px 0 6px">z &lt; −1.5 = unusually fast for this module</p>
        <table style="width:100%;border-collapse:collapse;font-size:11.5px">
          <thead><tr style="background:#e8ecf0;font-size:11px">
            <th style="padding:4px 10px;text-align:left">Module</th>
            <th style="padding:4px 10px;text-align:center">Mean (min)</th>
            <th style="padding:4px 10px;text-align:center">z-score</th>
          </tr></thead><tbody>${durRows||'<tr><td colspan="3" style="padding:6px;color:#888">No data</td></tr>'}</tbody>
        </table>
      </div>
      <div>
        <strong style="font-size:12px">Short-circuit gate rates</strong>
        <p style="font-size:10.5px;color:#888;margin:2px 0 6px">High rate vs fleet → possible survey shortcutting</p>
        <table style="width:100%;border-collapse:collapse;font-size:11.5px">
          <thead><tr style="background:#e8ecf0;font-size:11px">
            <th style="padding:4px 10px;text-align:left">Gate</th>
            <th style="padding:4px 10px;text-align:center">Operator</th>
            <th style="padding:4px 10px;text-align:center">Fleet avg</th>
            <th style="padding:4px 10px;text-align:center">Diff</th>
          </tr></thead><tbody>${scRows||'<tr><td colspan="4" style="padding:6px;color:#888">No data</td></tr>'}</tbody>
        </table>
        <div style="margin-top:10px"><strong style="font-size:12px">Skip violations by module</strong><br>${skipDetail}</div>
      </div>
      <div>
        <strong style="font-size:12px">Attrition by transition</strong>
        <p style="font-size:10.5px;color:#888;margin:2px 0 6px">% of their R_prev HHs that dropped in R_curr</p>
        <table style="width:100%;border-collapse:collapse;font-size:11.5px">
          <thead><tr style="background:#e8ecf0;font-size:11px">
            <th style="padding:4px 10px">Transition</th>
            <th style="padding:4px 10px;text-align:center">N</th>
            <th style="padding:4px 10px;text-align:center">Drop %</th>
          </tr></thead><tbody>${attrRows||'<tr><td colspan="3" style="padding:6px;color:#888">No data</td></tr>'}</tbody>
        </table>
      </div>
    </div>`;

  document.getElementById('ops-detail-body').innerHTML = body;
  const detailEl = document.getElementById('ops-detail');
  detailEl.style.display = 'block';
  detailEl.scrollIntoView({behavior:'smooth', block:'nearest'});
}

function buildPanel(){
  if(!PAN || !PAN.attrition) return;
  const p = PAN;
  const rounds = [1,2,3,4,5,6,7];
  const rLabels = rounds.map(r=>'R'+r);

  // ── Panel Health Scorecard ──
  const scEl = document.getElementById('panel-scorecard');
  if(scEl){
    const rc = p.round_counts||{};
    const totalHH = p.all_hhs||0;
    const alwaysIn = p.always_in||0;
    const alwaysPct = totalHH>0 ? (alwaysIn/totalHH*100).toFixed(1) : 0;
    const overallAttr = p.attrition||[];
    const retRates = overallAttr.map(a => a.n>0 ? (a.retained/(a.retained+a.dropped)*100) : 0);
    const avgRet = retRates.length>0 ? (retRates.reduce((s,v)=>s+v,0)/retRates.length).toFixed(1) : '—';
    const callViols = (p.call_interval_summary||[]).reduce((s,c)=>s+(c.n_under||0),0);
    const nPsuProblems = (p.psu_problem_list||[]).length;
    const biasVerdict = (p.attrition_bias||{}).verdict||'—';
    const biasCls = biasVerdict==='HIGH'?'red':biasVerdict==='MODERATE'?'yellow':'green';

    // R1 baseline survivors in R5
    const r1Count = parseInt(rc['1'])||0;
    const r5Count = parseInt(rc['5'])||0;

    scEl.innerHTML = `
      <div class="stat-box blue"><div class="num">${totalHH.toLocaleString()}</div><div class="lbl">Unique HHs (all rounds)</div></div>
      <div class="stat-box ${alwaysPct<30?'red':alwaysPct<50?'yellow':'green'}"><div class="num">${alwaysIn}</div><div class="lbl">All 5 rounds (${alwaysPct}%)</div></div>
      <div class="stat-box ${parseFloat(avgRet)<70?'red':parseFloat(avgRet)<80?'yellow':'green'}"><div class="num">${avgRet}%</div><div class="lbl">Avg retention rate</div></div>
      <div class="stat-box ${biasCls}"><div class="num">${biasVerdict}</div><div class="lbl">Selection bias</div></div>
      <div class="stat-box ${nPsuProblems>150?'red':nPsuProblems>50?'yellow':'green'}"><div class="num">${nPsuProblems}</div><div class="lbl">Under-target PSUs</div></div>
      <div class="stat-box ${callViols>50?'red':callViols>10?'yellow':'green'}"><div class="num">${callViols}</div><div class="lbl">Call interval violations</div></div>`;
  }

  // ── Sample Funnel Chart ──
  const funnelEl = document.getElementById('panelFunnelChart');
  if(funnelEl){
    const rc = p.round_counts||{};
    const attrProf = p.attrition_profile||[];
    // Compute R1 survivors through each round
    const r1Start = parseInt(rc['1'])||0;
    let r1Survivors = [r1Start];
    for(let i=0;i<attrProf.length;i++){
      const prev = r1Survivors[r1Survivors.length-1];
      const retRate = attrProf[i].retained.n / (attrProf[i].retained.n + attrProf[i].dropped.n);
      r1Survivors.push(Math.round(prev * retRate));
    }
    const newEntries = rounds.map((r,i)=>{
      if(i===0) return 0;
      const total = parseInt(rc[String(r)])||0;
      return Math.max(0, total - r1Survivors[i]);
    });

    new Chart(funnelEl, {
      type:'bar',
      data:{
        labels: rLabels,
        datasets:[
          { label:'R1 baseline survivors', data:r1Survivors, backgroundColor:'#3498db', borderRadius:3 },
          { label:'New entries', data:newEntries, backgroundColor:'#e67e22', borderRadius:3 }
        ]
      },
      options:{
        responsive:true, maintainAspectRatio:false,
        plugins:{
          legend:{position:'bottom',labels:{font:{size:10}}},
          tooltip:{callbacks:{afterBody:(items)=>{
            const idx=items[0].dataIndex;
            const total = r1Survivors[idx]+newEntries[idx];
            const survPct = r1Start>0 ? (r1Survivors[idx]/r1Start*100).toFixed(1) : 0;
            return `Total: ${total} HHs\\nR1 survivor rate: ${survPct}%`;
          }}}
        },
        scales:{
          x:{stacked:true,ticks:{font:{size:11}}},
          y:{stacked:true,title:{display:true,text:'Households',font:{size:10}},ticks:{font:{size:10}}}
        }
      }
    });
  }

  // ── Retention Rate Chart ──
  const retEl = document.getElementById('panelRetentionChart');
  if(retEl){
    const attrProf = p.attrition_profile||[];
    const labels = attrProf.map(a=>a.label);
    const retRates = attrProf.map(a=>{
      const total = a.retained.n + a.dropped.n;
      return total>0 ? parseFloat((a.retained.n/total*100).toFixed(1)) : 0;
    });
    const callData = (p.call_interval_summary||[]).map(c=>c.pct_under||0);
    // Align call data with transitions: R2→R3 uses R3 call data, etc.
    const callAligned = attrProf.map(a=>{
      const toR = a.to_round;
      const ci = (p.call_interval_summary||[]).find(c=>c.round===toR);
      return ci ? ci.pct_under : 0;
    });

    new Chart(retEl, {
      type:'line',
      data:{
        labels,
        datasets:[
          { label:'Retention %', data:retRates, borderColor:'#3498db', backgroundColor:'rgba(52,152,219,0.1)',
            fill:true, tension:0.3, pointRadius:5, pointBackgroundColor:retRates.map(v=>v<70?'#e74c3c':v<80?'#e67e22':'#27ae60'), borderWidth:2 },
          { label:'Call violations %', data:callAligned, borderColor:'#e74c3c', borderDash:[5,3],
            fill:false, tension:0.3, pointRadius:4, borderWidth:1.5, yAxisID:'y1' }
        ]
      },
      options:{
        responsive:true, maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:10}}}},
        scales:{
          y:{min:50,max:100,title:{display:true,text:'Retention %',font:{size:10}},ticks:{font:{size:10},callback:v=>v+'%'}},
          y1:{position:'right',min:0,max:10,title:{display:true,text:'Call violation %',font:{size:10}},ticks:{font:{size:10},callback:v=>v+'%'},grid:{drawOnChartArea:false}}
        }
      }
    });
  }

  // ── Attrition Bias Analysis ──
  const bias = p.attrition_bias;
  if(bias){
    // ── Banner ──────────────────────────────────────────────────────────────
    const bannerEl = document.getElementById('panel-bias-banner');
    if(bannerEl){
      const colors = {HIGH:'#e74c3c', MODERATE:'#e67e22', LOW:'#27ae60'};
      const icons  = {HIGH:'🔴', MODERATE:'🟡', LOW:'🟢'};
      const msgs   = {
        HIGH:     `HIGH — Significant selection bias in ${bias.n_biased_trans} of ${(bias.transitions||[]).length} round-to-round transitions. Households that stay are systematically different from those that leave in multiple waves.`,
        MODERATE: `MODERATE — Differences detected in ${bias.n_biased_trans} transition(s). Panel may be selectively retaining certain household types.`,
        LOW:      'LOW — No significant differences detected between retained and dropped households in any transition.',
      };
      const wvars = (bias.worst_vars||[]);
      bannerEl.innerHTML=`<div style="background:${colors[bias.verdict]}18;border-left:4px solid ${colors[bias.verdict]};padding:12px 16px;border-radius:4px;margin-bottom:12px">
        <strong style="color:${colors[bias.verdict]};font-size:13px">${icons[bias.verdict]} Selection Bias Verdict: ${bias.verdict}</strong>
        <p style="margin:4px 0 0 0;font-size:12.5px;color:#333">${msgs[bias.verdict]}</p>
        ${wvars.length ? `<p style="margin:4px 0 0 0;font-size:12px;color:#555">Variables significant in ≥1 transition: <strong>${wvars.join(', ')}</strong></p>` : ''}
      </div>`;
    }

    // ── Badge ────────────────────────────────────────────────────────────────
    const badgeEl = document.getElementById('panel-bias-verdict-badge');
    if(badgeEl){
      const bc={HIGH:'badge-red',MODERATE:'badge-yellow',LOW:'badge-green'};
      badgeEl.className=`badge ${bc[bias.verdict]||'badge-blue'}`;
      badgeEl.textContent=bias.verdict+' CONCERN';
    }
    // Populate collapse summary
    const biasSumEl = document.getElementById('panel-bias-summary');
    if(biasSumEl){
      const bc3={HIGH:'#e74c3c',MODERATE:'#e67e22',LOW:'#27ae60'};
      biasSumEl.innerHTML=`<span style="color:${bc3[bias.verdict]||'#888'}">${bias.verdict}</span>
        <span style="color:#888">${bias.n_biased_trans||0} of ${(bias.transitions||[]).length} transitions biased</span>`;
    }

    // ── Per-transition comparison table (tabbed) ──────────────────────────
    const biasTableEl = document.getElementById('panel-bias-table');
    const bTrans = bias.transitions||[];
    if(biasTableEl && bTrans.length){
      const VAR_LABELS = {
        urban:'% Urban', hhsize:'HH Size (mean)', female:'% Female head',
        employed:'% Employed', has_account:'% Has bank account',
        has_savings:'% Has savings',
      };
      // Tab buttons
      let tabHtml=`<div style="display:flex;gap:6px;margin-bottom:10px;flex-wrap:wrap">`;
      bTrans.forEach((t,idx)=>{
        const bc2=t.verdict==='BIASED'?'#e74c3c':'#27ae60';
        const isFirst=(idx===0);
        tabHtml+=`<button onclick="showBiasTab(${idx})" id="bias-tab-btn-${idx}"
          style="padding:5px 13px;border:2px solid ${bc2};border-radius:4px;
                 background:${isFirst?bc2:'transparent'};color:${isFirst?'#fff':bc2};
                 cursor:pointer;font-size:12px;font-weight:600">
          ${t.label} ${t.verdict==='BIASED'?'⚠️':'✓'}
        </button>`;
      });
      tabHtml+=`</div>`;

      // Per-transition panels
      let panelsHtml='';
      bTrans.forEach((t,idx)=>{
        const vars=t.vars||{};
        const varKeys=['urban','hhsize','female','employed','has_account'];
        panelsHtml+=`<div id="bias-panel-${idx}" style="display:${idx===0?'block':'none'}">
          <p style="font-size:11.5px;color:#666;margin:0 0 8px">
            Comparing <strong>${t.n_retained}</strong> retained vs
            <strong>${t.n_dropped}</strong> dropped households
            (n prev round = ${t.n_prev}) — characteristics from R${t.from_round}.
          </p>
          <div style="overflow-x:auto">
          <table style="width:100%;border-collapse:collapse;font-size:12.5px;min-width:540px">
          <thead><tr style="background:#1a2332;color:#fff">
            <th style="padding:7px 10px;text-align:left">Variable</th>
            <th style="padding:7px 10px;text-align:center">Retained</th>
            <th style="padding:7px 10px;text-align:center">Dropped</th>
            <th style="padding:7px 10px;text-align:center">Diff</th>
            <th style="padding:7px 10px;text-align:center">Test stat</th>
            <th style="padding:7px 10px;text-align:center">p-value</th>
            <th style="padding:7px 10px;text-align:center">Sig.</th>
          </tr></thead><tbody>`;

        varKeys.forEach((vk,ri)=>{
          const v=vars[vk]; if(!v) return;
          const isSig=(v.p<0.05);
          const rowBg=isSig?'#fff3cd':(ri%2===0?'#f8f9fa':'#fff');
          const sigClr=v.p<0.05?'#e74c3c':v.p<0.10?'#e67e22':'#555';
          const isCont=(vk==='hhsize');
          const rDisp=isCont?v.retained:(v.retained+'%');
          const dDisp=isCont?v.dropped:(v.dropped+'%');
          const rawDiff=v.diff;
          const diffSign=rawDiff>0?'+':'';
          const diffUnit=isCont?'':' pp';
          const diffClr=Math.abs(rawDiff)>(isCont?0.3:5)?'#e74c3c':'#555';
          const statDisp=isCont?`t=${v.t}`:`χ²=${v.chi2}`;
          panelsHtml+=`<tr style="background:${rowBg}">
            <td style="padding:7px 10px;font-weight:600">${VAR_LABELS[vk]||vk}${isSig?' ⚠️':''}</td>
            <td style="padding:7px 10px;text-align:center">${rDisp}</td>
            <td style="padding:7px 10px;text-align:center">${dDisp}</td>
            <td style="padding:7px 10px;text-align:center;color:${diffClr}">${diffSign}${rawDiff}${diffUnit}</td>
            <td style="padding:7px 10px;text-align:center;font-family:monospace;font-size:11px">${statDisp}</td>
            <td style="padding:7px 10px;text-align:center;font-family:monospace;font-size:11px;color:${sigClr}">${v.p}</td>
            <td style="padding:7px 10px;text-align:center;font-weight:700;color:${sigClr}">${v.sig}</td>
          </tr>`;
        });

        // Region row
        const reg=vars['region'];
        if(reg){
          const isSig=(reg.p<0.05);
          const rowBg=isSig?'#fff3cd':'#f8f9fa';
          const sigClr=reg.p<0.05?'#e74c3c':reg.p<0.10?'#e67e22':'#555';
          panelsHtml+=`<tr style="background:${rowBg}">
            <td style="padding:7px 10px;font-weight:600">Region (composition)${isSig?' ⚠️':''}</td>
            <td style="padding:7px 10px;text-align:center;color:#888" colspan="3">(see chart & heatmap below)</td>
            <td style="padding:7px 10px;text-align:center;font-family:monospace;font-size:11px">χ²=${reg.chi2}, V=${reg.v}</td>
            <td style="padding:7px 10px;text-align:center;font-family:monospace;font-size:11px;color:${sigClr}">${reg.p}</td>
            <td style="padding:7px 10px;text-align:center;font-weight:700;color:${sigClr}">${reg.sig}</td>
          </tr>`;
        }

        panelsHtml+=`</tbody></table></div>
          <p style="font-size:11px;color:#888;margin-top:6px">*** p&lt;0.001 &nbsp;** p&lt;0.01 &nbsp;* p&lt;0.05 &nbsp;† p&lt;0.10 &nbsp;ns=not significant</p>
        </div>`;
      });

      biasTableEl.innerHTML=tabHtml+panelsHtml;

      // Tab switching + chart update
      let _biasRegChartInst=null;
      window._biasUpdateRegChart=function(idx){
        const t=bTrans[idx]; if(!t) return;
        const rr=(t.reg_ret_rates||[]).slice().sort((a,b)=>(a.pct_retained||0)-(b.pct_retained||0));
        const labels=rr.map(r=>r.name);
        const vals=rr.map(r=>r.pct_retained||0);
        const colors2=vals.map(v=>v<50?'#e74c3c':v<65?'#e67e22':v<75?'#f39c12':'#27ae60');
        const lbl=document.getElementById('bias-reg-chart-label');
        if(lbl) lbl.textContent=`(${t.label})`;
        if(_biasRegChartInst){ _biasRegChartInst.destroy(); _biasRegChartInst=null; }
        const canvas=document.getElementById('panelBiasRegChart');
        if(!canvas) return;
        _biasRegChartInst=new Chart(canvas,{
          type:'bar',
          data:{
            labels:labels,
            datasets:[{
              label:`% Retained ${t.label}`,
              data:vals,
              backgroundColor:colors2,
              borderRadius:3,
            }]
          },
          options:{
            indexAxis:'y',
            responsive:true, maintainAspectRatio:false,
            plugins:{
              legend:{display:false},
              tooltip:{callbacks:{label:c=>{
                const r=rr[c.dataIndex];
                return `${c.raw}% retained (n=${r.n}, ret=${r.retained}, drop=${r.dropped})`;
              }}}
            },
            scales:{
              x:{min:0,max:100,title:{display:true,text:'% Retained',font:{size:11}},ticks:{callback:v=>v+'%'}},
              y:{ticks:{font:{size:10.5}}}
            }
          }
        });
      };

      window.showBiasTab=function(idx){
        bTrans.forEach((_,i)=>{
          const panel=document.getElementById('bias-panel-'+i);
          const btn=document.getElementById('bias-tab-btn-'+i);
          if(panel) panel.style.display=(i===idx)?'block':'none';
          if(btn){
            const bc2=bTrans[i].verdict==='BIASED'?'#e74c3c':'#27ae60';
            btn.style.background=(i===idx)?bc2:'transparent';
            btn.style.color=(i===idx)?'#fff':bc2;
          }
        });
        window._biasUpdateRegChart(idx);
      };
      window._biasUpdateRegChart(0);
    }

    // ── Composition drift chart (unchanged — uses bias.comp_drift) ─────────
    const driftData = [...(bias.comp_drift||[])].sort((a,b)=>a.drift-b.drift);
    const driftColors = driftData.map(r=>r.drift<0?'#e74c3c':'#27ae60');
    new Chart(document.getElementById('panelBiasDriftChart'),{
      type:'bar',
      data:{
        labels: driftData.map(r=>r.region_name),
        datasets:[{
          label:'Share change (pp)',
          data: driftData.map(r=>r.drift),
          backgroundColor: driftColors,
          borderRadius:3,
        }]
      },
      options:{
        indexAxis:'y',
        responsive:true, maintainAspectRatio:false,
        plugins:{
          legend:{display:false},
          tooltip:{callbacks:{label:c=>{
            const r=driftData[c.dataIndex];
            const rounds=Object.keys(r.pct_per_round||{}).sort();
            const r1=r.pct_per_round['1']||0;
            const rLast=r.pct_per_round[rounds[rounds.length-1]]||0;
            return `${c.raw>0?'+':''}${c.raw}pp  (R1=${r1}%  → R${rounds[rounds.length-1]}=${rLast}%)`;
          }}}
        },
        scales:{
          x:{title:{display:true,text:'Share change (pp)',font:{size:11}}},
          y:{ticks:{font:{size:10.5}}}
        }
      }
    });

    // ── Retention heatmap: Region × Transition ────────────────────────────
    const hmEl = document.getElementById('panel-trans-heatmap');
    if(hmEl && bTrans.length){
      // Collect all regions across all transitions
      const regKeySet=new Set();
      const regNameMap={};
      bTrans.forEach(t=>(t.reg_ret_rates||[]).forEach(r=>{
        const k=String(r.region);
        regKeySet.add(k);
        regNameMap[k]=r.name;
      }));
      const allRegKeys=[...regKeySet].sort((a,b)=>parseInt(a)-parseInt(b));

      let html=`<div style="overflow-x:auto"><table style="border-collapse:collapse;font-size:11.5px;min-width:480px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:6px 10px;text-align:left">Region</th>
          ${bTrans.map(t=>{
            const bc2=t.verdict==='BIASED'?'#e74c3c50':'#27ae6030';
            return `<th style="padding:6px 8px;text-align:center;background:${bc2}">${t.label}${t.verdict==='BIASED'?' ⚠️':''}</th>`;
          }).join('')}
        </tr></thead><tbody>`;

      allRegKeys.forEach((rk,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        html+=`<tr style="background:${bg}">
          <td style="padding:5px 10px;font-weight:600">${regNameMap[rk]||rk}</td>`;
        bTrans.forEach(t=>{
          const rd=(t.reg_ret_rates||[]).find(r=>String(r.region)===rk);
          const pct=rd?rd.pct_retained:null;
          const cellBg=pct===null?'#f0f0f0':pct<50?'#fde8e8':pct<65?'#fff3cd':pct<75?'#fef9e7':'#d4edda';
          const textClr=pct===null?'#aaa':pct<50?'#c0392b':pct<65?'#856404':'#155724';
          const tip=rd?`title="${rd.retained} ret / ${rd.dropped} drop (n=${rd.n})"` :'';
          html+=`<td style="padding:5px 8px;text-align:center;background:${cellBg};color:${textClr};font-weight:600" ${tip}>
            ${pct!==null?pct+'%':'—'}</td>`;
        });
        html+=`</tr>`;
      });
      html+=`</tbody></table></div>`;
      hmEl.innerHTML=html;
    }
  }

  // ── Stats row ──
  const statsEl = document.getElementById('panel-stats');
  if(statsEl){
    // Latest round = last entry in the attrition array (auto-scales with ROUNDS)
    const lastIdx = p.attrition.length - 1;
    const lastR   = ROUNDS[ROUNDS.length - 1];
    const retLast = p.attrition[lastIdx];
    const retPct  = retLast ? Math.round(retLast.retained/p.attrition[0].n*100) : '?';
    statsEl.innerHTML = [
      {n:p.all_hhs,    lbl:'Unique HHs (all rounds)', cls:'purple'},
      {n:p.attrition[0].n, lbl:'R1 Baseline', cls:'blue'},
      {n:p.always_in,  lbl:`Present in all ${ROUNDS.length} rounds`, cls:'green'},
      {n:p.r1_only,    lbl:'R1 only (never seen again)', cls:'yellow'},
      {n:p.never_r1,   lbl:'Never in R1 (new entries)', cls:''},
      {n:retPct+'%',   lbl:`R1 HHs retained by R${lastR}`, cls:retPct<60?'red':'green'},
    ].map(s=>`<div class="stat-box ${s.cls}"><div class="num">${s.n}</div><div class="lbl">${s.lbl}</div></div>`).join('');
  }

  // ── Attrition note ──
  const noteEl = document.getElementById('panel-attrition-note');
  if(noteEl){
    const lastIdx2 = p.attrition.length - 1;
    const lastR2   = ROUNDS[ROUNDS.length - 1];
    const aLast    = p.attrition[lastIdx2];
    noteEl.innerHTML = `<strong>R${lastR2} retention:</strong> ${aLast.retained} of ${p.attrition[0].n} R1 baseline households (${Math.round(aLast.retained/p.attrition[0].n*100)}%) were re-interviewed in Round ${lastR2}. ${aLast.new_in} households appear in R${lastR2} that were not in the R1 sample.`;
  }

  // ── Attrition stacked bar ──
  const retArr  = p.attrition.map(r=>r.retained);
  const newArr  = p.attrition.map(r=>r.new_in);
  const dropArr = p.attrition.map(r=>r.dropped);
  new Chart(document.getElementById('panelAttrChart'),{
    type:'bar',
    data:{
      labels: rLabels,
      datasets:[
        {label:'Retained from R1', data:retArr,  backgroundColor:'#2980b9'},
        {label:'New (not in R1)',   data:newArr,  backgroundColor:'#27ae60'},
      ]
    },
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{font:{size:11}}}},
      scales:{x:{stacked:true},y:{stacked:true,beginAtZero:true,
        title:{display:true,text:'Households',font:{size:11}}}}}
  });

  // ── Pattern distribution bar ──
  const topPat = p.pattern_dist.slice(0,10);
  new Chart(document.getElementById('panelPatChart'),{
    type:'bar',
    data:{
      labels: topPat.map(x=>x.pattern),
      datasets:[{label:'HHs', data:topPat.map(x=>x.n),
        backgroundColor: topPat.map(x=>{
          const ones = (x.pattern.match(/1/g)||[]).length;
          return ones===5?'#27ae60':ones>=3?'#2980b9':ones===1?'#e67e22':'#e74c3c';
        })}]
    },
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{display:false},
        tooltip:{callbacks:{label:c=>`${c.raw} HHs`}}},
      scales:{x:{ticks:{font:{size:10}}},
              y:{beginAtZero:true,title:{display:true,text:'Households',font:{size:11}}}}}
  });

  // ── Household panel tracker ───────────────────────────────────────────────
  if(p.hh_matrix && p.hh_matrix.length){
    const HHM = p.hh_matrix;
    const PAGE_SIZE = 50;
    const STATUS_COLORS = {
      'All Rounds':       {bg:'#d4edda', txt:'#155724'},
      'Left Panel':       {bg:'#fde8e8', txt:'#7b1c1c'},
      'New Entry':        {bg:'#cce5ff', txt:'#004085'},
      'New Entry → Left': {bg:'#fff3cd', txt:'#856404'},
      'Intermittent':     {bg:'#e2e3e5', txt:'#383d41'},
    };
    const ROUNDS_DISP = [1,2,3,4,5,6,7,8];

    // Badge
    const matBadge = document.getElementById('hh-matrix-badge');
    if(matBadge) matBadge.textContent = `${HHM.length} households`;

    // Populate region dropdown
    const regSel = document.getElementById('hh-region-filter');
    if(regSel){
      const regSet = {};
      HHM.forEach(h=>{if(h.region) regSet[h.region]=h.region_name;});
      Object.keys(regSet).sort((a,b)=>parseInt(a)-parseInt(b)).forEach(k=>{
        const opt=document.createElement('option');
        opt.value=k; opt.textContent=regSet[k];
        regSel.appendChild(opt);
      });
      regSel.addEventListener('change',()=>{_hhState.region=regSel.value;_hhState.page=0;renderHH();});
    }

    // Status filter buttons
    const statusFilterEl = document.getElementById('hh-status-filter');
    const STATUS_LIST = ['All Rounds','Left Panel','New Entry','New Entry → Left','Intermittent'];
    if(statusFilterEl){
      const counts = {};
      STATUS_LIST.forEach(s=>{counts[s]=HHM.filter(h=>h.status===s).length;});
      [['All',''],...STATUS_LIST.map(s=>[s,s])].forEach(([label,val])=>{
        const n = val===''?HHM.length:counts[val]||0;
        const btn=document.createElement('button');
        btn.textContent=`${label} (${n})`;
        btn.className='round-btn'+(val===''?' active':'');
        btn.style.fontSize='11px';
        btn.addEventListener('click',()=>{
          statusFilterEl.querySelectorAll('button').forEach(b=>b.classList.remove('active'));
          btn.classList.add('active');
          _hhState.status=val; _hhState.page=0; renderHH();
        });
        statusFilterEl.appendChild(btn);
      });
    }

    // Urban/Rural filter
    const urbanFilterEl = document.getElementById('hh-urban-filter');
    if(urbanFilterEl){
      [['All',''],['Urban','1'],['Rural','2']].forEach(([label,val])=>{
        const btn=document.createElement('button');
        btn.textContent=label; btn.className='round-btn'+(val===''?' active':'');
        btn.style.fontSize='11px';
        btn.addEventListener('click',()=>{
          urbanFilterEl.querySelectorAll('button').forEach(b=>b.classList.remove('active'));
          btn.classList.add('active');
          _hhState.urban=val; _hhState.page=0; renderHH();
        });
        urbanFilterEl.appendChild(btn);
      });
    }

    // Search box
    const searchEl = document.getElementById('hh-search');
    if(searchEl) searchEl.addEventListener('input',()=>{_hhState.query=searchEl.value.trim().toLowerCase();_hhState.page=0;renderHH();});

    // State
    const _hhState = window._hhState = {page:0, status:'', region:'', urban:'', query:''};

    function getFiltered(){
      return HHM.filter(h=>{
        if(_hhState.status && h.status!==_hhState.status) return false;
        if(_hhState.region && String(h.region)!==_hhState.region) return false;
        if(_hhState.urban && String(h.urban)!==_hhState.urban) return false;
        if(_hhState.query){
          const q=_hhState.query;
          if(!String(h.hhid).includes(q) && !h.psu.includes(q)) return false;
        }
        return true;
      });
    }

    window.renderHH = renderHH;
    function renderHH(){
      const filtered = getFiltered();
      const total = filtered.length;
      const start = _hhState.page * PAGE_SIZE;
      const page  = filtered.slice(start, start+PAGE_SIZE);
      const tableEl  = document.getElementById('hh-matrix-table');
      const pageEl   = document.getElementById('hh-matrix-pagination');

      // Table
      let html=`<p style="font-size:11.5px;color:#555;margin:0 0 6px">Showing ${Math.min(start+1,total)}–${Math.min(start+PAGE_SIZE,total)} of ${total} households</p>
        <div style="overflow-x:auto">
        <table style="width:100%;border-collapse:collapse;font-size:12px;min-width:640px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:6px 9px;text-align:left;cursor:pointer">HHID</th>
          <th style="padding:6px 9px;text-align:left">PSU</th>
          <th style="padding:6px 9px;text-align:left">Region</th>
          <th style="padding:6px 8px;text-align:center">Type</th>
          ${ROUNDS_DISP.map(r=>`<th style="padding:6px 8px;text-align:center;min-width:48px">R${r}${r>1?'<br><span style="font-weight:normal;font-size:9px;opacity:0.8">days</span>':''}</th>`).join('')}
          <th style="padding:6px 9px;text-align:center">Rounds</th>
          <th style="padding:6px 9px;text-align:center">Status</th>
        </tr></thead><tbody>`;

      page.forEach((h,i)=>{
        const rowBg = i%2===0?'#f8f9fa':'#fff';
        const uTag = h.urban===1
          ?`<span style="background:#cce5ff;color:#004085;border-radius:3px;padding:1px 5px;font-size:10px">Urban</span>`
          :`<span style="background:#d4edda;color:#155724;border-radius:3px;padding:1px 5px;font-size:10px">Rural</span>`;
        const sc = STATUS_COLORS[h.status]||{bg:'#eee',txt:'#333'};
        const dg = h.days_gap||{};
        const roundCells = ROUNDS_DISP.map(r=>{
          const present = h.presence[String(r)]===1;
          const days = dg[String(r)]!=null ? dg[String(r)] : null;
          // Cell colour when present: driven by days gap (R1 has no gap)
          let cellBg, cellTxt;
          if(!present){
            cellBg='#fde8e8'; cellTxt='#c0392b';
          } else if(r===1 || days===null){
            cellBg='#d4edda'; cellTxt='#155724';       // first appearance
          } else if(days < 20){
            cellBg='#e74c3c'; cellTxt='#fff';           // violation — red
          } else if(days < 30){
            cellBg='#fff3cd'; cellTxt='#856404';        // tight (20–29d, needs approval) — amber
          } else {
            cellBg='#d4edda'; cellTxt='#155724';        // normal — green
          }
          const daysLabel = (r>1 && days!==null)
            ? `<div style="font-size:9.5px;font-weight:400;margin-top:1px;opacity:0.85">${days}d</div>`
            : (r===1 && present ? `<div style="font-size:9px;opacity:0.6;margin-top:1px">—</div>` : '');
          const tick = present ? '✓' : '✗';
          return `<td style="padding:3px 5px;text-align:center;background:${cellBg};color:${cellTxt};font-weight:700;font-size:13px;line-height:1.2">
            ${tick}${daysLabel}</td>`;
        }).join('');
        html+=`<tr style="background:${rowBg}">
          <td style="padding:5px 9px;font-family:monospace;font-size:11px;font-weight:600">${h.hhid}</td>
          <td style="padding:5px 9px;font-family:monospace;font-size:10.5px;color:#555">${h.psu}</td>
          <td style="padding:5px 9px;font-size:11.5px">${h.region_name}</td>
          <td style="padding:5px 8px;text-align:center">${uTag}</td>
          ${roundCells}
          <td style="padding:5px 9px;text-align:center;font-weight:700">${h.rounds_present}/${ROUNDS.length}</td>
          <td style="padding:5px 9px;text-align:center">
            <span style="background:${sc.bg};color:${sc.txt};border-radius:3px;padding:2px 7px;font-size:10.5px;font-weight:600;white-space:nowrap"
              title="First: R${h.first_round||'?'}  Last: R${h.last_round||'?'}">${h.status}</span>
          </td>
        </tr>`;
      });
      html+=`</tbody></table></div>
        <div style="display:flex;gap:14px;flex-wrap:wrap;margin-top:7px;font-size:11px;align-items:center">
          <span style="color:#888">Days since previous interview:</span>
          <span><span style="background:#d4edda;color:#155724;padding:1px 7px;border-radius:3px;font-weight:600">✓ —</span> First appearance</span>
          <span><span style="background:#d4edda;color:#155724;padding:1px 7px;border-radius:3px;font-weight:600">✓ 30d+</span> Normal</span>
          <span><span style="background:#fff3cd;color:#856404;padding:1px 7px;border-radius:3px;font-weight:600">✓ 20–29d</span> Needs approval</span>
          <span><span style="background:#e74c3c;color:#fff;padding:1px 7px;border-radius:3px;font-weight:600">✓ &lt;20d</span> Violation</span>
          <span><span style="background:#fde8e8;color:#c0392b;padding:1px 7px;border-radius:3px;font-weight:600">✗</span> Absent</span>
          <span style="color:#aaa">| Hover status badge for first/last round</span>
        </div>`;
      if(tableEl) tableEl.innerHTML=html;

      // Pagination
      const totalPages = Math.ceil(total/PAGE_SIZE);
      let pHtml='';
      if(totalPages>1){
        pHtml+=`<span style="font-size:12px;color:#555">Page ${_hhState.page+1} of ${totalPages}</span>`;
        if(_hhState.page>0)
          pHtml+=`<button class="round-btn" style="font-size:11px" onclick="_hhState.page--;renderHH()">◀ Prev</button>`;
        // Show up to 7 page buttons around current
        const pStart=Math.max(0,_hhState.page-3);
        const pEnd=Math.min(totalPages-1,_hhState.page+3);
        for(let pg=pStart;pg<=pEnd;pg++){
          pHtml+=`<button class="round-btn${pg===_hhState.page?' active':''}" style="font-size:11px;min-width:30px"
            onclick="_hhState.page=${pg};renderHH()">${pg+1}</button>`;
        }
        if(_hhState.page<totalPages-1)
          pHtml+=`<button class="round-btn" style="font-size:11px" onclick="_hhState.page++;renderHH()">Next ▶</button>`;
      }
      if(pageEl) pageEl.innerHTML=pHtml;
    }

    // CSV download
    window._hhDownloadCsv=function(){
      const filtered=getFiltered();
      const header=['hhid','psu','region','region_name','urban_label','R1','R2','R3','R4','R5','R6','R7','R8','days_to_R2','days_to_R3','days_to_R4','days_to_R5','days_to_R6','days_to_R7','days_to_R8','rounds_present','first_round','last_round','status','pattern'];
      const rows=filtered.map(h=>{
        const dg=h.days_gap||{};
        return [
          h.hhid, h.psu, h.region, h.region_name, h.urban_label,
          h.presence['1'], h.presence['2'], h.presence['3'], h.presence['4'], h.presence['5'], h.presence['6'], h.presence['7'], h.presence['8'],
          dg['2']!=null?dg['2']:'', dg['3']!=null?dg['3']:'', dg['4']!=null?dg['4']:'', dg['5']!=null?dg['5']:'',
          dg['6']!=null?dg['6']:'', dg['7']!=null?dg['7']:'', dg['8']!=null?dg['8']:'',
          h.rounds_present, h.first_round||'', h.last_round||'', h.status, h.pattern
        ];
      });
      const csv=[header,...rows].map(r=>r.join(',')).join('\\n');
      const a=document.createElement('a');
      a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
      a.download='l2phl_panel_households.csv';
      a.click();
    };

    renderHH();
  }

  // ── PSU status table ──
  const psuEl = document.getElementById('panel-psu-status');
  if(psuEl && p.psu_status){
    const tgt = `Urban PSUs: target ${p.urban_target} HH &nbsp;|&nbsp; Rural PSUs: target ${p.rural_target} HH`;
    let html = `<p class="note-box note-info" style="margin-bottom:10px;font-size:12px">${tgt}</p>`;
    html += `<table style="width:100%;border-collapse:collapse;font-size:12.5px">
      <thead><tr style="background:#1a2332;color:#fff">
        <th style="padding:7px 10px;text-align:left">Round</th>
        <th style="padding:7px 10px;text-align:center">Total PSUs</th>
        <th style="padding:7px 10px;text-align:center;background:#27ae60">On Target</th>
        <th style="padding:7px 10px;text-align:center;background:#e67e22">Under Target</th>
        <th style="padding:7px 10px;text-align:center;background:#2980b9">Over Target</th>
        <th style="padding:7px 10px;text-align:center">% On Target</th>
      </tr></thead><tbody>`;
    p.psu_status.forEach((row,i)=>{
      const tot = (row.on_target||0)+(row.under||0)+(row.over||0);
      const pct = tot>0?Math.round(row.on_target/tot*100):0;
      const bg  = i%2===0?'#f8f9fa':'#fff';
      html += `<tr style="background:${bg}">
        <td style="padding:6px 10px;font-weight:600">R${row.round}</td>
        <td style="padding:6px 10px;text-align:center">${tot}</td>
        <td style="padding:6px 10px;text-align:center;color:#27ae60;font-weight:600">${row.on_target||0}</td>
        <td style="padding:6px 10px;text-align:center;color:#e67e22;font-weight:600">${row.under||0}</td>
        <td style="padding:6px 10px;text-align:center;color:#2980b9;font-weight:600">${row.over||0}</td>
        <td style="padding:6px 10px;text-align:center">
          <div style="background:#eee;border-radius:4px;overflow:hidden;height:14px;min-width:80px">
            <div style="background:${pct>=70?'#27ae60':'#e67e22'};width:${pct}%;height:100%"></div>
          </div>
          <span style="font-size:11px">${pct}%</span>
        </td>
      </tr>`;
    });
    html += `</tbody></table>`;
    psuEl.innerHTML = html;
  }

  // ── Region x Urban x Round table ──
  const regEl = document.getElementById('panel-region-table');
  if(regEl && p.reg_summary){
    // Group by region
    const regMap = {};
    p.reg_summary.forEach(r=>{
      const key = r.region_name;
      if(!regMap[key]) regMap[key] = {region:r.region, name:r.region_name, rows:[]};
      regMap[key].rows.push(r);
    });
    const sortedRegs = Object.values(regMap).sort((a,b)=>a.region-b.region);

    let html = `<table style="width:100%;border-collapse:collapse;font-size:12px">
      <thead><tr style="background:#1a2332;color:#fff">
        <th style="padding:7px 10px;text-align:left">Region</th>
        <th style="padding:7px 10px;text-align:left">Urban/Rural</th>`;
    rounds.forEach(r=>html+=`<th style="padding:7px 10px;text-align:center">R${r}</th>`);
    html += `<th style="padding:7px 10px;text-align:center">Total</th></tr></thead><tbody>`;

    sortedRegs.forEach((reg,ri)=>{
      reg.rows.sort((a,b)=>a.urban_label.localeCompare(b.urban_label));
      reg.rows.forEach((row,j)=>{
        const bg = ri%2===0?'#f8f9fa':'#fff';
        const rowTotal = rounds.reduce((s,r)=>s+(row.counts[String(r)]||0),0);
        html += `<tr style="background:${bg}">`;
        if(j===0) html += `<td style="padding:6px 10px;font-weight:600;vertical-align:top" rowspan="${reg.rows.length}">${reg.name}</td>`;
        const uLabel = row.urban_label==='Urban'
          ? `<span style="background:#cce5ff;color:#004085;border-radius:3px;padding:1px 5px;font-size:10.5px">Urban</span>`
          : `<span style="background:#d4edda;color:#155724;border-radius:3px;padding:1px 5px;font-size:10.5px">Rural</span>`;
        html += `<td style="padding:6px 10px">${uLabel}</td>`;
        rounds.forEach(r=>{
          const cnt = row.counts[String(r)]||0;
          const clr = cnt===0?'#ccc':'#222';
          html += `<td style="padding:6px 10px;text-align:center;color:${clr};font-weight:${cnt>0?600:400}">${cnt||'—'}</td>`;
        });
        html += `<td style="padding:6px 10px;text-align:center;font-weight:700">${rowTotal}</td>`;
        html += `</tr>`;
      });
    });
    html += `</tbody></table>`;
    regEl.innerHTML = html;
  }

  // ── In/Out summary table ──
  const inoutEl = document.getElementById('panel-inout-table');
  if(inoutEl && p.attrition){
    const baseN = p.attrition[0].n;
    let html = `<table style="width:100%;border-collapse:collapse;font-size:12.5px">
      <thead><tr style="background:#1a2332;color:#fff">
        <th style="padding:7px 10px">Round</th>
        <th style="padding:7px 10px;text-align:center">Total HHs</th>
        <th style="padding:7px 10px;text-align:center;background:#2980b9">Retained from R1</th>
        <th style="padding:7px 10px;text-align:center;background:#e74c3c">Not seen from R1</th>
        <th style="padding:7px 10px;text-align:center;background:#27ae60">New entries</th>
        <th style="padding:7px 10px;text-align:center">Retention %</th>
      </tr></thead><tbody>`;
    p.attrition.forEach((row,i)=>{
      const retPct = Math.round(row.retained/baseN*100);
      const bg = i%2===0?'#f8f9fa':'#fff';
      html += `<tr style="background:${bg}">
        <td style="padding:7px 10px;font-weight:700">R${row.round}</td>
        <td style="padding:7px 10px;text-align:center;font-weight:600">${row.n}</td>
        <td style="padding:7px 10px;text-align:center;color:#2980b9;font-weight:600">${row.retained}</td>
        <td style="padding:7px 10px;text-align:center;color:#e74c3c;font-weight:600">${row.dropped}</td>
        <td style="padding:7px 10px;text-align:center;color:#27ae60;font-weight:600">${row.new_in}</td>
        <td style="padding:7px 10px;text-align:center">
          <div style="background:#eee;border-radius:4px;overflow:hidden;height:14px;min-width:80px;display:inline-block;width:60px">
            <div style="background:${retPct>=70?'#27ae60':retPct>=50?'#e67e22':'#e74c3c'};width:${retPct}%;height:100%"></div>
          </div>
          <span style="font-size:11px;margin-left:4px">${retPct}%</span>
        </td>
      </tr>`;
    });
    html += `</tbody></table>`;
    inoutEl.innerHTML = html;
  }

  // ── Leavers vs New Entries ──
  if(p.leaver_vs_new){
    const lvn = p.leaver_vs_new;
    const transLabels = lvn.map(t=>t.label);

    // Overall verdict badge
    const allDiff = lvn.every(t=>t.verdict==='DIFFERENT');
    const anyDiff = lvn.some(t=>t.verdict==='DIFFERENT');
    const lvnBadge = document.getElementById('lvn-verdict-badge');
    if(lvnBadge){
      const cls = allDiff?'badge-red':anyDiff?'badge-yellow':'badge-green';
      const txt = allDiff?'REGIONAL BIAS IN EVERY ROUND':anyDiff?'DIFFERENCES FOUND':'SIMILAR';
      lvnBadge.className=`badge ${cls}`;
      lvnBadge.textContent=txt;
    }

    // Summary table
    const sumEl = document.getElementById('panel-lvn-summary');
    if(sumEl){
      const allVarKeys = [...new Set(lvn.flatMap(t=>Object.keys(t.vars).filter(k=>k!=='region')))];
      const varLabels = {urban:'% Urban',hhsize:'HH Size',female:'% Female',employed:'% Employed',has_account:'% Has Bank Acct'};

      let html=`<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px;min-width:600px">
        <thead>
        <tr style="background:#1a2332;color:#fff">
          <th style="padding:7px 10px" rowspan="2">Variable</th>
          ${lvn.map(t=>`<th colspan="3" style="padding:7px 8px;text-align:center;border-left:1px solid #334">${t.label}</th>`).join('')}
        </tr>
        <tr style="background:#2c3e50;color:#ccc;font-size:11px">
          ${lvn.map(()=>`<th style="padding:4px 6px;text-align:center;border-left:1px solid #334">Leavers</th><th style="padding:4px 6px;text-align:center">New</th><th style="padding:4px 6px;text-align:center">Δ</th>`).join('')}
        </tr></thead><tbody>`;

      allVarKeys.forEach((vk,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        html+=`<tr style="background:${bg}"><td style="padding:6px 10px;font-weight:600">${varLabels[vk]||vk}</td>`;
        lvn.forEach(t=>{
          const res=t.vars[vk];
          if(!res){
            html+=`<td colspan="3" style="padding:6px 8px;text-align:center;color:#ccc;border-left:1px solid #eee">n/a</td>`;
          } else {
            const diff=res.diff;
            const diffClr=res.p<0.05?'#e74c3c':res.p<0.10?'#e67e22':'#555';
            const diffStr=(diff>=0?'+':'')+diff+(vk!=='hhsize'?'pp':'');
            html+=`<td style="padding:6px 8px;text-align:center;border-left:1px solid #eee">${res.leavers}</td>
              <td style="padding:6px 8px;text-align:center">${res.new}</td>
              <td style="padding:6px 8px;text-align:center">
                <span style="color:${diffClr};font-weight:${res.p<0.10?700:400}">${diffStr}</span>
                <span style="font-size:10px;color:${diffClr};display:block">${res.sig}</span>
              </td>`;
          }
        });
        html+=`</tr>`;
      });

      // Region row
      html+=`<tr style="background:#fff3cd"><td style="padding:6px 10px;font-weight:700">Region composition</td>`;
      lvn.forEach(t=>{
        const r=t.vars.region;
        if(!r){ html+=`<td colspan="3" style="padding:6px 8px;text-align:center;color:#ccc;border-left:1px solid #eee">n/a</td>`; return; }
        const clr=r.p<0.001?'#e74c3c':r.p<0.01?'#e67e22':r.p<0.05?'#f39c12':'#27ae60';
        html+=`<td colspan="2" style="padding:6px 8px;text-align:center;border-left:1px solid #eee;font-size:11px">χ²=${r.chi2} V=${r.v}</td>
          <td style="padding:6px 8px;text-align:center"><span style="color:${clr};font-weight:700">${r.sig}</span><span style="font-size:10px;color:${clr};display:block">p=${r.p}</span></td>`;
      });
      html+=`</tr></tbody></table></div>
        <p style="font-size:11px;color:#888;margin-top:6px">Δ = new minus leavers &nbsp;|&nbsp; *** p&lt;0.001 &nbsp;** p&lt;0.01 &nbsp;* p&lt;0.05 &nbsp;† p&lt;0.10 &nbsp;ns=not significant</p>`;
      sumEl.innerHTML=html;
    }

    // Urban chart: leavers vs new per transition
    new Chart(document.getElementById('lvnUrbanChart'),{
      type:'bar',
      data:{
        labels: transLabels,
        datasets:[
          {label:'Leavers % Urban', data: lvn.map(t=>t.vars.urban?.leavers||null), backgroundColor:'#e74c3c', borderRadius:3},
          {label:'New entries % Urban', data: lvn.map(t=>t.vars.urban?.new||null), backgroundColor:'#27ae60', borderRadius:3},
        ]
      },
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:11}}},
          tooltip:{callbacks:{label:c=>`${c.dataset.label}: ${c.raw}%`}}},
        scales:{x:{},y:{min:40,max:80,title:{display:true,text:'% Urban',font:{size:11}},ticks:{callback:v=>v+'%'}}}}
    });

    // HHsize chart
    new Chart(document.getElementById('lvnHhsizeChart'),{
      type:'bar',
      data:{
        labels: transLabels,
        datasets:[
          {label:'Leavers', data: lvn.map(t=>t.vars.hhsize?.leavers||null), backgroundColor:'#e74c3c', borderRadius:3},
          {label:'New entries', data: lvn.map(t=>t.vars.hhsize?.new||null), backgroundColor:'#27ae60', borderRadius:3},
        ]
      },
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:11}}}},
        scales:{x:{},y:{min:3,title:{display:true,text:'Mean HH size',font:{size:11}}}}}
    });

    // Regional composition table per transition
    const regLvnEl = document.getElementById('panel-lvn-region');
    if(regLvnEl){
      const allRegKeys = [...new Set(lvn.flatMap(t=>Object.keys(t.vars.region?.by_region||{})))].sort((a,b)=>+a-+b);
      const regNames = {};
      lvn.forEach(t=>Object.entries(t.vars.region?.by_region||{}).forEach(([k,v])=>{ if(!regNames[k]) regNames[k]=v.name; }));

      let html=`<div style="overflow-x:auto"><table style="border-collapse:collapse;font-size:11px;min-width:600px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:6px 10px">Region</th>
          ${lvn.map(t=>`<th colspan="2" style="padding:6px 8px;text-align:center;border-left:1px solid #334">${t.label}</th>`).join('')}
        </tr>
        <tr style="background:#2c3e50;color:#ccc;font-size:10.5px">
          <th style="padding:4px 10px"></th>
          ${lvn.map(()=>`<th style="padding:4px 6px;text-align:center;border-left:1px solid #334">Leavers%</th><th style="padding:4px 6px;text-align:center">New%</th>`).join('')}
        </tr></thead><tbody>`;

      allRegKeys.forEach((rk,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        html+=`<tr style="background:${bg}"><td style="padding:5px 10px;font-weight:600">${regNames[rk]||rk}</td>`;
        lvn.forEach(t=>{
          const rd=t.vars.region?.by_region?.[rk];
          if(!rd){ html+=`<td colspan="2" style="padding:5px 6px;text-align:center;color:#ccc;border-left:1px solid #eee">—</td>`; return; }
          const diff=rd.n_pct-rd.l_pct;
          const diffClr=Math.abs(diff)>5?'#e74c3c':Math.abs(diff)>2?'#e67e22':'#555';
          html+=`<td style="padding:5px 6px;text-align:center;border-left:1px solid #eee">${rd.l_pct}%</td>
            <td style="padding:5px 6px;text-align:center"><span style="color:${diffClr};font-weight:${Math.abs(diff)>5?700:400}">${rd.n_pct}%</span></td>`;
        });
        html+=`</tr>`;
      });
      // Chi2 row
      html+=`<tr style="background:#fff3cd;font-weight:700">
        <td style="padding:6px 10px">Regional χ²</td>
        ${lvn.map(t=>{const r=t.vars.region; const clr=r?.p<0.001?'#e74c3c':r?.p<0.01?'#e67e22':'#f39c12'; return `<td colspan="2" style="padding:6px 8px;text-align:center;border-left:1px solid #eee"><span style="color:${clr}">${r?.sig||'—'} (p=${r?.p||'?'})</span></td>`;}).join('')}
      </tr>`;
      html+=`</tbody></table></div>`;
      regLvnEl.innerHTML=html;
    }
  }

  // ── Attrition composition profile ──
  if(p.attrition_profile){
    const prof = p.attrition_profile;
    const transLabels = prof.map(t=>t.label);

    // Urban % by group per transition
    const pctUrban = grp => grp.n>0 ? Math.round(grp.n_urban/grp.n*100) : 0;
    const retPct  = prof.map(t=>pctUrban(t.retained));
    const dropPct = prof.map(t=>pctUrban(t.dropped));
    const newPct  = prof.map(t=>pctUrban(t.new_in));

    new Chart(document.getElementById('panelAttrProfileChart'),{
      type:'bar',
      data:{
        labels: transLabels,
        datasets:[
          {label:'Retained % Urban', data:retPct,  backgroundColor:'#2980b9'},
          {label:'Dropped % Urban',  data:dropPct, backgroundColor:'#e74c3c'},
          {label:'New-entry % Urban',data:newPct,  backgroundColor:'#27ae60'},
        ]
      },
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:11}}},
          tooltip:{callbacks:{label:c=>`${c.dataset.label}: ${c.raw}%`}}},
        scales:{x:{},y:{beginAtZero:true,max:100,
          title:{display:true,text:'% Urban',font:{size:11}}}}}
    });

    // Volume chart
    const retN  = prof.map(t=>t.retained.n);
    const dropN = prof.map(t=>t.dropped.n);
    const newN  = prof.map(t=>t.new_in.n);
    new Chart(document.getElementById('panelAttrVolumeChart'),{
      type:'bar',
      data:{
        labels: transLabels,
        datasets:[
          {label:'Retained', data:retN,  backgroundColor:'#2980b9'},
          {label:'Dropped',  data:dropN, backgroundColor:'#e74c3c'},
          {label:'New',      data:newN,  backgroundColor:'#27ae60'},
        ]
      },
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{position:'bottom',labels:{font:{size:11}}}},
        scales:{x:{},y:{beginAtZero:true,
          title:{display:true,text:'Households',font:{size:11}}}}}
    });

    // Composition table
    const profEl = document.getElementById('panel-attrition-profile');
    if(profEl){
      let html=`<table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:6px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:7px 10px">Transition</th>
          <th colspan="2" style="padding:7px 10px;text-align:center;background:#2980b9">Retained</th>
          <th colspan="2" style="padding:7px 10px;text-align:center;background:#e74c3c">Dropped</th>
          <th colspan="2" style="padding:7px 10px;text-align:center;background:#27ae60">New Entry</th>
        </tr>
        <tr style="background:#2c3e50;color:#ccc;font-size:11px">
          <th style="padding:4px 10px"></th>
          <th style="padding:4px 10px;text-align:center">n</th><th style="padding:4px 10px;text-align:center">%Urban</th>
          <th style="padding:4px 10px;text-align:center">n</th><th style="padding:4px 10px;text-align:center">%Urban</th>
          <th style="padding:4px 10px;text-align:center">n</th><th style="padding:4px 10px;text-align:center">%Urban</th>
        </tr></thead><tbody>`;
      prof.forEach((t,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        const pu=g=>g.n>0?Math.round(g.n_urban/g.n*100)+'%':'—';
        html+=`<tr style="background:${bg}">
          <td style="padding:6px 10px;font-weight:700">${t.label}</td>
          <td style="padding:6px 10px;text-align:center;color:#2980b9;font-weight:600">${t.retained.n}</td>
          <td style="padding:6px 10px;text-align:center">${pu(t.retained)}</td>
          <td style="padding:6px 10px;text-align:center;color:#e74c3c;font-weight:600">${t.dropped.n}</td>
          <td style="padding:6px 10px;text-align:center">${pu(t.dropped)}</td>
          <td style="padding:6px 10px;text-align:center;color:#27ae60;font-weight:600">${t.new_in.n}</td>
          <td style="padding:6px 10px;text-align:center">${pu(t.new_in)}</td>
        </tr>`;
      });
      html+=`</tbody></table>`;
      profEl.innerHTML=html;
    }
  }

  // ── PSU problem tracker ──
  if(p.psu_problem_list){
    const filterEl = document.getElementById('panel-psu-filter');
    const probEl   = document.getElementById('panel-psu-problems');
    let psuFilter  = 'all';

    function renderPsuProblems(){
      let list = p.psu_problem_list;
      if(psuFilter==='urban') list=list.filter(x=>x.urban===1);
      if(psuFilter==='rural') list=list.filter(x=>x.urban===2);
      const show = list.slice(0,100);
      let html=`<p style="font-size:11.5px;color:#555;margin-bottom:8px">Showing ${show.length} of ${list.length} under-target PSUs (out of ${p.psu_problem_list.length} total). Over-target PSUs are excluded.</p>`;
      html+=`<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:11.5px;min-width:600px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:6px 8px;text-align:left">PSU</th>
          <th style="padding:6px 8px;text-align:left">Region</th>
          <th style="padding:6px 8px;text-align:center">Type</th>
          <th style="padding:6px 8px;text-align:center">Target</th>
          ${rounds.map(r=>`<th style="padding:6px 8px;text-align:center">R${r}</th>`).join('')}
          <th style="padding:6px 8px;text-align:center;background:#e74c3c">Rounds Under</th>
          <th style="padding:6px 8px;text-align:center;background:#c0392b">Rounds Zero</th>
        </tr></thead><tbody>`;
      show.forEach((psu,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        const uTag=psu.urban===1
          ?`<span style="background:#cce5ff;color:#004085;border-radius:3px;padding:1px 5px;font-size:10px">Urban</span>`
          :`<span style="background:#d4edda;color:#155724;border-radius:3px;padding:1px 5px;font-size:10px">Rural</span>`;
        const cellColor=cnt=>cnt===0?'background:#fde8e8;color:#c0392b;font-weight:700':
                              cnt<psu.target?'background:#fff3cd;color:#856404;font-weight:600':
                              cnt===psu.target?'background:#d4edda;color:#155724;font-weight:600':
                              'color:#2980b9;font-weight:600';
        html+=`<tr style="background:${bg}">
          <td style="padding:5px 8px;font-family:monospace;font-size:10.5px">${psu.psu}</td>
          <td style="padding:5px 8px">${psu.region_name}</td>
          <td style="padding:5px 8px;text-align:center">${uTag}</td>
          <td style="padding:5px 8px;text-align:center;font-weight:700">${psu.target}</td>
          ${rounds.map(r=>{
            const cnt=psu.counts[String(r)]??'—';
            return `<td style="padding:5px 8px;text-align:center;${typeof cnt==='number'?cellColor(cnt):''}">${cnt}</td>`;
          }).join('')}
          <td style="padding:5px 8px;text-align:center">
            <span style="background:${psu.n_under===ROUNDS.length?'#e74c3c':psu.n_under>=Math.ceil(ROUNDS.length*0.6)?'#e67e22':'#f39c12'};color:#fff;border-radius:3px;padding:2px 7px;font-size:11px;font-weight:700">${psu.n_under}/${ROUNDS.length}</span>
          </td>
          <td style="padding:5px 8px;text-align:center">
            ${psu.n_zero>0?`<span style="background:#c0392b;color:#fff;border-radius:3px;padding:2px 7px;font-size:11px;font-weight:700">${psu.n_zero}</span>`:`<span style="color:#888">0</span>`}
          </td>
        </tr>`;
      });
      html+=`</tbody></table></div>`;
      if(probEl) probEl.innerHTML=html;
    }

    if(filterEl){
      filterEl.innerHTML=[
        {v:'all',l:`All PSUs (${p.psu_problem_list.length})`},
        {v:'urban',l:`Urban only (${p.psu_problem_list.filter(x=>x.urban===1).length})`},
        {v:'rural',l:`Rural only (${p.psu_problem_list.filter(x=>x.urban===2).length})`},
      ].map(b=>`<button onclick="window._psuFilter='${b.v}';document.querySelectorAll('#panel-psu-filter button').forEach(x=>x.classList.remove('active'));this.classList.add('active');${''}"
        class="round-btn${psuFilter===b.v?' active':''}" style="margin-right:6px" id="psu-btn-${b.v}">${b.l}</button>`).join('');
      filterEl.querySelectorAll('button').forEach(btn=>{
        btn.addEventListener('click',()=>{
          psuFilter=btn.id.replace('psu-btn-','');
          filterEl.querySelectorAll('button').forEach(b=>b.classList.remove('active'));
          btn.classList.add('active');
          renderPsuProblems();
        });
      });
    }
    renderPsuProblems();
  }

  // ── Within-PSU refusal clustering risk ───────────────────────────────────
  const refusalEl = document.getElementById('panel-refusal-risk');
  if(refusalEl && p.psu_problem_list && bias && (bias.transitions||[]).length){

    // Step 1: aggregate PSU problem counts by region
    const regPsu = {};
    p.psu_problem_list.forEach(psu=>{
      const rk=String(psu.region);
      if(!regPsu[rk]) regPsu[rk]={name:psu.region_name, total:0, chronic:0, zero:0};
      regPsu[rk].total++;
      if(psu.n_under>=3) regPsu[rk].chronic++;
      if(psu.n_zero>0)   regPsu[rk].zero++;
    });

    // Step 2: build per-region retention lookup from bias.transitions
    const regRetMap = {};
    bias.transitions.forEach(t=>{
      (t.reg_ret_rates||[]).forEach(r=>{
        const rk=String(r.region);
        if(!regRetMap[rk]) regRetMap[rk]={name:r.name};
        regRetMap[rk][t.label]=r.pct_retained;
      });
    });

    // Step 3: merge all regions and classify risk
    const allRks=new Set([...Object.keys(regPsu),...Object.keys(regRetMap)]);
    const rows=[];
    allRks.forEach(rk=>{
      const pi=regPsu[rk]; const ri=regRetMap[rk];
      if(!pi && !ri) return;
      const name=(pi||ri).name||rk;
      const nTotal  = pi?pi.total:0;
      const nChronic= pi?pi.chronic:0;
      const nZero   = pi?pi.zero:0;
      // Lowest retention across any transition
      let lowestRet=100;
      if(ri) bias.transitions.forEach(t=>{
        const pct=ri[t.label];
        if(pct!==undefined && pct<lowestRet) lowestRet=pct;
      });
      const hasPsu=(nChronic>0||nZero>0);
      const hasAttr=(lowestRet<65);
      const risk=hasPsu&&hasAttr?'HIGH':hasPsu||hasAttr?'MODERATE':'LOW';
      rows.push({rk,name,nTotal,nChronic,nZero,lowestRet,risk,ri});
    });
    const rOrder={HIGH:0,MODERATE:1,LOW:2};
    rows.sort((a,b)=>{
      if(rOrder[a.risk]!==rOrder[b.risk]) return rOrder[a.risk]-rOrder[b.risk];
      return b.nChronic-a.nChronic;
    });

    const nHigh=rows.filter(r=>r.risk==='HIGH').length;
    const nMod =rows.filter(r=>r.risk==='MODERATE').length;
    const tLabels=bias.transitions.map(t=>t.label);

    let html=`<div style="background:#fde8e8;border-left:4px solid #e74c3c;padding:12px 16px;border-radius:4px;margin-bottom:14px;font-size:12px;color:#333">
      <strong>⚠️ Methodological note for fieldwork team</strong><br>
      Post-stratification weights correct for <em>between-PSU</em> regional imbalances, but
      <em>within-PSU</em> refusal clustering cannot be fixed by weighting. If the households
      that refused replacement share characteristics (same community norms, socioeconomic profile,
      or local trust in the survey), the responding sample within those PSUs is already self-selected
      before any weight is applied. <strong>${nHigh} region${nHigh!==1?'s':''} below are flagged HIGH</strong>
      — they show both elevated round-to-round attrition <em>and</em> chronic PSU shortfalls.
      These require field-level intervention (re-contact, community entry strategy) rather than
      statistical correction. <strong>${nMod} region${nMod!==1?'s':''} are MODERATE</strong> — one
      of the two problems is present.
    </div>
    <div style="overflow-x:auto">
    <table style="width:100%;border-collapse:collapse;font-size:12px;min-width:640px">
    <thead><tr style="background:#1a2332;color:#fff">
      <th style="padding:7px 10px;text-align:left">Region</th>
      <th style="padding:7px 10px;text-align:center">Under-target<br><span style="font-weight:normal;font-size:10px">PSUs</span></th>
      <th style="padding:7px 10px;text-align:center;background:#8e1e1e">Chronic<br><span style="font-weight:normal;font-size:10px">≥3 rounds under</span></th>
      <th style="padding:7px 10px;text-align:center;background:#6d1212">Ever-zero<br><span style="font-weight:normal;font-size:10px">0 HHs in any round</span></th>
      ${tLabels.map(l=>`<th style="padding:7px 8px;text-align:center">% Ret<br><span style="font-weight:normal;font-size:10px">${l}</span></th>`).join('')}
      <th style="padding:7px 10px;text-align:center">Risk</th>
    </tr></thead><tbody>`;

    rows.forEach((r,i)=>{
      const bg=i%2===0?'#f8f9fa':'#fff';
      const rBadge={HIGH:'badge-red',MODERATE:'badge-yellow',LOW:'badge-green'};
      html+=`<tr style="background:${bg}">
        <td style="padding:6px 10px;font-weight:600">${r.name}</td>
        <td style="padding:6px 10px;text-align:center">${r.nTotal>0?r.nTotal:'—'}</td>
        <td style="padding:6px 10px;text-align:center">${r.nChronic>0
          ?`<span style="background:#e74c3c;color:#fff;border-radius:3px;padding:1px 7px;font-weight:700">${r.nChronic}</span>`:'—'}</td>
        <td style="padding:6px 10px;text-align:center">${r.nZero>0
          ?`<span style="background:#c0392b;color:#fff;border-radius:3px;padding:1px 7px;font-weight:700">${r.nZero}</span>`:'—'}</td>
        ${bias.transitions.map(t=>{
          const pct=r.ri?r.ri[t.label]:null;
          if(pct==null) return `<td style="padding:6px 8px;text-align:center;color:#aaa">—</td>`;
          const cBg=pct<50?'#fde8e8':pct<65?'#fff3cd':pct<75?'#fef9e7':'#d4edda';
          const cTxt=pct<50?'#c0392b':pct<65?'#856404':'#155724';
          return `<td style="padding:6px 8px;text-align:center;background:${cBg};color:${cTxt};font-weight:600">${pct}%</td>`;
        }).join('')}
        <td style="padding:6px 10px;text-align:center">
          <span class="badge ${rBadge[r.risk]}">${r.risk}</span>
        </td>
      </tr>`;
    });

    html+=`</tbody></table></div>
    <p style="font-size:11px;color:#888;margin-top:7px">
      <strong>Risk = HIGH</strong>: region has ≥1 chronic or ever-zero PSU AND retention dropped below 65% in at least one transition. &nbsp;
      <strong>MODERATE</strong>: one of the two problems present. &nbsp;
      Retention columns show % of previous-round HHs still present in each transition (colour scale: red &lt;50%, orange 50–65%, green ≥75%).
    </p>`;

    refusalEl.innerHTML=html;
  }

  // ── Call interval summary and violations ──
  if(p.call_interval_summary){
    const totalViol = (p.call_violations||[]).length;
    const badgeEl = document.getElementById('panel-call-badge');
    if(badgeEl) badgeEl.textContent=`${totalViol} violations across all rounds`;

    const sumEl = document.getElementById('panel-call-summary');
    if(sumEl){
      let html=`<table style="width:100%;border-collapse:collapse;font-size:12.5px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:7px 10px">Round</th>
          <th style="padding:7px 10px;text-align:center">HHs with prev. interview</th>
          <th style="padding:7px 10px;text-align:center;background:#e74c3c">Called &lt;20 days</th>
          <th style="padding:7px 10px;text-align:center">% Early</th>
          <th style="padding:7px 10px;text-align:center">Median gap (days)</th>
          <th style="padding:7px 10px;text-align:center">Min</th>
          <th style="padding:7px 10px;text-align:center">Max</th>
        </tr></thead><tbody>`;
      p.call_interval_summary.forEach((row,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        const pct=row.pct_under||0;
        const pctClr=pct>50?'#e74c3c':pct>20?'#e67e22':'#27ae60';
        html+=`<tr style="background:${bg}">
          <td style="padding:7px 10px;font-weight:700">R${row.round}</td>
          <td style="padding:7px 10px;text-align:center">${row.n_total}</td>
          <td style="padding:7px 10px;text-align:center;color:${(row.n_under||0)>0?'#e74c3c':'#27ae60'};font-weight:700">${row.n_under||0}</td>
          <td style="padding:7px 10px;text-align:center">
            <span style="color:${pctClr};font-weight:700">${pct}%</span>
          </td>
          <td style="padding:7px 10px;text-align:center;font-weight:600">${row.median??'—'}</td>
          <td style="padding:7px 10px;text-align:center;color:${row.min<20?'#e74c3c':'#222'};font-weight:${row.min<20?700:400}">${row.min??'—'}</td>
          <td style="padding:7px 10px;text-align:center">${row.max??'—'}</td>
        </tr>`;
      });
      html+=`</tbody></table>`;
      sumEl.innerHTML=html;
    }

    // Per-round tab buttons + violation table
    const allViols = p.call_violations || [];
    const rounds = [...new Set(allViols.map(v=>v.round))].sort((a,b)=>a-b);
    let _ciRound = rounds[0] || null;   // currently selected round
    Object.defineProperty(window, '_ciRound', {get(){return _ciRound}, set(v){_ciRound=v}});

    window.renderViolTabs = renderViolTabs;
    function renderViolTabs(){
      const tabEl = document.getElementById('panel-call-round-tabs');
      const violEl = document.getElementById('panel-call-violations');
      if(!tabEl || !violEl) return;

      // Build round tab buttons
      let tabHtml = `<div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:6px">
        <span style="font-size:11px;color:#888;margin-right:2px">Filter by round:</span>`;
      const allBtn = `<button onclick="_ciRound=null;renderViolTabs()" style="padding:3px 10px;border-radius:3px;font-size:11.5px;font-weight:600;cursor:pointer;border:1.5px solid ${_ciRound===null?'#2d3f55':'#aab8c8'};background:${_ciRound===null?'#2d3f55':'#f8f9fa'};color:${_ciRound===null?'#fff':'#444'}">All (${allViols.length})</button>`;
      tabHtml += allBtn;
      rounds.forEach(r=>{
        const cnt = allViols.filter(v=>v.round===r).length;
        const active = _ciRound===r;
        tabHtml+=`<button onclick="_ciRound=${r};renderViolTabs()" style="padding:3px 10px;border-radius:3px;font-size:11.5px;font-weight:600;cursor:pointer;border:1.5px solid ${active?'#e74c3c':'#aab8c8'};background:${active?'#e74c3c':'#f8f9fa'};color:${active?'#fff':'#444'}">R${r} (${cnt})</button>`;
      });
      tabHtml += `</div>`;
      tabEl.innerHTML = tabHtml;

      // Filter violations
      const viols = (_ciRound===null ? allViols : allViols.filter(v=>v.round===_ciRound)).slice(0,300);
      const total  = _ciRound===null ? allViols.length : allViols.filter(v=>v.round===_ciRound).length;

      if(viols.length===0){
        violEl.innerHTML=`<p style="color:#27ae60;font-size:12px;padding:8px 0">✓ No violations for this round.</p>`;
        return;
      }
      let html=`<p style="font-size:11.5px;color:#555;margin-bottom:8px">Showing ${viols.length} of ${total} violations (&lt;20 days since last interview).</p>`;
      html+=`<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:11.5px;min-width:500px">
        <thead><tr style="background:#1a2332;color:#fff">
          <th style="padding:6px 10px">HH ID</th>
          <th style="padding:6px 10px;text-align:center">Round</th>
          <th style="padding:6px 10px;text-align:center">Prev Round</th>
          <th style="padding:6px 10px;text-align:center;background:#e74c3c">Days Gap</th>
          <th style="padding:6px 10px;text-align:center">Urban/Rural</th>
          <th style="padding:6px 10px;text-align:left">Region</th>
          <th style="padding:6px 10px;text-align:left">PSU</th>
        </tr></thead><tbody>`;
      viols.forEach((v,i)=>{
        const bg=i%2===0?'#f8f9fa':'#fff';
        const gapClr=v.days_gap<10?'#7b241c':v.days_gap<15?'#c0392b':'#e74c3c';
        html+=`<tr style="background:${bg}">
          <td style="padding:5px 10px;font-weight:600">${v.hhid}</td>
          <td style="padding:5px 10px;text-align:center;font-weight:700">R${v.round}</td>
          <td style="padding:5px 10px;text-align:center;color:#888">R${v.prev_round}</td>
          <td style="padding:5px 10px;text-align:center">
            <span style="background:${gapClr};color:#fff;border-radius:3px;padding:2px 8px;font-weight:700">${v.days_gap}d</span>
          </td>
          <td style="padding:5px 10px;text-align:center">
            ${v.urban_label==='Urban'
              ?`<span style="background:#cce5ff;color:#004085;border-radius:3px;padding:1px 5px;font-size:10px">Urban</span>`
              :`<span style="background:#d4edda;color:#155724;border-radius:3px;padding:1px 5px;font-size:10px">Rural</span>`}
          </td>
          <td style="padding:5px 10px">${v.region_name}</td>
          <td style="padding:5px 10px;font-family:monospace;font-size:10.5px">${v.psu}</td>
        </tr>`;
      });
      html+=`</tbody></table></div>`;
      violEl.innerHTML=html;
    }
    renderViolTabs();
  }
}

// ── MISSING HEATMAP ──────────────────────────────────────────────────────────
let hmCurrentMod='M00';
function buildMissing(){
  const tabs=document.getElementById('hm-tabs');
  tabs.innerHTML=MODULES.map(m=>`<span class="mtab${m===hmCurrentMod?' active':''}" onclick="switchHM('${m}')">${m}</span>`).join('');
  renderHM(hmCurrentMod);
}
function switchHM(m){
  hmCurrentMod=m;
  document.querySelectorAll('#hm-tabs .mtab').forEach((t,i)=>t.classList.toggle('active',MODULES[i]===m));
  renderHM(m);
}
function renderHM(mod){
  const rows = DQ.heatmap_data[mod]||[];
  if(!rows.length){document.getElementById('hm-content').innerHTML='<p style="color:#888;padding:10px">No data.</p>';return;}
  let html=`<div class="heatmap-wrap"><table class="heatmap"><thead><tr>
    <th class="left" style="min-width:150px">Variable</th>
    ${ROUNDS.map(r=>`<th>R${r}</th>`).join('')}
    <th>RAG</th></tr></thead><tbody>`;
  rows.forEach(row=>{
    const rag=row.rag||'green';
    const ragDot={red:'🔴',yellow:'🟡',green:'🟢'}[rag]||'⚪';
    html+=`<tr><td class="vn">${row.var}</td>`;
    ROUNDS.forEach(r=>{
      const v=row[r];
      const bg=hmColor(v), tc=v>50?'#333':'#333';
      html+=`<td><div class="hm-cell" style="background:${bg};color:${tc}">${hmText(v)}</div></td>`;
    });
    html+=`<td>${ragDot}</td></tr>`;
  });
  html+='</tbody></table></div>';
  document.getElementById('hm-content').innerHTML=html;
}

// ── INTERVIEW QUALITY ────────────────────────────────────────────────────────
function buildInterview(){
  const meta=DQ.interview_meta;
  const dur=meta.duration?.by_round||{};
  const modDur=meta.module_durations?.by_module||{};
  const partial=meta.partial_interviews?.by_round||{};
  const excess=meta.excess_interviews?.by_round||{};
  const calls=meta.call_attempts?.by_round||{};

  const totalShort=ROUNDS.reduce((s,r)=>{const d=dur[r];return s+(d?(d.very_short||0)+(d.short||0):0)},0);
  const totalPartial=Object.values(partial).reduce((s,v)=>s+(v||0),0);
  const totalExcess=Object.values(excess).reduce((s,v)=>s+(v||0),0);

  document.getElementById('int-stats').innerHTML=`
    <div class="stat-box ${totalShort>10?'red':'yellow'}"><div class="num">${totalShort}</div><div class="lbl">Short interviews (&lt;20 min)</div></div>
    <div class="stat-box ${totalPartial>0?'yellow':'green'}"><div class="num">${totalPartial}</div><div class="lbl">Partial interviews</div></div>
    <div class="stat-box ${totalExcess>0?'yellow':'green'}"><div class="num">${totalExcess}</div><div class="lbl">Excess interviews</div></div>
    <div class="stat-box blue"><div class="num">${ROUNDS.length}</div><div class="lbl">Rounds covered</div></div>
  `;

  setTimeout(()=>{
    makeBar('durChart',ROUNDS.map(r=>`R${r}`),[
      {label:'P25',data:ROUNDS.map(r=>dur[r]?.p25??null),backgroundColor:'rgba(52,152,219,.3)',borderRadius:0},
      {label:'Median',data:ROUNDS.map(r=>dur[r]?.p50??null),backgroundColor:'rgba(52,152,219,.85)',borderRadius:4},
      {label:'P75',data:ROUNDS.map(r=>dur[r]?.p75??null),backgroundColor:'rgba(52,152,219,.3)',borderRadius:0},
    ],{y:{ticks:{callback:v=>`${v}m`}},plugins:{legend:{display:true,labels:{font:{size:10},boxWidth:8}}}});

    const stacks=['very_short','short','normal','long','very_long'];
    const sColors=['#e74c3c','#e67e22','#2ecc71','#f39c12','#c0392b'];
    const sLabels=['<10m','10–20m','20–60m','60–120m','>2h'];
    makeBar('durCatChart',ROUNDS.map(r=>`R${r}`),
      stacks.map((k,i)=>({label:sLabels[i],
        data:ROUNDS.map(r=>{const d2=dur[r];return d2&&d2.n?Math.round((d2[k]||0)/d2.n*100):0}),
        backgroundColor:sColors[i],borderRadius:2})),
      {y:{stacked:true,ticks:{callback:v=>`${v}%`}},scales:{x:{stacked:true},y:{stacked:true}},
       plugins:{legend:{display:true,labels:{font:{size:10},boxWidth:8}}}});

    const mLabels=Object.keys(modDur);
    makeBar('modDurChart',mLabels,ROUNDS.map((r,i)=>({
      label:`R${r}`,data:mLabels.map(m=>modDur[m]?.[r]??null),
      backgroundColor:R_COLORS[i],borderRadius:3})),
      {y:{ticks:{callback:v=>`${v}m`}},plugins:{legend:{display:true,labels:{font:{size:10},boxWidth:8}}}});

    makeBar('partialChart',ROUNDS.map(r=>`R${r}`),[{
      label:'Partial',data:ROUNDS.map(r=>partial[r]??0),
      backgroundColor:ROUNDS.map(r=>(partial[r]||0)>5?'#e74c3c':'#f39c12'),borderRadius:4}]);

    const cMean=Object.fromEntries(Object.entries(calls).map(([r,v])=>[r,v?.mean??null]));
    const c3p=Object.fromEntries(Object.entries(calls).map(([r,v])=>[r,v?.attempts_3plus??null]));
    makeBar('callChart',ROUNDS.map(r=>`R${r}`),[
      {label:'Mean attempts',data:ROUNDS.map(r=>cMean[r]??null),backgroundColor:'#3498db',borderRadius:4},
      {label:'3+ attempts',data:ROUNDS.map(r=>c3p[r]??null),backgroundColor:'#e74c3c',borderRadius:4}],
      {y:{min:0},plugins:{legend:{display:true,labels:{font:{size:10},boxWidth:8}}}});
  },50);
}

// ── ALL CHANGES PAGE ──────────────────────────────────────────────────────────
let changesRoundFilter = 'all';
function buildChanges(){
  const filters = document.getElementById('changes-filters');
  const filterRounds = ['all', ...ROUNDS.slice(1).map(r=>`R${r}`)];  // R2..Rmax
  filters.innerHTML = `<span class="rfil-label">Show round:</span>
    ${filterRounds.map(r=>`<span class="rfil${r===changesRoundFilter?' active':''}" onclick="setChangesFilter('${r}')">${r==='all'?'All Rounds':r}</span>`).join('')}`;
  // Update the subtitle round range to match ROUNDS
  const _rSpan = document.getElementById('all-changes-round-range');
  if(_rSpan) _rSpan.textContent = `R2–R${ROUNDS.length}`;
  renderChanges();
}
function setChangesFilter(r){
  changesRoundFilter=r;
  buildChanges();
}
function renderChanges(){
  const allRounds = ROUNDS.slice(1).map(r=>`R${r}`);  // R2..Rmax
  // Build prev map from ROUNDS so future rounds (R8+) extend automatically
  const prevR = {};
  for(let i=1; i<ROUNDS.length; i++) prevR[`R${ROUNDS[i]}`] = `R${ROUNDS[i-1]}`;
  const toShow = changesRoundFilter==='all'?allRounds:[changesRoundFilter];
  let html='';

  toShow.forEach(rnd=>{
    let changes=[];
    MODULES.forEach(m=>{
      (MT[m]||[]).forEach(row=>{
        const inC=row[`in_${rnd}`]==='✓', inP=row[`in_${prevR[rnd]}`]==='✓';
        const rIdx=['R1','R2','R3','R4','R5','R6','R7','R8'].indexOf(rnd);
        const isFirst=['R1','R2','R3','R4','R5','R6','R7','R8'].slice(0,rIdx).every(pr=>row[`in_${pr}`]!=='✓');
        if(inC&&isFirst) changes.push({type:'new',var:row.variable,mod:m,title:row.question_title,text:row.english_text,detail:`First appears in ${rnd}`});
        if(!inC&&inP) changes.push({type:'drop',var:row.variable,mod:m,title:row.question_title,text:row.english_text,detail:`Present in ${prevR[rnd]}, absent in ${rnd}+`});
        if(inC&&inP){
          if(row.title_changes&&row.title_changes.includes(`${prevR[rnd]}→${rnd}`)){
            const detail=row.title_changes.split('|').find(x=>x.includes(`${prevR[rnd]}→${rnd}`))||row.title_changes;
            changes.push({type:'title',var:row.variable,mod:m,title:row.question_title,text:'',detail:detail.trim()});
          }
          if(row.skip_changes&&row.skip_changes.includes(rnd)){
            const detail=row.skip_changes.split('|').find(x=>x.includes(rnd))||row.skip_changes;
            changes.push({type:'skip',var:row.variable,mod:m,title:row.question_title,text:'',detail:detail.trim()});
          }
          if(row.option_changes&&row.option_changes.includes(rnd)){
            const detail=row.option_changes.split('|').filter(x=>x.includes(rnd)).join('; ')||row.option_changes;
            changes.push({type:'opts',var:row.variable,mod:m,title:row.question_title,text:'',detail:detail.trim()});
          }
        }
      });
    });

    if(!changes.length) return;
    const nNew=changes.filter(c=>c.type==='new').length;
    const nDrop=changes.filter(c=>c.type==='drop').length;
    const nOther=changes.length-nNew-nDrop;
    const ragColor=nDrop>5?'#e74c3c':nNew>10?'#f39c12':'#27ae60';
    html+=`<div class="card" style="margin-bottom:12px">
    <h2 style="cursor:pointer;user-select:none;font-size:14px" onclick="toggleSection(this)">
      <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${ragColor}">▼</span>
      <span class="badge badge-blue">${rnd}</span>
      <span class="rch-arrow" style="margin:0 4px">vs</span>
      <span class="badge badge-grey">${prevR[rnd]}</span>
      <span class="section-summary" style="margin-left:8px;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
        <span style="color:#888">${changes.length} change${changes.length>1?'s':''}</span>
        ${nNew>0?'<span style="color:#27ae60">'+nNew+' new</span>':''}
        ${nDrop>0?'<span style="color:#e74c3c">'+nDrop+' dropped</span>':''}
        ${nOther>0?'<span style="color:#e67e22">'+nOther+' modified</span>':''}
      </span>
    </h2>
    <div class="collapsible-body">`;

    const typeIcon={new:'➕',drop:'➖',title:'✏️',skip:'🔀',code:'📋',opts:'🔢'};
    const typeTag={new:'chg-new',drop:'chg-drop',title:'chg-title',skip:'chg-skip',code:'chg-code',opts:'chg-opts'};
    const typeLabel={new:'New Question',drop:'Dropped',title:'Wording Change',skip:'Skip Logic',code:'Code Change',opts:'Options Changed'};

    html+=`<table style="width:100%;border-collapse:collapse;font-size:12px">
      <thead><tr style="background:#f7f9fc">
        <th style="padding:5px 8px;text-align:left;font-size:10.5px;color:#555;text-transform:uppercase">Mod</th>
        <th style="padding:5px 8px;text-align:left">Variable</th>
        <th style="padding:5px 8px;text-align:left">Change</th>
        <th style="padding:5px 8px;text-align:left">Question Title</th>
        <th style="padding:5px 8px;text-align:left">Detail</th>
      </tr></thead><tbody>`;

    changes.forEach((c,ci)=>{
      const bg=ci%2===0?'#fff':'#f9f9f9';
      html+=`<tr style="background:${bg};border-bottom:1px solid #eee">
        <td style="padding:5px 8px"><span class="badge badge-blue" style="font-size:10px">${c.mod}</span></td>
        <td style="padding:5px 8px;font-family:monospace;font-weight:700;font-size:12px">${c.var}</td>
        <td style="padding:5px 8px"><span class="chg-tag ${typeTag[c.type]}">${typeIcon[c.type]} ${typeLabel[c.type]}</span></td>
        <td style="padding:5px 8px;font-size:11.5px;color:#333">${c.title}</td>
        <td style="padding:5px 8px;font-size:11px;color:#555">${c.detail}</td>
      </tr>`;
    });

    html+='</tbody></table></div></div>';
  });

  document.getElementById('changes-content').innerHTML=html||'<p style="color:#888;padding:20px">No changes detected for selected filters.</p>';
}

// ══════════════════════════════════════════════════════════════════════════════
// MODULE DEEP-DIVE PAGES
// ══════════════════════════════════════════════════════════════════════════════
function buildAllModulePages(){
  const container = document.getElementById('module-pages-container');
  MODULES.forEach(m=>{
    const div = document.createElement('div');
    div.id = `page-mod-${m}`;
    div.className = 'page';
    div.innerHTML = buildModulePage(m);
    container.appendChild(div);
  });
}

function buildModulePage(mod){
  const vmap = _buildVarMaps(mod);
  const rows = (MT[mod]||[]).slice().sort((a,b)=>_varSort(a.variable,vmap)-_varSort(b.variable,vmap));
  const s = DQ.module_summary[mod]||{};

  // Augment heatmap with synthetic 0% rows for all MODULE_VAR_ORDER vars not in pipeline data
  let _rawHeat = (DQ.heatmap_data[mod]||[]).slice();
  if(MODULE_VAR_ORDER[mod]){
    // Build allowed-var set from MODULE_VAR_ORDER
    const _allowedVars = new Set();
    MODULE_VAR_ORDER[mod].forEach(entry=>{
      _allowedVars.add(entry.kobo.toLowerCase());
      _allowedVars.add(entry.stata.toLowerCase());
    });
    // Filter: only keep heatmap rows that match MODULE_VAR_ORDER entries
    _rawHeat = _rawHeat.filter(r=>{
      const v=(r.var||'').toLowerCase();
      if(_allowedVars.has(v)) return true;
      const pm=v.match(/\(([^)]+)\)/);
      if(pm && _allowedVars.has(pm[1].toLowerCase())) return true;
      const base=v.replace(/\s*\(.*?\)\s*/g,'').trim();
      if(base && _allowedVars.has(base)) return true;
      return false;
    });
    // Build existing-var set for augmentation
    const _existVars = new Set();
    _rawHeat.forEach(r=>{
      const v=r.var||'';
      _existVars.add(v.toLowerCase());
      const pm=v.match(/\(([^)]+)\)/);
      if(pm) _existVars.add(pm[1].toLowerCase());
      const base=v.replace(/\s*\(.*?\)\s*/g,'').trim().toLowerCase();
      if(base) _existVars.add(base);
    });
    // Inject synthetic 0% rows for MODULE_VAR_ORDER vars not in pipeline data
    MODULE_VAR_ORDER[mod].forEach(entry=>{
      const kl=entry.kobo.toLowerCase(), sl=entry.stata.toLowerCase();
      if(!_existVars.has(kl) && !_existVars.has(sl)){
        const label = kl===sl ? entry.kobo : entry.kobo+' ('+entry.stata+')';
        const synth = {var:label, rag:'green'};
        ROUNDS.forEach(r=>{ synth[String(r)]=0; });
        _rawHeat.push(synth);
      }
    });
  }
  // Sort: regular variables by Kobo order, then dur_* variables at the bottom
  const heatRows = _rawHeat.sort((a,b)=>{
    const aIsDur = (a.var||'').toLowerCase().startsWith('dur_');
    const bIsDur = (b.var||'').toLowerCase().startsWith('dur_');
    if(aIsDur && !bIsDur) return 1;
    if(!aIsDur && bIsDur) return -1;
    return _varSort(a.var,vmap)-_varSort(b.var,vmap);
  });

  const skipIssues = DQ.skip_issues.filter(x=>x.module===mod);
  const mandIssues = DQ.mandatory_issues.filter(x=>x.module===mod);
  // OOR: filter to only show this module's own duration var (not all dur_* from M00)
  const _ownDur = MODULE_DUR_VAR[mod];
  const oorIssues  = DQ.oor_issues.filter(x=>{
    if(x.module!==mod) return false;
    const v = (x.variable||'').toLowerCase();
    // Keep non-duration vars always
    if(!v.startsWith('dur_')) return true;
    // For duration vars, only keep this module's own + dur_tot
    return v===_ownDur || v==='dur_tot';
  }).slice().sort((a,b)=>_varSort(a.variable,vmap)-_varSort(b.variable,vmap));

  // Build DQ lookup by variable (lowercase)
  // Also extract questionnaire codes in parentheses, e.g. "(D27)" → "d27"
  const dqByVar = {};
  function _addDQ(vl, cat, x){
    if(!dqByVar[vl]) dqByVar[vl]={skip:[],mand:[],oor:[]};
    dqByVar[vl][cat].push(x);
  }
  function _extractKeys(str){
    // Split on arrows/commas/spaces, keep tokens starting with a letter
    const keys = str.split(/[→,\s]+/).filter(v=>v.match(/^[A-Za-z]/)).map(v=>v.toLowerCase());
    // Also extract codes in parentheses: (D27), (M10_1), (M13_oth) etc.
    const paren = str.match(/\(([A-Za-z][A-Za-z0-9_]*)\)/g);
    if(paren) paren.forEach(p=>keys.push(p.replace(/[()]/g,'').toLowerCase()));
    return [...new Set(keys)];
  }
  skipIssues.forEach(x=>{
    _extractKeys(x.variable).forEach(vl=>_addDQ(vl,'skip',x));
  });
  mandIssues.forEach(x=>{
    _extractKeys(x.variable).forEach(vl=>_addDQ(vl,'mand',x));
  });
  oorIssues.forEach(x=>{
    const vl=(x.variable||'').toLowerCase();
    _addDQ(vl,'oor',x);
    // Also extract parenthetical codes from OOR variable names
    const paren = (x.variable||'').match(/\(([A-Za-z][A-Za-z0-9_]*)\)/g);
    if(paren) paren.forEach(p=>_addDQ(p.replace(/[()]/g,'').toLowerCase(),'oor',x));
  });

  const totalSkipViol = skipIssues.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0),0);
  const totalMandMiss = mandIssues.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0),0);
  const newQs    = rows.filter(r=>r.status&&r.status.startsWith('New')).length;
  const droppedQs= rows.filter(r=>r.status&&r.status.startsWith('Dropped')).length;
  const changedQs= rows.filter(r=>r.title_changes||r.skip_changes||r.option_changes).length;

  // Per-round question presence counts
  const presCount = {};
  ['R1','R2','R3','R4','R5','R6','R7','R8'].forEach(r=>{presCount[r]=rows.filter(x=>x[`in_${r}`]==='✓').length});

  // Build the HTML
  let html = `
  <div class="mod-page-header">
    <div>
      <h1>${mod} – ${MOD_NAMES[mod]}</h1>
      <p class="subtitle">Per-question tracker · Questionnaire changes · Data quality issues</p>
    </div>
    <div class="mod-round-summary">
      ${['R1','R2','R3','R4','R5','R6','R7','R8'].map(r=>{
        const n=presCount[r];
        const cls = n===0?'mrs-na':'mrs-ok';
        return `<div class="${cls} mrs-pill">R${r.slice(1)}: ${n}q</div>`;
      }).join('')}
    </div>
  </div>

  <div class="mod-dq-bar">
    <div class="mdb-item ${totalSkipViol>0?'red':'green'}"><div class="mdb-num">${totalSkipViol}</div><div class="mdb-lbl">Skip violations</div></div>
    <div class="mdb-item ${totalMandMiss>0?'yellow':'green'}"><div class="mdb-num">${totalMandMiss}</div><div class="mdb-lbl">Mandatory missing</div></div>
    <div class="mdb-item blue"><div class="mdb-num">${rows.length}</div><div class="mdb-lbl">Unique questions</div></div>
    <div class="mdb-item ${newQs>0?'purple':'green'}"><div class="mdb-num">${newQs}</div><div class="mdb-lbl">New questions</div></div>
    <div class="mdb-item ${droppedQs>0?'yellow':'green'}"><div class="mdb-num">${droppedQs}</div><div class="mdb-lbl">Dropped questions</div></div>
    <div class="mdb-item ${changedQs>0?'yellow':'green'}"><div class="mdb-num">${changedQs}</div><div class="mdb-lbl">Changed</div></div>
  </div>`;

  // RAG explanation for this module
  const rag = s.rag||'green';
  const maxMiss = (s.max_missing_pct||0).toFixed(1);
  if(rag==='red'){
    const reasons = [];
    if(totalSkipViol>100) reasons.push(`<strong>${totalSkipViol}</strong> skip violations (threshold: 100)`);
    if(totalMandMiss>100) reasons.push(`<strong>${totalMandMiss}</strong> mandatory missing (threshold: 100)`);
    if(parseFloat(maxMiss)>=30) reasons.push(`worst variable at <strong>${maxMiss}%</strong> missing (threshold: 30%)`);
    html+=`<div class="note-box" style="background:rgba(231,76,60,0.08);border-left:4px solid #e74c3c;margin-bottom:10px">
      <strong style="color:#e74c3c">🔴 This module is RED because:</strong>
      <ul style="margin:4px 0 0 16px;font-size:12px;color:#444">${reasons.map(r=>`<li>${r}</li>`).join('')}</ul>
    </div>`;
  } else if(rag==='yellow'){
    const reasons = [];
    if(totalSkipViol>0) reasons.push(`<strong>${totalSkipViol}</strong> skip violation${totalSkipViol>1?'s':''}`);
    if(totalMandMiss>0) reasons.push(`<strong>${totalMandMiss}</strong> mandatory missing`);
    if((s.n_oor_values||0)>0) reasons.push(`<strong>${s.n_oor_values}</strong> out-of-range value${s.n_oor_values>1?'s':''}`);
    if(parseFloat(maxMiss)>=10) reasons.push(`worst variable at <strong>${maxMiss}%</strong> missing (threshold: 10%)`);
    html+=`<div class="note-box" style="background:rgba(243,156,18,0.08);border-left:4px solid #f39c12;margin-bottom:10px">
      <strong style="color:#f39c12">🟡 This module is YELLOW because:</strong>
      <ul style="margin:4px 0 0 16px;font-size:12px;color:#444">${reasons.map(r=>`<li>${r}</li>`).join('')}</ul>
    </div>`;
  } else {
    html+=`<div class="note-box" style="background:rgba(46,204,113,0.08);border-left:4px solid #2ecc71;margin-bottom:10px">
      <strong style="color:#2ecc71">🟢 This module is GREEN:</strong>
      <span style="font-size:12px;color:#444">Zero violations · all variables below 10% missing (worst: ${maxMiss}%)</span>
    </div>`;
  }

  // Module questionnaire notes (Kobo XLSForm evolution)
  const modNotes = MODULE_NOTES[mod];
  if(modNotes && modNotes.items && modNotes.items.length>0){
    html+=`<div class="card" style="border:1px solid #e8d5b0;background:linear-gradient(135deg,#fffef8,#fdf6e3)">
      <h2 style="color:#8b6914;font-size:14px;margin-bottom:8px">📋 ${modNotes.title}</h2>
      <div style="font-size:11px;color:#666;margin-bottom:8px">
        Cross-round questionnaire changes identified from Kobo XLSForm audit. These affect data interpretation and skip logic validation.
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:12px">
        <thead>
          <tr style="border-bottom:2px solid #e8d5b0;text-align:left">
            <th style="padding:4px 8px;width:100px;color:#8b6914">Rounds</th>
            <th style="padding:4px 8px;width:90px;color:#8b6914">Type</th>
            <th style="padding:4px 8px;color:#8b6914">Detail</th>
          </tr>
        </thead>
        <tbody>`;
    modNotes.items.forEach((n,i)=>{
      const bg = i%2===0?'transparent':'rgba(139,105,20,0.03)';
      const tagColor = n.tag.includes('⚠')? '#e74c3c':'#8b6914';
      html+=`<tr style="background:${bg};border-bottom:1px solid #f0e6cc">
        <td style="padding:5px 8px;font-family:monospace;font-size:11px;white-space:nowrap">${n.rounds}</td>
        <td style="padding:5px 8px"><span style="display:inline-block;background:rgba(139,105,20,0.1);color:${tagColor};padding:1px 6px;border-radius:3px;font-size:10px;font-weight:600">${n.tag}</span></td>
        <td style="padding:5px 8px;color:#444;line-height:1.4">${n.text}</td>
      </tr>`;
    });
    html+=`</tbody></table></div>`;
  }

  // Question tracker table
  html += `
  <div class="card">
    <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
      <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px">▼</span>
      Question-Level Cross-Round Tracker
      <span class="badge badge-grey" style="cursor:pointer" onclick="event.stopPropagation();toggleAllDetails('${mod}')">Toggle All Details</span>
    </h2>
    <div class="collapsible-body">
    <div class="note-info note-box" style="margin-bottom:10px">
      ✓ = present in that round &nbsp;|&nbsp; — = absent &nbsp;|&nbsp; Click any row for full details incl. skip logic per round.
      Coloured tags show questionnaire changes; 🔴/🟡 icons flag DQ issues in the actual data.
    </div>

    <div class="heatmap-wrap">
    <table class="qtrack-table" id="qtable-${mod}">
    <thead class="sticky-round">
      <tr>
        <th class="left" style="min-width:80px">Variable</th>
        <th class="left" style="min-width:160px">Question Title</th>
        ${['R1','R2','R3','R4','R5','R6','R7','R8'].map(r=>`<th style="min-width:42px">${r}</th>`).join('')}
        <th class="left" style="min-width:130px">Changes</th>
        <th class="left" style="min-width:100px">DQ Issues</th>
        <th class="left" style="min-width:90px">Type</th>
      </tr>
    </thead>
    <tbody>`;

  rows.forEach((row,ri)=>{
    const v = row.variable||'';
    const vl = v.toLowerCase();
    const dq = dqByVar[vl]||{skip:[],mand:[],oor:[]};
    const hasDQ = dq.skip.length+dq.mand.length+dq.oor.length>0;
    const rowBg = row.status&&row.status.startsWith('New')?'#f0fff4':
                  row.status&&row.status.startsWith('Dropped')?'#fff5f5':
                  (row.title_changes||row.skip_changes||row.option_changes)?'#fffef0':'';

    // Build change tags
    let changeTags='';
    if(row.status&&row.status.startsWith('New')) changeTags+=`<span class="chg-tag chg-new">➕ ${row.status}</span>`;
    if(row.status&&row.status.startsWith('Dropped')) changeTags+=`<span class="chg-tag chg-drop">➖ ${row.status}</span>`;
    if(row.title_changes) changeTags+=`<span class="chg-tag chg-title" title="${row.title_changes.replace(/"/g,"'")}">✏️ Wording</span>`;
    if(row.skip_changes) changeTags+=`<span class="chg-tag chg-skip" title="${row.skip_changes.replace(/"/g,"'")}">🔀 Skip logic</span>`;
    if(row.option_changes) changeTags+=`<span class="chg-tag chg-opts" title="${row.option_changes.replace(/"/g,"'")}">🔢 Options</span>`;

    // DQ badges
    let dqTags='';
    const skipViol = dq.skip.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v2)=>a+(v2||0),0),0);
    const mandViol = dq.mand.reduce((s,x)=>s+Object.values(x.counts_by_round).reduce((a,v2)=>a+(v2||0),0),0);
    if(skipViol>0) dqTags+=`<span class="dq-inline dq-skip">🔴 Skip: ${skipViol}</span>`;
    if(mandViol>0) dqTags+=`<span class="dq-inline dq-mand">🟡 Mand: ${mandViol}</span>`;
    if(dq.oor.length>0) dqTags+=`<span class="dq-inline dq-oor">⚠ OOR</span>`;

    // Presence cells
    const pres = ['R1','R2','R3','R4','R5','R6','R7','R8'].map(r=>{
      const p = row[`in_${r}`]==='✓';
      return `<td style="text-align:center"><span class="pres ${p?'yes':'no'}">${p?'✓':'—'}</span></td>`;
    }).join('');

    // Expandable detail row id
    const did=`det-${mod}-${ri}`;

    html+=`
    <tr class="toggle-row" onclick="toggleDetail('${did}')" style="${rowBg?`background:${rowBg}`:''}">
      <td class="var-name">${_varLabel(v,vmap)} <button class="toggle-btn">▾</button></td>
      <td class="q-title">${(row.question_title||'').substring(0,60)}${(row.question_title||'').length>60?'…':''}</td>
      ${pres}
      <td>${changeTags||'<span style="color:#aaa;font-size:11px">—</span>'}</td>
      <td>${dqTags||'<span style="color:#aaa;font-size:11px">—</span>'}</td>
      <td><span class="qtype">${row.question_type||'—'}</span></td>
    </tr>
    <tr><td colspan="12" style="padding:0">
      <div class="q-detail" id="${did}">${buildDetailPanel(row,dq,mod)}</div>
    </td></tr>`;
  });

  html+=`</tbody></table></div></div></div>`;

  // ─────── SKIP VIOLATIONS SECTION ───────
  // Hide fully-green checks (all rounds have data AND all counts are 0).
  // Keep checks with N/A rounds — the questionnaire difference is informative.
  const skipList = skipIssues.map(x=>{
    const total = Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
    const hasNA = ROUNDS.some(r=>(x.counts_by_round[r]??x.counts_by_round[String(r)]??null)===null);
    // Fully green = zero violations AND no N/A rounds → hide
    if(total===0 && !hasNA) return '';
    const sev = x.severity||'medium';
    const icon = total===0?'🟢':sev==='high'?'🔴':'🟡';
    const rowCls = total===0?'clean':sev;
    const pills = ROUNDS.map(r=>{
      const raw = x.counts_by_round[r]??x.counts_by_round[String(r)]??null;
      const hasData = raw!=null;
      const v = raw||0;
      const p = x.pct_by_round?.[r]??x.pct_by_round?.[String(r)]??0;
      const cls = !hasData?'vpill-grey':v>0?'vpill-red':'vpill-green';
      return `<span class="vpill ${cls}">R${r}: ${hasData?`${v} (${(p||0).toFixed(1)}%)`:'N/A'}</span>`;
    }).join('');
    return `<div class="viol-row ${rowCls}">
      <div class="viol-icon">${icon}</div>
      <div class="viol-text">
        <strong>${x.rule}</strong>
        <div class="viol-path">${x.variable}</div>
        <div class="viol-note">${x.note||''}</div>
        <div class="viol-pills">${pills}</div>
      </div>
    </div>`;
  }).join('');

  // Count visible checks (exclude fully-green with no N/A)
  const visibleSkip = skipIssues.filter(x=>{
    const total = Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
    const hasNA = ROUNDS.some(r=>(x.counts_by_round[r]??x.counts_by_round[String(r)]??null)===null);
    return total>0 || hasNA;
  });
  if(visibleSkip.length>0){
    const nWithViol = visibleSkip.filter(x=>Object.values(x.counts_by_round).some(v=>v>0)).length;
    html += `
  <div class="card">
    <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
      <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${totalSkipViol>0?'#e74c3c':'#27ae60'}">▼</span>
      Skip Pattern Violations
      <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
        <span style="color:#888">${visibleSkip.length} checks</span>
        <span style="color:${totalSkipViol>0?'#e74c3c':'#27ae60'}">${totalSkipViol} violations</span>
      </span>
    </h2>
    <div class="collapsible-body">
    <p class="subtitle">Questions filled despite gate answer routing away from them</p>
    <div class="note-warn note-box">
      E.g. A1=2 routes to A26, skipping A10/A11. If A10 is filled, it is a violation.
    </div>
    <div class="stats-row">
      <div class="stat-box ${totalSkipViol>0?'red':'green'}"><div class="num">${totalSkipViol}</div><div class="lbl">Total violations</div></div>
      <div class="stat-box ${nWithViol>0?'yellow':'green'}"><div class="num">${nWithViol}</div><div class="lbl">Checks with violations</div></div>
      <div class="stat-box blue"><div class="num">${visibleSkip.length}</div><div class="lbl">Visible checks</div></div>
    </div>
    ${skipList}
  </div></div>`;
  }

  // ─────── OUT-OF-RANGE SECTION ───────
  // Hide fully-green checks (all rounds 0, no N/A). Keep checks with
  // N/A rounds or actual violations visible.
  const oorList = oorIssues.map(x=>{
    const total=Object.values(x.counts||{}).reduce((a,v)=>a+(v||0),0);
    const hasNA = ROUNDS.some(r=>(x.counts?.[r]??x.counts?.[String(r)]??null)===null);
    if(total===0 && !hasNA) return '';
    const sev=x.severity||'medium';
    const icon = total===0?'🟢':sev==='high'?'🔴':'🟡';
    const rowCls = total===0?'clean':sev;
    const pills = ROUNDS.map(r=>{
      const raw = x.counts?.[r]??x.counts?.[String(r)]??null;
      const hasData = raw!=null;
      const v = raw||0;
      const cls = !hasData?'vpill-grey':v>0?'vpill-red':'vpill-green';
      return `<span class="vpill ${cls}">R${r}: ${hasData?v:'N/A'}</span>`;
    }).join('');
    return `<div class="viol-row ${rowCls}">
      <div class="viol-icon">${icon}</div>
      <div class="viol-text">
        <strong>${_varLabel(x.variable,vmap)}</strong>
        <div class="viol-path">${x.rule}</div>
        <div class="viol-note">${x.label||''}</div>
        <div class="viol-pills">${pills}</div>
      </div>
    </div>`;
  }).join('');

  const visibleOOR = oorIssues.filter(x=>{
    const total=Object.values(x.counts||{}).reduce((a,v)=>a+(v||0),0);
    const hasNA = ROUNDS.some(r=>(x.counts?.[r]??x.counts?.[String(r)]??null)===null);
    return total>0 || hasNA;
  });
  const totalOOR = oorIssues.reduce((s,x)=>s+Object.values(x.counts||{}).reduce((a,v)=>a+(v||0),0),0);
  if(visibleOOR.length>0){
    html += `
  <div class="card">
    <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
      <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${totalOOR>0?'#f39c12':'#27ae60'}">▼</span>
      Out-of-Range Values
      <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
        <span style="color:#888">${visibleOOR.length} checks</span>
        <span style="color:${totalOOR>0?'#e67e22':'#27ae60'}">${totalOOR} out-of-range</span>
      </span>
    </h2>
    <div class="collapsible-body">
    <p class="subtitle">Values outside specified bounds in questionnaire</p>
    <div class="note-info note-box">
      Valid ranges from questionnaire notes, e.g. "A10: 0–7 days", "A11: 0–168 hours".
    </div>
    <div class="stats-row">
      <div class="stat-box ${totalOOR>0?'yellow':'green'}"><div class="num">${totalOOR}</div><div class="lbl">Total out-of-range</div></div>
      <div class="stat-box blue"><div class="num">${visibleOOR.length}</div><div class="lbl">Visible checks</div></div>
    </div>
    ${oorList}
  </div></div>`;
  }

  // ─────── MANDATORY MISSING SECTION ───────
  const mandList = mandIssues.map(x=>{
    const total = Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
    if(total===0) return '';  // All green → hide
    const sev = x.severity||'medium';
    const icon = sev==='high'?'🔴':'🟡';
    const rowCls = sev;
    const pills = ROUNDS.map(r=>{
      const raw = x.counts_by_round[r]??x.counts_by_round[String(r)]??null;
      const hasData = raw!=null;
      const v = raw||0;
      const cls = !hasData?'vpill-grey':v>0?'vpill-red':'vpill-green';
      return `<span class="vpill ${cls}">R${r}: ${hasData?v:'N/A'}</span>`;
    }).join('');
    return `<div class="viol-row ${rowCls}">
      <div class="viol-icon">${icon}</div>
      <div class="viol-text">
        <strong>${x.rule}</strong>
        <div class="viol-path">${x.variable}</div>
        <div class="viol-note">${x.note||''}</div>
        <div class="viol-pills">${pills}</div>
      </div>
    </div>`;
  }).join('');

  if(mandIssues.length>0){
    html += `
  <div class="card">
    <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
      <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${totalMandMiss>0?'#f39c12':'#27ae60'}">▼</span>
      Mandatory Fields — Unexpectedly Missing
      <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
        <span style="color:#888">${mandIssues.length} checks</span>
        <span style="color:${totalMandMiss>0?'#e67e22':'#27ae60'}">${totalMandMiss} missing</span>
      </span>
    </h2>
    <div class="collapsible-body">
    <p class="subtitle">Follow-up questions that should be filled but are blank</p>
    <div class="note-warn note-box">
      E.g. H2=1/2/3 means care was needed → H2A must be answered. Blank H2A is a failure.
    </div>
    <div class="stats-row">
      <div class="stat-box ${totalMandMiss>0?'yellow':'green'}"><div class="num">${totalMandMiss}</div><div class="lbl">Total failures</div></div>
      <div class="stat-box blue"><div class="num">${mandIssues.length}</div><div class="lbl">Total checks</div></div>
    </div>
    ${mandList}
  </div></div>`;
  }

  // ─────── INTERVIEW QUALITY SECTION (M00 ONLY) ───────
  if(mod==='M00' && DQ.interview_meta){
    const intMeta = DQ.interview_meta;
    html += `
  <div class="card">
    <h2>Interview Quality</h2>
    <p class="subtitle">Duration distributions, partial interviews, call attempts</p>
    <div class="stats-row">
      <div class="stat-box blue"><div class="num">${Math.round(intMeta.median_duration||0)}</div><div class="lbl">Median duration (min)</div></div>
      <div class="stat-box yellow"><div class="num">${intMeta.partial_count||0}</div><div class="lbl">Partial interviews</div></div>
      <div class="stat-box blue"><div class="num">${Math.round(intMeta.mean_call_attempts||0)}</div><div class="lbl">Mean call attempts</div></div>
    </div>
  </div>`;
  }


  // Missing data mini-heatmap for this module
  if(heatRows.length>0){
    html+=`<div class="card">
      <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
        <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${heatRows.some(r=>r.rag==='red')?'#e74c3c':heatRows.some(r=>r.rag==='yellow')?'#f39c12':'#27ae60'}">▼</span>
        Missing Data — ${mod} Variables
        <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
          <span style="color:#888">${heatRows.length} variables</span>
          <span style="color:${heatRows.some(r=>r.rag==='red')?'#e74c3c':heatRows.some(r=>r.rag==='yellow')?'#e67e22':'#27ae60'}">${heatRows.filter(r=>r.rag==='red').length} red · ${heatRows.filter(r=>r.rag==='yellow').length} yellow · ${heatRows.filter(r=>r.rag==='green').length} green</span>
        </span>
      </h2>
      <div class="collapsible-body">
      <div class="heatmap-wrap"><table class="heatmap"><thead><tr>
        <th class="left">Variable</th>
        ${ROUNDS.map(r=>`<th>R${r}</th>`).join('')}
        <th>RAG</th></tr></thead><tbody>`;
    heatRows.forEach(row2=>{
      const rag=row2.rag||'green';
      let varLabel=_varLabel(row2.var,vmap);
      if(row2.sata) varLabel+=` <span class="gate-tag" title="Select-all-that-apply: missing = no sub-item answered">SATA</span>`;
      if(row2.conditional) varLabel+=` <span class="gate-tag" title="Conditional: missing rate computed only among ${row2.gate}">if ${row2.gate}</span>`;
      html+=`<tr><td class="vn">${varLabel}</td>`;
      const sv=row2.skip_viol||{};
      ROUNDS.forEach(r=>{
        const v=row2[r];
        const skipN=sv[String(r)]||0;
        let cell=`<div class="hm-cell" style="background:${hmColor(v)}">${hmText(v)}`;
        if(skipN>0) cell+=` <span class="skip-badge" title="${skipN} skip violation${skipN>1?'s':''}: filled when skip logic says empty">⚠${skipN}</span>`;
        cell+=`</div>`;
        html+=`<td>${cell}</td>`;
      });
      const ragDot={red:'🔴',yellow:'🟡',green:'🟢'}[rag]||'⚪';
      html+=`<td>${ragDot}</td></tr>`;
    });
    html+=`</tbody></table></div></div></div>`;
  }

  // ─────── INTERVIEW QUALITY BY INTERVIEWER ───────
  const iqData = DQ.interview_meta?.interviewer_quality?.[mod];
  if(iqData && iqData.interviewers && iqData.interviewers.length>0){
    const ints = iqData.interviewers;
    const gMed = iqData.global_median;
    const gP25 = iqData.global_p25;
    const gP75 = iqData.global_p75;
    const nFast = ints.filter(i=>i.flag==='fast').length;
    const nSlow = ints.filter(i=>i.flag==='slow').length;
    const nNorm = ints.filter(i=>i.flag==='normal').length;
    const durLabel = iqData.dur_col;

    const sortBtns = ['overall',...ROUNDS.map(r=>'R'+r)].map(s=>{
      const isAct = s==='overall';
      return `<button class="iq-sort-btn${isAct?' active':''}" onclick="sortIQTable('${mod}','${s}')"
        style="padding:2px 10px;border:1.5px solid #2d3f55;border-radius:3px;font-size:11px;cursor:pointer;
               background:${isAct?'#2d3f55':'#f8f9fa'};color:${isAct?'#fff':'#2d3f55'}">${s==='overall'?'Overall':s}</button>`;
    }).join('');

    html+=`<div class="card">
      <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
        <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${nFast+nSlow>10?'#f39c12':'#27ae60'}">▼</span>
        Interview Quality — ${MOD_NAMES[mod]} Duration by Interviewer
        <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
          <span style="color:#888">${ints.length} interviewers · ${durLabel}</span>
          <span style="color:#e74c3c">${nFast} fast</span>
          <span style="color:#27ae60">${nNorm} normal</span>
          <span style="color:#3498db">${nSlow} slow</span>
        </span>
      </h2>
      <div class="collapsible-body">
      <div class="note-info note-box" style="font-size:11px">
        Global median: <strong>${gMed} min</strong> · P25: ${gP25} min · P75: ${gP75} min.
        <span style="color:#e74c3c">Fast</span> = below P25,
        <span style="color:#27ae60">Normal</span> = P25–P75,
        <span style="color:#3498db">Slow</span> = above P75.
        Interviewers with unusually fast times may indicate quality concerns (rushing, skipping).
      </div>
      <div style="display:flex;gap:6px;align-items:center;margin-bottom:8px;flex-wrap:wrap">
        <span style="font-size:11px;color:#888;font-weight:600">Sort by:</span>
        ${sortBtns}
      </div>
      <div id="iq-table-${mod}"></div>
      </div></div>`;
  }

  // ─────── KOBO SKIP LOGIC MAP ───────
  const koboMod = KOBO[mod];
  if(koboMod && koboMod.variables && koboMod.variables.length>0){
    const vars = koboMod.variables.slice().sort((a,b)=>_varSort(a.name,vmap)-_varSort(b.name,vmap));
    const nWithSkip = vars.filter(v=>ROUNDS.some(r=>{const rr=v.rules_by_round[r]; return rr&&rr.relevant})).length;
    const nReq = vars.filter(v=>ROUNDS.some(r=>{const rr=v.rules_by_round[r]; return rr&&rr.required})).length;
    const nConst = vars.filter(v=>ROUNDS.some(r=>{const rr=v.rules_by_round[r]; return rr&&rr.constraint})).length;
    const nChanged = vars.filter(v=>{
      const rules = ROUNDS.map(r=>v.rules_by_round[r]).filter(Boolean);
      if(rules.length<2) return false;
      return rules.some((r,i)=>i>0 && (r.relevant||'')!==(rules[0].relevant||''));
    }).length;
    // Detect type changes across rounds (e.g. select_one → select_multiple)
    function _hasTypeChange(v){
      const types = ROUNDS.map(r=>{const rr=v.rules_by_round[r]; return rr&&rr.type ? (rr.type+'').split(' ')[0] : null}).filter(Boolean);
      if(types.length<2) return false;
      return types.some((t,i)=>i>0 && t!==types[0]);
    }
    const nTypeChanged = vars.filter(v=>_hasTypeChange(v)).length;

    html+=`<div class="card">
      <h2 style="cursor:pointer;user-select:none" onclick="toggleSection(this)">
        <span class="collapse-arrow" style="display:inline-block;transition:transform 0.2s;margin-right:4px;color:${nChanged>0?'#f39c12':'#27ae60'}">▼</span>
        Kobo Skip Logic Map
        <span class="section-summary" style="margin-left:auto;display:inline-flex;gap:8px;font-size:11px;font-weight:500">
          <span style="color:#888">${vars.length} variables</span>
          <span style="color:${nChanged>0?'#e67e22':'#27ae60'}">${nChanged} changed</span>
        </span>
      </h2>
      <div class="collapsible-body">
      <p class="subtitle">Complete routing rules from KoboToolbox XLSForm questionnaires (R1–R${ROUNDS.length})</p>
      <div class="stats-row" style="margin-bottom:10px">
        <div class="stat-box blue"><div class="num">${vars.length}</div><div class="lbl">Total variables</div></div>
        <div class="stat-box blue"><div class="num">${nWithSkip}</div><div class="lbl">With skip logic</div></div>
        <div class="stat-box blue"><div class="num">${nReq}</div><div class="lbl">Required</div></div>
        <div class="stat-box ${nConst>0?'purple':'blue'}"><div class="num">${nConst}</div><div class="lbl">With constraints</div></div>
        <div class="stat-box ${nChanged>0?'yellow':'green'}"><div class="num">${nChanged}</div><div class="lbl">Skip logic changed</div></div>
        <div class="stat-box ${nTypeChanged>0?'purple':'blue'}"><div class="num">${nTypeChanged}</div><div class="lbl">Type changed</div></div>
      </div>
      <div class="kobo-filter-bar" id="kobo-filter-${mod}">
        <span style="font-size:10.5px;color:#888;font-weight:600">Show:</span>
        <button class="kobo-filter-btn active" onclick="filterKobo('${mod}','all')">All</button>
        <button class="kobo-filter-btn" onclick="filterKobo('${mod}','skip')">With skip logic</button>
        <button class="kobo-filter-btn" onclick="filterKobo('${mod}','changed')">Changed across rounds</button>
        <button class="kobo-filter-btn" onclick="filterKobo('${mod}','constraint')">With constraints</button>
        <button class="kobo-filter-btn" onclick="filterKobo('${mod}','typechange')">Type changed</button>
      </div>
      <div class="note-info note-box" style="margin-bottom:8px">
        <strong>Skip logic</strong> = Kobo <code>relevant</code> column (when is this question asked?).
        <strong>Req</strong> = mandatory field.
        <strong>Constraint</strong> = data validation rule.
        Yellow rows = skip logic changed across rounds.
        <span style="color:#8e44ad;font-weight:600">Purple type</span> = question type changed across rounds (e.g. text to select_one).
      </div>
      <div class="heatmap-wrap">
      <table class="kobo-table" id="kobo-table-${mod}">
      <thead>
        <tr>
          <th class="left" style="min-width:90px">Variable</th>
          <th class="left" style="min-width:60px">Type</th>
          ${ROUNDS.map((r,i)=>`<th class="left" style="min-width:180px">R${r}${i===0?' Skip / Req / Constraint':''}</th>`).join('')}
        </tr>
      </thead>
      <tbody>`;

    // ── Inline diff helper: highlight parts of B that differ from A ──
    function highlightDiff(a, b){
      if(!a || !b || a===b) return b;
      // Word-level diff: split on boundaries, mark changed words
      const wa = a.match(/[\\w$.{}()=]+|[^\\w$.{}()=]+/g)||[];
      const wb = b.match(/[\\w$.{}()=]+|[^\\w$.{}()=]+/g)||[];
      // Simple LCS-based diff
      const m=wa.length, nn=wb.length;
      const dp=Array.from({length:m+1},()=>new Uint16Array(nn+1));
      for(let i=1;i<=m;i++) for(let j=1;j<=nn;j++)
        dp[i][j]=wa[i-1]===wb[j-1]?dp[i-1][j-1]+1:Math.max(dp[i-1][j],dp[i][j-1]);
      // Backtrack to find which words in B are new
      const keep=new Set();
      let i=m,j=nn;
      while(i>0&&j>0){
        if(wa[i-1]===wb[j-1]){keep.add(j-1);i--;j--}
        else if(dp[i-1][j]>=dp[i][j-1]) i--;
        else j--;
      }
      return wb.map((w,idx)=>keep.has(idx)?w:`<span class="kobo-diff">${w}</span>`).join('');
    }

    vars.forEach((v,vi)=>{
      // Detect if skip logic changed across rounds
      const existingRules = ROUNDS.map(r=>v.rules_by_round[r]).filter(Boolean);
      const hasSkipChange = existingRules.length>=2 && existingRules.some((rr,i)=>i>0 && (rr.relevant||'')!==(existingRules[0].relevant||''));
      const hasReqChange = existingRules.length>=2 && existingRules.some((rr,i)=>i>0 && (!!rr.required)!==(!!existingRules[0].required));
      const hasConsChange = existingRules.length>=2 && existingRules.some((rr,i)=>i>0 && (rr.constraint||'')!==(existingRules[0].constraint||''));
      const hasSkip = ROUNDS.some(r=>{const rr=v.rules_by_round[r]; return rr&&rr.relevant});
      const hasCons = ROUNDS.some(r=>{const rr=v.rules_by_round[r]; return rr&&rr.constraint});

      // Find reference texts (first round that has data)
      let refRelevant=null, refRequired=null, refConstraint=null;
      for(const rr of existingRules){
        if(refRelevant===null) refRelevant=rr.relevant||'';
        if(refRequired===null) refRequired=!!rr.required;
        if(refConstraint===null) refConstraint=rr.constraint||'';
        break;
      }

      // Detect type change across rounds
      const typeChanged = _hasTypeChange(v);

      // Data attributes for filtering
      const cls = [hasSkipChange?'kobo-changed':'', typeChanged?'kobo-type-changed':''].filter(Boolean).join(' ');
      const dSkip = hasSkip?'1':'0';
      const dChanged = hasSkipChange?'1':'0';
      const dCons = hasCons?'1':'0';
      const dType = typeChanged?'1':'0';

      html+=`<tr class="${cls}" data-skip="${dSkip}" data-changed="${dChanged}" data-cons="${dCons}" data-typechange="${dType}">`;
      html+=`<td class="kv-name">${_varLabel(v.name,vmap)}</td>`;
      // Show per-round types if type changed, otherwise single type
      if(typeChanged){
        const perRound = ROUNDS.map(r=>{const rr=v.rules_by_round[r]; return rr&&rr.type ? (rr.type+'').split(' ')[0] : null});
        const typeStr = perRound.map((t,i)=>t ? `<span style="font-size:9px">R${i+1}:</span>${t}` : null).filter(Boolean).join('<br>');
        html+=`<td class="kv-type" style="color:#8e44ad;font-weight:700" title="Type changed across rounds">${typeStr}</td>`;
      } else {
        html+=`<td class="kv-type">${(v.type||'').split(' ')[0]}</td>`;
      }

      let prevRelevant=refRelevant, prevRequired=refRequired, prevConstraint=refConstraint;

      ROUNDS.forEach((r,ri)=>{
        const rr = v.rules_by_round[r];
        if(!rr){
          html+=`<td><span class="kobo-na">not in R${r}</span></td>`;
          return;
        }
        let cellParts=[];
        const curRelevant = rr.relevant||'';
        if(curRelevant){
          // Highlight diff if skip logic changed from previous round
          const diffed = (hasSkipChange && prevRelevant!==null && curRelevant!==prevRelevant)
            ? highlightDiff(prevRelevant, curRelevant) : curRelevant;
          cellParts.push(`<div class="kobo-rule skip" title="Skip logic (relevant)">${diffed}</div>`);
        }
        const reqDot = rr.required ?
          '<span class="kobo-req-dot yes" title="Required"></span><span style="font-size:10px;color:#155724;font-weight:600">Required</span>' :
          '<span class="kobo-req-dot no" title="Not required"></span><span style="font-size:10px;color:#aaa">Optional</span>';
        cellParts.push(`<div style="margin:2px 0">${reqDot}</div>`);
        const curConstraint = rr.constraint||'';
        if(curConstraint){
          const diffedCons = (hasConsChange && prevConstraint!==null && curConstraint!==prevConstraint)
            ? highlightDiff(prevConstraint, curConstraint) : curConstraint;
          cellParts.push(`<div class="kobo-rule constraint" title="Constraint: ${(rr.constraint_message||'').replace(/"/g,"'")}">${diffedCons}</div>`);
        }
        html+=`<td>${cellParts.join('')}</td>`;
        // Update previous for next round comparison
        if(curRelevant) prevRelevant=curRelevant;
        prevRequired=!!rr.required;
        if(curConstraint) prevConstraint=curConstraint;
      });
      html+=`</tr>`;
    });

    html+=`</tbody></table></div></div></div>`;
  }

  return html;
}

function buildDetailPanel(row, dq, mod){
  const rounds = ['R1','R2','R3','R4','R5','R6','R7','R8'];

  // English text (most recent)
  const engText = row.english_text||'';

  // ── Response options with cross-round change tracking ──
  let codesHTML = '';
  const varName = (row.variable||'').toUpperCase();
  // Collect codes from all rounds
  const allRoundCodes = {};
  for (const rnd of rounds){
    // Try from full_codes stored in module_tables first, then fall back to AQ
    const fc = row[`full_codes_${rnd.toLowerCase()}`];
    if(fc && fc.length){
      allRoundCodes[rnd] = fc;
    } else {
      const modQs = (AQ[rnd]||{})[mod]||[];
      const match = modQs.find(q=>(q.qnum||'').toUpperCase()===varName || (q.qnum||'').toUpperCase()===varName+'_');
      if(match && match.codes && match.codes.length) allRoundCodes[rnd] = match.codes;
    }
  }
  // Find the most complete round for display
  const bestRnd = ['R5','R4','R3','R2','R1'].find(r=>allRoundCodes[r])||'';
  if(bestRnd && allRoundCodes[bestRnd]){
    const bestCodes = allRoundCodes[bestRnd];
    // Build a map of which codes exist in which rounds
    const codePresence = {};  // code → {round → label}
    for(const rnd of rounds){
      (allRoundCodes[rnd]||[]).forEach(c=>{ if(!codePresence[c.code]) codePresence[c.code]={}; codePresence[c.code][rnd]=c.label; });
    }
    // Detect per-code: added in which round, label changes
    codesHTML=`<div style="margin-top:6px"><strong>Response Options:</strong><table style="margin:4px 0 0 8px;border-collapse:collapse;font-size:11.5px">`;
    bestCodes.forEach(c=>{
      const pres = codePresence[c.code]||{};
      const presRounds = rounds.filter(r=>pres[r]);
      const firstRnd = presRounds[0]||'';
      const isNew = firstRnd && firstRnd !== rounds[0] && allRoundCodes[rounds[0]];  // Added after R1
      // Check label changes
      const labels = [...new Set(Object.values(pres))];
      const hasLabelChange = labels.length > 1;
      let annotation = '';
      if(isNew) annotation += `<span style="color:#155724;font-size:10px;font-weight:600;margin-left:6px;background:#d4edda;padding:1px 5px;border-radius:8px">+ Added ${firstRnd}</span>`;
      if(hasLabelChange){
        const changes = [];
        for(let i=1;i<rounds.length;i++){
          const pr=rounds[i-1], r=rounds[i];
          if(pres[pr]&&pres[r]&&pres[pr]!==pres[r]) changes.push(`${pr}→${r}`);
        }
        if(changes.length) annotation += `<span style="color:#856404;font-size:10px;font-weight:600;margin-left:6px;background:#fff3cd;padding:1px 5px;border-radius:8px">✏️ Relabeled ${changes.join(', ')}</span>`;
      }
      codesHTML+=`<tr><td style="padding:1px 10px 1px 0;font-weight:600;color:#1a5276;font-family:monospace">${c.code}</td><td style="padding:1px 0">${c.label}${annotation}</td></tr>`;
    });
    codesHTML+=`</table></div>`;
  }

  // Skip logic per round
  let skipHTML='';
  const hasSkip = rounds.some(r=>row[`skip_${r.toLowerCase()}`]);
  if(hasSkip){
    skipHTML=`<div class="skip-rules"><strong>Skip Logic by Round:</strong>`;
    rounds.forEach(r=>{
      const sr = row[`skip_${r.toLowerCase()}`];
      if(sr) skipHTML+=`<div class="skip-rule-row">
        <span class="skip-round">${r}</span>
        <span class="skip-text">${sr}</span>
      </div>`;
    });
    skipHTML+=`</div>`;
  }

  // Data check notes
  const dcNotes = ['R3','R4','R5'].map(r=>row[`data_check_${r.toLowerCase()}`]).filter(Boolean);
  let dcHTML='';
  if(dcNotes.length){
    dcHTML=`<div style="margin-top:6px"><strong>Data Check Notes (from questionnaire):</strong><ul style="margin:3px 0 0 16px">`;
    dcNotes.forEach(n=>dcHTML+=`<li style="font-size:11px;color:#1a5276">${n}</li>`);
    dcHTML+=`</ul></div>`;
  }

  // DQ issue details
  let dqHTML='';
  if(dq.skip.length||dq.mand.length||dq.oor.length){
    dqHTML=`<div style="margin-top:6px"><strong>Data Quality Issues (from actual data):</strong>`;
    dq.skip.forEach(x=>{
      const total=Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
      dqHTML+=`<div class="dq-inline dq-skip" style="display:block;border-radius:5px;margin:3px 0;padding:4px 8px;max-width:100%">
        <strong>Skip violation:</strong> ${x.rule} — total ${total} rows
        <div style="font-size:10px;margin-top:2px">${[1,2,3,4,5,6,7].map(r=>{const v=x.counts_by_round[r]??x.counts_by_round[String(r)]??0;return `R${r}:${v}`}).join(' · ')}</div>
      </div>`;
    });
    dq.mand.forEach(x=>{
      const total=Object.values(x.counts_by_round).reduce((a,v)=>a+(v||0),0);
      dqHTML+=`<div class="dq-inline dq-mand" style="display:block;border-radius:5px;margin:3px 0;padding:4px 8px;max-width:100%">
        <strong>Mandatory missing:</strong> ${x.rule} — total ${total} rows
        <div style="font-size:10px;margin-top:2px">${[1,2,3,4,5,6,7].map(r=>{const v=x.counts_by_round[r]??x.counts_by_round[String(r)]??0;return `R${r}:${v}`}).join(' · ')}</div>
      </div>`;
    });
    dq.oor.forEach(x=>{
      dqHTML+=`<div class="dq-inline dq-oor" style="display:block;border-radius:5px;margin:3px 0;padding:4px 8px;max-width:100%">
        <strong>Out-of-range:</strong> ${x.rule} — ${x.label||x.variable}
        <div style="font-size:10px;margin-top:2px">${[1,2,3,4,5,6,7].map(r=>{const v=x.counts?.[r]??x.counts?.[String(r)]??0;return `R${r}:${v}`}).join(' · ')}</div>
      </div>`;
    });
    dqHTML+=`</div>`;
  }

  return `
    <dl>
      <dt>Title</dt><dd>${row.question_title||'—'}</dd>
      <dt>Type</dt><dd>${row.question_type||'—'}</dd>
      <dt>Status</dt><dd>${row.status||'—'}</dd>
      <dt>Rounds</dt><dd>${['R1','R2','R3','R4','R5','R6','R7','R8'].filter(r=>row[`in_${r}`]==='✓').join(', ')}</dd>
      ${row.english_text?`<dt>Question Text</dt><dd style="font-style:italic">${engText}</dd>`:''}
      ${row.title_changes?`<dt>Title Change</dt><dd style="color:#856404">${row.title_changes}</dd>`:''}
      ${row.option_changes?`<dt>Option Changes</dt><dd style="color:#4a235a">${row.option_changes}</dd>`:''}
      ${row.remarks?`<dt>Remarks</dt><dd style="color:#555">${row.remarks}</dd>`:''}
    </dl>
    ${codesHTML}${skipHTML}${dcHTML}${dqHTML}`;
}

function toggleSection(headerEl){
  const body = headerEl.parentElement.querySelector('.collapsible-body');
  if(!body) return;
  const arrow = headerEl.querySelector('.collapse-arrow');
  const isOpen = body.style.display !== 'none';
  body.style.display = isOpen ? 'none' : '';
  if(arrow) arrow.style.transform = isOpen ? 'rotate(-90deg)' : '';
}

function toggleDetail(id){
  const el=document.getElementById(id);
  if(el) el.classList.toggle('open');
}
function toggleAllDetails(mod){
  const table=document.getElementById(`qtable-${mod}`);
  if(!table)return;
  const panels=table.querySelectorAll('.q-detail');
  const anyOpen=[...panels].some(p=>p.classList.contains('open'));
  panels.forEach(p=>p.classList.toggle('open',!anyOpen));
}

function filterKobo(mod, mode){
  const table=document.getElementById(`kobo-table-${mod}`);
  if(!table) return;
  const rows=table.querySelectorAll('tbody tr');
  rows.forEach(row=>{
    if(mode==='all') { row.style.display=''; return; }
    if(mode==='skip')       row.style.display=row.dataset.skip==='1'?'':'none';
    if(mode==='changed')    row.style.display=row.dataset.changed==='1'?'':'none';
    if(mode==='constraint') row.style.display=row.dataset.cons==='1'?'':'none';
    if(mode==='typechange') row.style.display=row.dataset.typechange==='1'?'':'none';
  });
  // Update button states
  const bar=document.getElementById(`kobo-filter-${mod}`);
  if(bar) bar.querySelectorAll('.kobo-filter-btn').forEach(btn=>{
    const t=btn.textContent.toLowerCase();
    const match = mode==='all'?t.includes('all') : mode==='skip'?t.includes('skip') : mode==='changed'?t.includes('changed across') : mode==='constraint'?t.includes('constraint') : mode==='typechange'?t.includes('type') : false;
    btn.classList.toggle('active', match);
  });
}

// ── INTERVIEW QUALITY SORT BY ROUND ──────────────────────────────────────────
function sortIQTable(mod, sortKey){
  const iqData = DQ.interview_meta?.interviewer_quality?.[mod];
  if(!iqData) return;
  const container = document.getElementById('iq-table-'+mod);
  if(!container) return;
  const ints = iqData.interviewers.slice();
  const gP25 = iqData.global_p25, gP75 = iqData.global_p75;

  // Sort
  if(sortKey==='overall'){
    ints.sort((a,b)=>a.median-b.median);
  } else {
    const rKey = sortKey.replace('R','');
    ints.sort((a,b)=>{
      const aR = a.by_round?.[rKey];
      const bR = b.by_round?.[rKey];
      const aV = aR ? aR.median : 9999;
      const bV = bR ? bR.median : 9999;
      return aV - bV;
    });
  }

  // Update button styles
  const card = container.closest('.card');
  if(card) card.querySelectorAll('.iq-sort-btn').forEach(btn=>{
    const bKey = btn.textContent.trim()==='Overall'?'overall':btn.textContent.trim();
    const isAct = bKey===sortKey;
    btn.classList.toggle('active', isAct);
    btn.style.background = isAct?'#2d3f55':'#f8f9fa';
    btn.style.color = isAct?'#fff':'#2d3f55';
  });

  // Render table
  let h = '<div class="heatmap-wrap"><table class="heatmap"><thead><tr>';
  h += '<th class="left">Interviewer</th><th>N</th><th>Median (min)</th><th>Min</th><th>Max</th>';
  ROUNDS.forEach(r=>{ h += '<th' + (sortKey===('R'+r)?' style="background:#e8f4fd"':'') + '>R'+r+'</th>'; });
  h += '<th>Flag</th></tr></thead><tbody>';

  ints.forEach(i=>{
    const flagColor = i.flag==='fast'?'#e74c3c':i.flag==='slow'?'#3498db':'#27ae60';
    const flagIcon = i.flag==='fast'?'\\u26A1':i.flag==='slow'?'\\uD83D\\uDC22':'\\u2713';
    const bgColor = i.flag==='fast'?'rgba(231,76,60,0.06)':i.flag==='slow'?'rgba(52,152,219,0.06)':'';
    h += '<tr style="'+(bgColor?'background:'+bgColor:'')+'">';
    h += '<td class="vn" style="font-weight:600">'+i.int_id+'</td>';
    h += '<td style="text-align:center">'+i.n+'</td>';
    h += '<td style="text-align:center;font-weight:600;color:'+flagColor+'">'+i.median+'</td>';
    h += '<td style="text-align:center;font-size:11px;color:#888">'+i.min+'</td>';
    h += '<td style="text-align:center;font-size:11px;color:#888">'+i.max+'</td>';
    ROUNDS.forEach(r=>{
      const rd = i.by_round?.[String(r)];
      const isSort = sortKey===('R'+r);
      if(rd){
        const rdColor = rd.median < gP25 ? '#e74c3c' : rd.median > gP75 ? '#3498db' : '#27ae60';
        h += '<td style="text-align:center'+(isSort?';background:#e8f4fd':'')+'"><span style="color:'+rdColor+';font-weight:500">'+rd.median+'</span> <span style="font-size:9px;color:#aaa">('+rd.n+')</span></td>';
      } else {
        h += '<td style="text-align:center'+(isSort?';background:#e8f4fd':'')+';color:#ccc">\\u2014</td>';
      }
    });
    h += '<td style="text-align:center"><span style="color:'+flagColor+';font-weight:600">'+flagIcon+' '+i.flag+'</span></td></tr>';
  });
  h += '</tbody></table></div>';
  container.innerHTML = h;
}

// Initialize all IQ tables after module pages are built
function initIQTables(){
  MODULES.forEach(mod=>{
    if(DQ.interview_meta?.interviewer_quality?.[mod]) sortIQTable(mod,'overall');
  });
}

// ── INIT ──────────────────────────────────────────────────────────────────────
buildOverview();
buildChanges();
buildAllModulePages();
initIQTables();
buildPanel();
buildOperators();
</script>
</body>
</html>"""

out = _os.path.join(_OUTPUT, 'l2ph_dq_dashboard.html')
# Replace placeholders with actual JSON data
content = HTML
content = content.replace('""" + DQ   + """', DQ)
content = content.replace('""" + MT   + """', MT)
content = content.replace('""" + AQ   + """', AQ)
content = content.replace('""" + PAN  + """', PAN)
content = content.replace('""" + IVIEW+ """', IVIEW)
with open(out,'w') as f:
    f.write(content)
print(f'Generated: {out} ({round(os.path.getsize(out)/1024,1)} KB)')
