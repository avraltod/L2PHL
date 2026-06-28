#!/usr/bin/env python3
"""
L2PHL CATI Data Quality Pipeline
==================================
Regenerates the QC dashboard and Excel cross-round report.
Run from any directory — all paths are resolved relative to this file.

Usage:
  python3 update_pipeline.py [--dta] [--questionnaire] [--dofiles] [--all]

Flags:
  --dta           Re-run DQ checks from pooled .dta files in Analysis/HF/
  --questionnaire Re-parse questionnaire Excel files from CATI/Questionnaire/
  --dofiles       Re-parse Stata do-files from CATI/Round*/do/
  --all           Run all of the above (default if no flags given)

The HTML dashboard and Excel report are always regenerated at the end.

Outputs:
  output/l2ph_dq_dashboard.html
  output/L2PHL_Questionnaire_Cross_Round_Report.xlsx

Cache (intermediate JSON):
  cache/dq_data.json
  cache/panel_data.json
  cache/interviewer_data.json
  cache/all_questions.json
  cache/module_tables.json
  cache/do_modules.json
"""
import sys, os, subprocess, json, time, re, glob
from pathlib import Path
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────────────
HERE     = Path(__file__).resolve().parent          # Analysis/QC/
CATI     = HERE.parent.parent                       # CATI/
SCRIPTS  = HERE / 'scripts'
CACHE    = HERE / 'cache'
OUTPUT   = HERE / 'output'
HF_DIR   = HERE.parent / 'HF'                      # Analysis/HF/ — pooled .dta files
QUEST_DIR= CATI / 'Questionnaire'                  # CATI/Questionnaire/
ROUNDS_DIR= CATI                                   # CATI/Round*/do/

CACHE.mkdir(exist_ok=True)
OUTPUT.mkdir(exist_ok=True)


# ── Helpers ────────────────────────────────────────────────────────────────────
def log(msg, level='INFO'):
    icons = {'INFO': 'ℹ️ ', 'OK': '✅', 'WARN': '⚠️ ', 'ERROR': '❌', 'STEP': '🔷'}
    print(f"  {icons.get(level, '  ')}{msg}")

def step(n, msg):
    print(f"\n{'='*60}")
    print(f"  🔷 STEP {n}: {msg}")
    print(f"{'='*60}")

def run_script(name):
    result = subprocess.run(
        [sys.executable, str(SCRIPTS / name)],
        capture_output=True, text=True, cwd=str(HERE)
    )
    if result.returncode != 0:
        log(f"{name} failed:", 'ERROR')
        for line in result.stderr.splitlines()[-15:]:
            print(f"    {line}")
        return False
    for line in result.stdout.splitlines():
        print(f"    {line}")
    return True


# ── STEP 1 — Detect available data ────────────────────────────────────────────
def detect_files():
    step(1, "Detecting available files")

    # Pooled .dta files in Analysis/HF/
    dta_files = sorted(HF_DIR.glob('l2phl_M*.dta'))
    log(f"Pooled .dta modules in Analysis/HF/: {len(dta_files)}", 'OK' if dta_files else 'WARN')
    for f in dta_files:
        log(f"  {f.name}")

    # Per-round .dta files in Analysis/HF/R*/
    round_dta = {}
    for rdir in sorted(HF_DIR.glob('R*')):
        if rdir.is_dir():
            rdtas = sorted(rdir.glob('l2phl_*_M*.dta'))
            if rdtas:
                round_dta[rdir.name] = rdtas
                log(f"  {rdir.name}/: {len(rdtas)} module files")

    # Questionnaire Excel files
    quest_files = sorted(QUEST_DIR.glob('*.xlsx'))
    quest_files = [f for f in quest_files if 'questionnaire' in f.name.lower()
                   or ('cati' in f.name.lower() and 'r' in f.name.lower())]
    log(f"Questionnaire files in CATI/Questionnaire/: {len(quest_files)}", 'OK' if quest_files else 'WARN')
    for f in quest_files:
        log(f"  {f.name}")

    # Do-files in CATI/Round*/do/
    # Match both the AP master naming (L2PHL_CATI@…) and the firm delivery
    # naming (L2PH_CATI@…, e.g. R08's L2PH_CATI@R08@CB@…). Round dedup below
    # still prefers @AP when present, so R1–R7 stay on their AP masters.
    do_files = []
    for rdir in sorted(CATI.glob('Round*/do')):
        seen = set()
        for pat in ('L2PHL_CATI@*.do', 'L2PH_CATI@*.do'):
            for df in sorted(rdir.glob(pat)):
                if df not in seen:
                    seen.add(df)
                    do_files.append(df)
    log(f"Do-files in CATI/Round*/do/: {len(do_files)}", 'OK' if do_files else 'WARN')
    for f in do_files:
        log(f"  {f.parent.parent.name}/{f.name}")

    return dta_files, quest_files, do_files, round_dta


# ── STEP 2 — Rebuild DQ from pooled .dta files ────────────────────────────────
def rebuild_dq(dta_files):
    step(2, "Rebuilding DQ data from pooled .dta files")
    if not dta_files:
        log("No pooled .dta files found in Analysis/HF/ — skipping", 'WARN')
        return False
    ok = run_script('build_dq.py')
    if ok:
        size = (CACHE / 'dq_data.json').stat().st_size // 1024
        log(f"cache/dq_data.json written ({size} KB)", 'OK')
    return ok


# ── STEP 2b — Rebuild panel tracking data ─────────────────────────────────────
def rebuild_panel(dta_files):
    step('2b', "Rebuilding panel tracking data")
    passport = next((f for f in dta_files if 'M00' in f.name), None)
    if not passport:
        log("l2phl_M00_passport.dta not found — skipping panel build", 'WARN')
        return False
    ok = run_script('build_panel.py')
    if ok:
        size = (CACHE / 'panel_data.json').stat().st_size // 1024
        log(f"cache/panel_data.json written ({size} KB)", 'OK')
    return ok


# ── STEP 2c — Rebuild interviewer/operator performance data ───────────────────
def rebuild_interviewer(dta_files):
    step('2c', "Rebuilding operator performance data")
    passport = next((f for f in dta_files if 'M00' in f.name), None)
    if not passport:
        log("l2phl_M00_passport.dta not found — skipping interviewer build", 'WARN')
        return False
    ok = run_script('build_interviewer.py')
    if ok:
        ipath = CACHE / 'interviewer_data.json'
        if ipath.exists():
            size = ipath.stat().st_size // 1024
            log(f"cache/interviewer_data.json written ({size} KB)", 'OK')
    return ok


# ── Kobo utilities ────────────────────────────────────────────────────────────
def normalize_kobo_varname(var_name):
    """
    Normalize Kobo variable names to match module_tables format.
    - Strip trailing underscores
    - Remove numeric suffixes (_1, _2, etc.)
    - Remove _oth, _new suffixes
    - Uppercase
    Example: ED16_1_ → ED16, SH1b_ → SH1B, F17_new → F17
    """
    v = str(var_name).strip()
    # Strip trailing underscores
    v = v.rstrip('_')
    # Remove numeric suffixes like _1, _2, _oth, _new
    v = re.sub(r'_[0-9]+$', '', v)
    v = re.sub(r'_(oth|new)$', '', v, flags=re.IGNORECASE)
    return v.upper()

def parse_kobo_variable_order():
    """
    Parse all 5 Kobo XLSForm files and extract variable order per module.
    Returns: {mod: [var1, var2, ...]} where vars are normalized to uppercase.
    """
    KOBO_DIR = CATI / 'KOBO'
    if not KOBO_DIR.exists():
        log(f"KOBO directory not found at {KOBO_DIR}", 'WARN')
        return {}

    # Map Kobo group names to modules
    GROUP_TO_MOD = {
        'group_intro': 'M00',
        'group_demogs': 'M01',
        'group_educ': 'M02',
        'group_shocks': 'M03',
        'group_employment': 'M04',
        'group_employment_new': 'M04',
        'group_income': 'M05',
        'group_income_new': 'M05',
        'group_finance': 'M06',
        'group_health': 'M07',
        'group_h9': 'M07',
        'group_H12toH16': 'M07',
        'group_f08': 'M08',
        'group_opinions-views': 'M09',
        'group_opinions_views': 'M09',
    }

    kobo_files = list(KOBO_DIR.glob('*.xlsx'))
    if not kobo_files:
        log(f"No Kobo files found in {KOBO_DIR}", 'WARN')
        return {}

    # Extract variable order from all Kobo files
    # Use R5 (latest) as primary, fill gaps from earlier rounds
    import openpyxl
    kobo_var_order = defaultdict(lambda: defaultdict(list))

    for kobo_path in sorted(kobo_files):
        # Detect round
        fname_lower = kobo_path.name.lower()
        if 'r5' in fname_lower or 'r05' in fname_lower:
            rnd = 'R5'
        elif 'r4' in fname_lower or 'r04' in fname_lower:
            rnd = 'R4'
        elif 'r3' in fname_lower or 'r03' in fname_lower:
            rnd = 'R3'
        elif 'r2' in fname_lower or 'r02' in fname_lower or 'follow-up_r2' in fname_lower:
            rnd = 'R2'
        elif 'r1' in fname_lower or 'r01' in fname_lower or 'follow-up1' in fname_lower or 'phase 1' in fname_lower:
            rnd = 'R1'
        else:
            continue

        try:
            wb = openpyxl.load_workbook(str(kobo_path), read_only=True, data_only=True)
            if 'survey' not in wb.sheetnames:
                continue

            ws = wb['survey']
            rows = list(ws.values)
            current_group = None

            for row in rows:
                if not row or len(row) < 2:
                    continue

                row_type = str(row[0]).strip().lower() if row[0] else ''
                row_name = str(row[1]).strip() if row[1] else ''

                if row_type == 'begin_group':
                    current_group = row_name.lower()
                elif row_type == 'end_group':
                    current_group = None
                elif current_group and row_name and row_type:
                    # Skip non-question types
                    if row_type in ('begin_group', 'end_group', 'begin_repeat', 'end_repeat',
                                   'start', 'end', 'deviceid', 'calculate', 'note'):
                        continue

                    # Map group to module
                    mod = GROUP_TO_MOD.get(current_group, None)
                    if mod:
                        norm_var = normalize_kobo_varname(row_name)
                        if norm_var not in kobo_var_order[mod][rnd]:
                            kobo_var_order[mod][rnd].append(norm_var)

            wb.close()
            log(f"  {kobo_path.name} ({rnd}): parsed")

        except Exception as e:
            log(f"Error parsing {kobo_path.name}: {e}", 'WARN')

    # Build master ordering: use R5 as primary, fill gaps from earlier rounds
    kobo_master = {}
    for mod in ['M00', 'M01', 'M02', 'M03', 'M04', 'M05', 'M06', 'M07', 'M08', 'M09']:
        master_order = []
        seen = set()
        # R5 (latest) first
        for var in kobo_var_order[mod].get('R5', []):
            if var not in seen:
                master_order.append(var)
                seen.add(var)
        # Fill gaps from earlier rounds
        for rnd in ['R4', 'R3', 'R2', 'R1']:
            for var in kobo_var_order[mod].get(rnd, []):
                if var not in seen:
                    master_order.append(var)
                    seen.add(var)
        if master_order:
            kobo_master[mod] = master_order

    return kobo_master


# ── STEP 3 — Re-parse questionnaire Excel files ───────────────────────────────
def rebuild_questionnaire(quest_files):
    step(3, "Parsing questionnaire Excel files")
    if not quest_files:
        log("No questionnaire files found in CATI/Questionnaire/ — skipping", 'WARN')
        return False

    import openpyxl

    ROUND_MAP = {
        'R1': ['phase 1', 'phase1', 'cati phase', 'cati r1'],
        'R2': ['r2 questionnaire', 'cati r2'],
        'R3': ['r3 questionnaire', 'cati r3'],
        'R4': ['r4 questionnaire', 'cati r4'],
        'R5': ['r5 questionnaire', 'cati r5'],
        'R6': ['r6 questionnaire', 'cati r6'],
        'R7': ['r7 questionnaire', 'cati r7'],
    }
    SHEETS = {
        'Introduction': 'M00', 'Demographics': 'M01', 'Education': 'M02',
        'Shocks': 'M03', 'Employment': 'M04', 'Income': 'M05',
        'Finance': 'M06', 'Health': 'M07', 'Food & Non-Food': 'M08',
        'Opinions & Views': 'M09',
    }

    files_map = {}
    for fpath in sorted(quest_files, reverse=True):
        fl = fpath.name.lower()
        if 'trailer' in fl:
            continue
        for rnd, patterns in ROUND_MAP.items():
            if any(p in fl for p in patterns) and rnd not in files_map:
                files_map[rnd] = fpath

    log(f"Rounds mapped: {sorted(files_map.keys())}")
    for r, f in files_map.items():
        log(f"  {r}: {f.name}")

    if len(files_map) < 2:
        log("Need at least 2 questionnaire files — skipping", 'WARN')
        return False

    def get_col(headers, *names):
        for name in names:
            for i, h in enumerate(headers):
                if name.upper() in str(h).upper():
                    return i
        return None

    def val(row, idx):
        if idx is not None and idx < len(row) and row[idx] is not None:
            v = str(row[idx]).strip()
            return v if v and v.lower() not in ('none', 'nan') else ''
        return ''

    def extract_questions(wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.values)
        if not rows:
            return []
        header_idx = 0
        for i, row in enumerate(rows[:8]):
            rv = [str(c).upper().strip() if c else '' for c in row]
            if 'Q#' in rv or 'ENGLISH' in rv or any('QUESTION' in v for v in rv):
                header_idx = i
                break
        headers = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(rows[header_idx])]
        _qc = get_col(headers, 'Q#', 'Q #')
        qnum_col  = _qc if _qc is not None else 1
        title_col = get_col(headers, 'QUESTION TITLE', 'TITLE')
        prog_col  = get_col(headers, 'PROGRAMMING')
        type_col  = get_col(headers, 'TYPE OF QUESTION', 'TYPE')
        codes_col = get_col(headers, 'CODES')
        eng_col   = get_col(headers, 'ENGLISH')
        remarks_col = next((i for i, h in enumerate(headers)
                           if 'REMARK' in h.upper() and 'PSRC' not in h.upper()
                           and 'DATA CHECK' not in h.upper()), None)
        check_col = get_col(headers, 'DATA CHECK')

        questions = []
        current_q = None
        for row in rows[header_idx + 1:]:
            if not any(c for c in row):
                continue
            qnum = val(row, qnum_col)
            if qnum and re.match(r'^[A-Za-z][A-Za-z0-9_]*$', qnum) and qnum.upper() not in ('NONE', 'COL_0', 'COL_1'):
                current_q = {
                    'qnum': qnum, 'title': val(row, title_col),
                    'type': val(row, type_col), 'english': val(row, eng_col)[:300],
                    'codes': [], 'skip_rules': [], 'remarks': val(row, remarks_col)[:300],
                    'data_check': val(row, check_col)[:200], 'prog_note': '',
                }
                questions.append(current_q)
                codes = val(row, codes_col)
                if codes:
                    current_q['codes'].append({'code': codes, 'label': val(row, eng_col)[:100]})
                prog = val(row, prog_col)
                if prog and ('go to' in prog.lower() or 'skip' in prog.lower()):
                    current_q['skip_rules'].append(prog[:150])
                elif prog:
                    current_q['prog_note'] = prog[:150]
            elif current_q:
                codes = val(row, codes_col)
                eng   = val(row, eng_col)
                prog  = val(row, prog_col)
                if codes and eng:
                    current_q['codes'].append({'code': codes, 'label': eng[:100]})
                if prog and ('go to' in prog.lower() or 'if' in prog.lower()):
                    current_q['skip_rules'].append(
                        f"Code {codes}: {prog[:120]}" if codes else prog[:120])
                if val(row, remarks_col) and not current_q['remarks']:
                    current_q['remarks'] = val(row, remarks_col)[:300]
                if val(row, check_col) and not current_q['data_check']:
                    current_q['data_check'] = val(row, check_col)[:200]
        return questions

    ROUNDS = sorted(files_map.keys())
    all_qs = {}
    for rnd, fpath in files_map.items():
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        all_qs[rnd] = {}
        for sheet, mod in SHEETS.items():
            all_qs[rnd][mod] = extract_questions(wb, sheet)
        wb.close()
        n = sum(len(v) for v in all_qs[rnd].values())
        log(f"  {rnd}: {n} questions extracted")

    with open(CACHE / 'all_questions.json', 'w') as f:
        json.dump(all_qs, f, indent=2)
    log(f"cache/all_questions.json written ({(CACHE/'all_questions.json').stat().st_size//1024} KB)", 'OK')

    # R8 stand-in: no R8 questionnaire is in the repo yet (only on Drive), and the
    # R8 Kobo form is the R7 clone — so mirror R7 as R8 for the question tracker /
    # change panels so every section shows R8 (= R7). Replace when the real R8
    # questionnaire is added. (all_questions.json above stays truthful R1–R7.)
    if 'R7' in all_qs and 'R8' not in all_qs:
        all_qs['R8'] = all_qs['R7']
        if 'R8' not in ROUNDS:
            ROUNDS = sorted(ROUNDS + ['R8'])

    # Build module_tables
    MODULES = ['M00', 'M01', 'M02', 'M03', 'M04', 'M05', 'M06', 'M07', 'M08', 'M09']
    module_tables = {}
    for mod in MODULES:
        seen = {}
        order = []
        for r in ROUNDS:
            for q in all_qs.get(r, {}).get(mod, []):
                v = q['qnum'].upper()
                if v not in seen:
                    seen[v] = r
                    order.append(v)
        rows_out = []
        for v in order:
            row = {'variable': v, 'first_round': seen[v]}
            titles = {}; types = {}; skips = {}; eng = {}; codes = {}; remarks = {}; checks = {}
            full_codes = {}  # Store actual code lists per round for change detection
            for r in ROUNDS:
                q = next((x for x in all_qs.get(r, {}).get(mod, []) if x['qnum'].upper() == v), None)
                row[f'in_{r}'] = '✓' if q else ''
                if q:
                    titles[r]  = q.get('title', '')
                    types[r]   = q.get('type', '')
                    skips[r]   = '; '.join(q.get('skip_rules', []))[:200]
                    eng[r]     = q.get('english', '')[:200]
                    codes[r]   = len(q.get('codes', []))
                    full_codes[r] = q.get('codes', [])  # [{code, label}, ...]
                    remarks[r] = q.get('remarks', '')
                    checks[r]  = q.get('data_check', '')
            def best(d): return next((d[r] for r in reversed(ROUNDS) if d.get(r, '')), '')
            row['question_title'] = best(titles)
            row['question_type']  = best(types)
            row['english_text']   = best(eng)
            for r in ROUNDS:
                row[f'codes_{r.lower()}']      = codes.get(r, '')
                row[f'skip_{r.lower()}']       = skips.get(r, '')
                row[f'data_check_{r.lower()}'] = checks.get(r, '')
            title_changes = []
            for i, r in enumerate(ROUNDS[1:], 1):
                pr = ROUNDS[i - 1]
                if titles.get(r) and titles.get(pr) and titles[r] != titles[pr]:
                    title_changes.append(f"{pr}→{r}: '{titles[pr]}' → '{titles[r]}'")
            row['title_changes'] = ' | '.join(title_changes)
            skip_changes = []
            for i, r in enumerate(ROUNDS[1:], 1):
                pr = ROUNDS[i - 1]
                s_r = skips.get(r, ''); s_pr = skips.get(pr, '')
                if s_r != s_pr:
                    if s_r and not s_pr:   skip_changes.append(f"Added in {r}")
                    elif s_pr and not s_r: skip_changes.append(f"Removed in {r}")
                    else:                  skip_changes.append(f"Changed {pr}→{r}")
            row['skip_changes'] = ' | '.join(skip_changes)
            # ── Response option changes ──
            option_changes = []
            for i, r in enumerate(ROUNDS[1:], 1):
                pr = ROUNDS[i - 1]
                fc_r = full_codes.get(r, []); fc_pr = full_codes.get(pr, [])
                if not fc_r and not fc_pr: continue
                codes_r  = {str(c.get('code','')): c.get('label','') for c in fc_r}
                codes_pr = {str(c.get('code','')): c.get('label','') for c in fc_pr}
                added   = [c for c in codes_r  if c not in codes_pr]
                removed = [c for c in codes_pr if c not in codes_r]
                relabeled = [c for c in codes_r if c in codes_pr and codes_r[c] != codes_pr[c]]
                parts = []
                for c in added:
                    parts.append(f"{r}: +Code {c} '{codes_r[c]}'")
                for c in removed:
                    parts.append(f"{r}: -Code {c} '{codes_pr[c]}'")
                for c in relabeled:
                    parts.append(f"{r}: Code {c} relabeled '{codes_pr[c]}' → '{codes_r[c]}'")
                option_changes.extend(parts)
            row['option_changes'] = ' | '.join(option_changes)
            # Store full codes per round for dashboard display
            for r in ROUNDS:
                row[f'full_codes_{r.lower()}'] = full_codes.get(r, [])
            present = [r for r in ROUNDS if row.get(f'in_{r}') == '✓']
            if not present:             row['status'] = 'NOT FOUND'
            elif present == ROUNDS:     row['status'] = 'All rounds'
            elif present[0] != ROUNDS[0]: row['status'] = f'New in {present[0]}'
            elif present[-1] != ROUNDS[-1]: row['status'] = f'Dropped after {present[-1]}'
            else:                       row['status'] = 'All rounds'
            row['remarks'] = best(remarks)
            rows_out.append(row)
        module_tables[mod] = rows_out

    # ── Reorder variables by Kobo XLSForm order ─────────────────────────────────────
    log("Parsing Kobo XLSForm files for variable ordering...")
    kobo_master = parse_kobo_variable_order()
    if kobo_master:
        reordered_count = 0
        for mod in module_tables:
            if mod in kobo_master:
                kobo_order = kobo_master[mod]
                # Build a map of variables in kobo_order with their positions
                kobo_positions = {var: i for i, var in enumerate(kobo_order)}
                # Separate rows into ordered and unordered
                ordered_rows = []
                unordered_rows = []
                for row in module_tables[mod]:
                    if row['variable'] in kobo_positions:
                        ordered_rows.append((kobo_positions[row['variable']], row))
                    else:
                        unordered_rows.append(row)
                # Sort by Kobo position
                ordered_rows.sort(key=lambda x: x[0])
                # Combine: Kobo-ordered variables first, then unmatched variables
                reordered = [row for _, row in ordered_rows] + unordered_rows
                module_tables[mod] = reordered
                n_ordered = len(ordered_rows)
                n_total = len(module_tables[mod])
                reordered_count += n_ordered
                if n_ordered > 0:
                    log(f"  {mod}: {n_ordered}/{n_total} vars reordered by Kobo", 'OK')
        if reordered_count > 0:
            log(f"Total variables reordered by Kobo: {reordered_count}", 'OK')
    else:
        log("No Kobo variable ordering found — using questionnaire order", 'WARN')

    # ── Round-presence overrides ──────────────────────────────────────────────
    # Variables that exist in the data for a given round but were not an explicit
    # question row in the questionnaire workbook (e.g., backfilled from baseline,
    # derived in Kobo, or present under a different label).
    ROUND_OVERRIDES = {
        ('M00', 'FMID'): {'R1': True},   # FMID present in R1 data (Kobo: member_called / derived)
    }
    for (mod, var), rounds_to_add in ROUND_OVERRIDES.items():
        if mod not in module_tables:
            continue
        row = next((r for r in module_tables[mod] if r['variable'].upper() == var.upper()), None)
        if row:
            for rnd, present in rounds_to_add.items():
                if present:
                    row[f'in_{rnd}'] = '✓'
            # Recalculate first_round and status
            present_rounds = [r for r in ROUNDS if row.get(f'in_{r}') == '✓']
            if present_rounds:
                row['first_round'] = present_rounds[0]
                if present_rounds == ROUNDS:
                    row['status'] = 'All rounds'
                elif present_rounds[0] != ROUNDS[0]:
                    row['status'] = f'New in {present_rounds[0]}'
                elif present_rounds[-1] != ROUNDS[-1]:
                    row['status'] = f'Dropped after {present_rounds[-1]}'
                else:
                    row['status'] = 'All rounds'
            log(f"  Override: {mod}.{var} now present in {list(rounds_to_add.keys())}", 'OK')

    # ── M01 sub-variable expansion ───────────────────────────────────────────
    # The questionnaire has aggregate questions (D25, D26, M10, M13) but the
    # pooled data splits them into sub-variables (_oth, _1/_2/_3). Expand so
    # the tracker matches the DQ panels.
    if 'M01' in module_tables:
        def _clone_row(parent_row, new_var, new_title, new_type=None):
            """Clone a parent tracker row with a new variable name and title."""
            r = dict(parent_row)
            r['variable'] = new_var
            r['question_title'] = new_title
            if new_type:
                r['question_type'] = new_type
            return r

        M01_EXPANSIONS = {
            # parent_var: [(new_var, title, type_override), ...]
            'D25': [('D25_OTH', 'OTHER REASON FOR LEAVING', 'Open-end')],
            'D26': [
                ('D26_1', 'COUNTRY WHERE MOVED', None),
                ('D26_2', 'PROVINCE WHERE MOVED', None),
                ('D26_3', 'CITY WHERE MOVED', None),
            ],
            'M13': [('M13_OTH', 'OTHER REASON FOR MOVING IN', 'Open-end')],
            'M10': [
                ('M10_1', 'COUNTRY WHERE CAME FROM', None),
                ('M10_2', 'PROVINCE WHERE CAME FROM', None),
                ('M10_3', 'CITY WHERE CAME FROM', None),
            ],
        }

        new_m01 = []
        for row in module_tables['M01']:
            v = row['variable'].upper()
            if v in ('D26', 'M10'):
                # Replace parent with sub-variables
                for new_var, title, typ in M01_EXPANSIONS[v]:
                    new_m01.append(_clone_row(row, new_var, title, typ))
            else:
                new_m01.append(row)
                if v in M01_EXPANSIONS:
                    for new_var, title, typ in M01_EXPANSIONS[v]:
                        new_m01.append(_clone_row(row, new_var, title, typ))
        module_tables['M01'] = new_m01
        log(f"  M01: expanded D25/D26/M10/M13 into sub-variables ({len(new_m01)} rows)", 'OK')

    with open(CACHE / 'module_tables.json', 'w') as f:
        json.dump(module_tables, f, indent=2)
    log(f"cache/module_tables.json written ({(CACHE/'module_tables.json').stat().st_size//1024} KB)", 'OK')
    return True


# ── STEP 3b — Parse Kobo XLSForm skip logic ──────────────────────────────────
def rebuild_kobo_skip_logic():
    """Parse all Kobo XLSForm files and extract skip logic, mandatory, and constraint rules."""
    kobo_dir = CATI / 'KOBO'
    if not kobo_dir.exists():
        log("No CATI/KOBO/ directory found — skipping Kobo skip logic", 'WARN')
        return False
    kobo_script = Path(__file__).parent / 'scripts' / 'parse_kobo.py'
    if not kobo_script.exists():
        log("scripts/parse_kobo.py not found — skipping Kobo skip logic", 'WARN')
        return False
    import subprocess
    result = subprocess.run(
        ['python3', str(kobo_script)],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        log(f"parse_kobo.py failed: {result.stderr[:300]}", 'WARN')
        return False
    # parse_kobo.py writes directly to CACHE/kobo_skip_logic.json
    dst = CACHE / 'kobo_skip_logic.json'
    if not dst.exists():
        log("kobo_skip_logic.json not generated", 'WARN')
        return False
    log(f"cache/kobo_skip_logic.json written ({dst.stat().st_size//1024} KB)", 'OK')
    return True


# ── STEP 4 — Re-parse do-files ────────────────────────────────────────────────
def rebuild_dofiles(do_files):
    step(4, "Parsing Stata do-files")
    if not do_files:
        log("No do-files found in CATI/Round*/do/ — skipping", 'WARN')
        return False

    ROUND_RE = re.compile(r'@R0?(\d)', re.I)
    SECTION_PATTERNS = [
        (r'\(M1\)|ROSTER|M01', 'M01'), (r'\(M2\)|EDUCATION|M02', 'M02'),
        (r'\(M3\)|SHOCKS?|M03', 'M03'), (r'\(M4\)|EMPLOYMENT|M04', 'M04'),
        (r'\(M5\)|INCOME|M05', 'M05'), (r'\(M6\)|FINANCE|M06', 'M06'),
        (r'\(M7\)|HEALTH|M07', 'M07'), (r'\(M8\)|FOOD|M08', 'M08'),
        (r'\(M9\)|OPINION|VIEW|M09', 'M09'),
    ]

    def parse_do(path):
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        result = {}
        current_mod = None
        for line in lines:
            raw = line.strip()
            if raw.startswith('*') or raw.startswith('//'):
                text = raw.lstrip('*').lstrip('/').strip().upper()
                for pattern, mod in SECTION_PATTERNS:
                    if re.search(pattern, text, re.I):
                        if any(kw in text for kw in ['(M', 'MODULE', 'ROUNDS']):
                            current_mod = mod
                            if mod not in result:
                                result[mod] = {'vars': set(), 'destring': [], 'tostring': [], 'generate': []}
                            break
                continue
            if current_mod is None:
                continue
            code = re.sub(r'\s*//.*$', '', raw).strip()
            if not code:
                continue
            m = re.match(r'(?:cap(?:ture)?\s+)?destring\s+([\w\s*?]+?)(?:,|$)', code, re.I)
            if m:
                for v in m.group(1).strip().split():
                    if re.match(r'^[a-zA-Z_]', v) and v.lower() not in ('replace', 'force', 'ignore', 'float', 'double'):
                        result[current_mod]['vars'].add(v.lower())
                        result[current_mod]['destring'].append(v.lower())
            m = re.match(r'tostring\s+([\w\s]+?)(?:,|$)', code, re.I)
            if m:
                for v in m.group(1).strip().split():
                    if re.match(r'^[a-zA-Z_]', v):
                        result[current_mod]['vars'].add(v.lower())
                        result[current_mod]['tostring'].append(v.lower())
            m = re.match(r'gen(?:erate)?\s+(\w+)\s*=', code, re.I)
            if m:
                var = m.group(1).lower()
                result[current_mod]['vars'].add(var)
                result[current_mod]['generate'].append({'var': var, 'expr': code[m.end():].strip()[:80]})
        for mod in result:
            result[mod]['vars'] = sorted(result[mod]['vars'])
        return result

    # Map by round number — prefer the @AP master, else the most recent firm
    # delivery (rank = (is_AP, date)). R1–R7 have an @AP file so it always wins;
    # R08 ships only firm files (L2PH_CATI@R08@CB@{0611,0619,0626}) so the latest
    # date (20260626, matching the pooled master's R8) is chosen.
    DATE_RE = re.compile(r'@(\d{8})\.do$', re.I)
    def _rank(fp):
        fl = fp.name.lower()
        dm = DATE_RE.search(fp.name)
        return (1 if '@ap@' in fl else 0, int(dm.group(1)) if dm else 0)
    do_map = {}
    for fpath in do_files:
        m = ROUND_RE.search(fpath.name)
        if m:
            rnd = f"R{int(m.group(1))}"
            if rnd not in do_map or _rank(fpath) > _rank(do_map[rnd]):
                do_map[rnd] = fpath

    do_modules = {}
    for rnd, fpath in sorted(do_map.items()):
        mods = parse_do(fpath)
        do_modules[rnd] = mods
        n = sum(len(d['vars']) for d in mods.values())
        log(f"  {rnd} ({fpath.name}): {n} vars across {list(mods.keys())}")

    with open(CACHE / 'do_modules.json', 'w') as f:
        json.dump(do_modules, f, indent=2)
    log(f"cache/do_modules.json written ({(CACHE/'do_modules.json').stat().st_size//1024} KB)", 'OK')
    return True


# ── STEP 4b — Build issue intelligence ───────────────────────────────────────
def rebuild_issues():
    step("4b", "Building issue intelligence (issues.json)")
    r = subprocess.run([sys.executable, str(SCRIPTS / 'build_issues.py')],
                       capture_output=True, text=True)
    if r.returncode != 0:
        log(f"build_issues failed:\n{r.stderr[-800:]}", 'ERROR'); return False
    log(r.stdout.strip() or "issues.json written", 'OK'); return True


# ── STEP 5 — Regenerate HTML dashboard ───────────────────────────────────────
def rebuild_dashboard():
    step(5, "Regenerating HTML dashboard")
    ok = run_script('gen_dashboard.py')
    if ok:
        size = (OUTPUT / 'l2ph_dq_dashboard.html').stat().st_size // 1024
        log(f"output/l2ph_dq_dashboard.html ({size} KB)", 'OK')
    return ok


# ── STEP 6 — Regenerate Excel report ─────────────────────────────────────────
def rebuild_excel():
    step(6, "Regenerating Excel cross-round report")
    ok = run_script('build_report.py')
    if ok:
        size = (OUTPUT / 'L2PHL_Questionnaire_Cross_Round_Report.xlsx').stat().st_size // 1024
        log(f"output/L2PHL_Questionnaire_Cross_Round_Report.xlsx ({size} KB)", 'OK')
    return ok


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    args = sys.argv[1:] or ['--all']
    do_dta = '--all' in args or '--dta' in args
    do_q   = '--all' in args or '--questionnaire' in args
    do_do  = '--all' in args or '--dofiles' in args

    print()
    print("╔══════════════════════════════════════════════╗")
    print("║  L2PHL CATI Data Quality Pipeline           ║")
    print("╚══════════════════════════════════════════════╝")
    print(f"  Root  : {CATI}")
    print(f"  HF    : {HF_DIR}")
    print(f"  QC    : {HERE}")

    dta_files, quest_files, do_files, round_dta = detect_files()

    errors = []
    t0 = time.time()

    if do_dta:
        if not rebuild_dq(dta_files):          errors.append('DQ rebuild failed')
        if not rebuild_panel(dta_files):       errors.append('Panel rebuild failed')
        if not rebuild_interviewer(dta_files): errors.append('Interviewer rebuild failed')

    if do_q:
        if not rebuild_questionnaire(quest_files): errors.append('Questionnaire parse failed')
        rebuild_kobo_skip_logic()  # Parse Kobo XLSForms for skip logic map (non-blocking)

    if do_do:
        if not rebuild_dofiles(do_files): errors.append('Do-file parse failed')

    if not rebuild_issues():     errors.append('Issue intelligence build failed')
    if not rebuild_dashboard(): errors.append('Dashboard generation failed')
    if not rebuild_excel():     errors.append('Excel report generation failed')

    elapsed = time.time() - t0
    print()
    print("╔══════════════════════════════════════════════╗")
    if errors:
        print(f"║  ⚠️  Finished with {len(errors)} error(s) in {elapsed:.1f}s")
        for e in errors: print(f"║  ❌ {e}")
    else:
        print(f"║  ✅ All outputs updated in {elapsed:.1f}s")
        print("║")
        print("║  📊 output/l2ph_dq_dashboard.html")
        print("║  📋 output/L2PHL_Questionnaire_Cross_Round_Report.xlsx")
    print("╚══════════════════════════════════════════════╝")
    print()
