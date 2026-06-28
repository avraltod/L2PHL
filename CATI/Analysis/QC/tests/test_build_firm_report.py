import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
from build_firm_report import build, HEADERS
from openpyxl import load_workbook

def rec(**kw):
    base = {"report_to_firm": True, "status": "acknowledged", "verdict": "A2",
            "module":"M04","variable":"a18","label":"A1=2 but A18 filled","owner":"firm-field",
            "counts_by_round":{"6":12},"evidence":{"kobo":{},"dofile":{"ever_touched":False}},
            "notes":"fix it"}
    base.update(kw); return base

def test_workbook_structure(tmp_path):
    wb = build([rec(), rec(report_to_firm=False)], "20260628")
    p = tmp_path/"t.xlsx"; wb.save(p)
    ws = load_workbook(p).active
    headers = [ws.cell(4, c).value for c in range(1, len(HEADERS)+1)]
    assert headers == HEADERS                              # header row at row 4
    assert ws.cell(5, 2).value == "M04" and ws.cell(5, 3).value == "a18"  # 1 data row
    assert ws.cell(6, 2).value is None                    # only the open firm one
    assert ws.cell(4, 13).value == "Firm response / fixed?"
    assert ws.cell(5, 13).value in (None, "")             # response col empty
    assert "Firm Data Quality Tracker" in ws.cell(1, 1).value
    assert "1 open firm issue" in ws.cell(2, 1).value
